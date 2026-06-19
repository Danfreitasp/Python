import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import base64
import csv
import uuid
import secrets
import string
from datetime import datetime

from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

from openpyxl import Workbook, load_workbook


# ==========================
# Configurações gerais
# ==========================

APP_TITLE = "Gerenciador de Senhas Local"
VAULT_FILE = "senhas_vault.enc"

PBKDF2_ITERATIONS = 390_000

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


# ==========================
# Funções de criptografia
# ==========================

def gerar_salt():
    return os.urandom(16)


def derivar_chave(senha_mestra: str, salt: bytes) -> bytes:
    """
    Gera uma chave Fernet a partir da senha mestra.
    A senha mestra não é salva em lugar nenhum.
    """
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=PBKDF2_ITERATIONS,
    )

    chave = base64.urlsafe_b64encode(kdf.derive(senha_mestra.encode("utf-8")))
    return chave


def criptografar_dados(dados: list, senha_mestra: str, salt: bytes) -> str:
    chave = derivar_chave(senha_mestra, salt)
    fernet = Fernet(chave)

    conteudo_json = json.dumps(dados, ensure_ascii=False, indent=2)
    token = fernet.encrypt(conteudo_json.encode("utf-8"))

    pacote = {
        "salt": base64.b64encode(salt).decode("utf-8"),
        "token": token.decode("utf-8")
    }

    return json.dumps(pacote, ensure_ascii=False, indent=2)


def descriptografar_arquivo(caminho: str, senha_mestra: str):
    with open(caminho, "r", encoding="utf-8") as arquivo:
        pacote = json.load(arquivo)

    salt = base64.b64decode(pacote["salt"])
    token = pacote["token"].encode("utf-8")

    chave = derivar_chave(senha_mestra, salt)
    fernet = Fernet(chave)

    dados_json = fernet.decrypt(token).decode("utf-8")
    dados = json.loads(dados_json)

    return dados, salt


# ==========================
# Janela de senha mestra
# ==========================

class MasterPasswordDialog(ctk.CTkToplevel):
    def __init__(self, parent, modo_criacao=False):
        super().__init__(parent)

        self.parent = parent
        self.modo_criacao = modo_criacao
        self.resultado = None

        self.title("Senha Mestra")
        self.geometry("420x260" if modo_criacao else "420x200")
        self.resizable(False, False)

        self.transient(parent)
        self.grab_set()

        titulo = "Criar senha mestra" if modo_criacao else "Digite sua senha mestra"

        ctk.CTkLabel(
            self,
            text=titulo,
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=(25, 10))

        texto_info = (
            "Essa senha protegerá seu cofre local.\n"
            "Se você esquecê-la, não será possível recuperar os dados."
            if modo_criacao
            else "Informe a senha mestra para abrir seu cofre."
        )

        ctk.CTkLabel(self, text=texto_info, justify="center").pack(pady=(0, 15))

        self.entry_senha = ctk.CTkEntry(self, placeholder_text="Senha mestra", show="*")
        self.entry_senha.pack(padx=30, fill="x", pady=5)
        self.entry_senha.focus()

        if modo_criacao:
            self.entry_confirmar = ctk.CTkEntry(self, placeholder_text="Confirmar senha mestra", show="*")
            self.entry_confirmar.pack(padx=30, fill="x", pady=5)
        else:
            self.entry_confirmar = None

        botoes_frame = ctk.CTkFrame(self, fg_color="transparent")
        botoes_frame.pack(pady=20)

        ctk.CTkButton(
            botoes_frame,
            text="Confirmar",
            command=self.confirmar
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            botoes_frame,
            text="Cancelar",
            fg_color="gray",
            command=self.cancelar
        ).pack(side="left", padx=8)

        self.bind("<Return>", lambda event: self.confirmar())
        self.protocol("WM_DELETE_WINDOW", self.cancelar)

    def confirmar(self):
        senha = self.entry_senha.get().strip()

        if not senha:
            messagebox.showwarning("Atenção", "Digite a senha mestra.")
            return

        if self.modo_criacao:
            confirmar = self.entry_confirmar.get().strip()

            if senha != confirmar:
                messagebox.showerror("Erro", "As senhas não conferem.")
                return

            if len(senha) < 6:
                messagebox.showwarning(
                    "Senha fraca",
                    "Use uma senha mestra com pelo menos 6 caracteres.\n"
                    "O ideal é usar uma frase longa e difícil de adivinhar."
                )
                return

        self.resultado = senha
        self.destroy()

    def cancelar(self):
        self.resultado = None
        self.destroy()


# ==========================
# Aplicativo principal
# ==========================

class GerenciadorSenhas(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title(APP_TITLE)
        self.geometry("1200x700")
        self.minsize(1050, 620)

        self.senha_mestra = None
        self.salt = None
        self.registros = []
        self.registro_selecionado_id = None

        if not self.iniciar_cofre():
            self.destroy()
            return

        self.criar_interface()
        self.atualizar_tabela()

    # ==========================
    # Inicialização do cofre
    # ==========================

    def iniciar_cofre(self):
        if os.path.exists(VAULT_FILE):
            return self.abrir_cofre_existente()
        else:
            return self.criar_novo_cofre()

    def abrir_cofre_existente(self):
        dialog = MasterPasswordDialog(self, modo_criacao=False)
        self.wait_window(dialog)

        if not dialog.resultado:
            return False

        try:
            dados, salt = descriptografar_arquivo(VAULT_FILE, dialog.resultado)
            self.senha_mestra = dialog.resultado
            self.salt = salt
            self.registros = dados
            return True

        except InvalidToken:
            messagebox.showerror(
                "Erro",
                "Senha mestra incorreta ou arquivo criptografado inválido."
            )
            return False

        except Exception as erro:
            messagebox.showerror(
                "Erro",
                f"Não foi possível abrir o cofre.\n\nDetalhes: {erro}"
            )
            return False

    def criar_novo_cofre(self):
        dialog = MasterPasswordDialog(self, modo_criacao=True)
        self.wait_window(dialog)

        if not dialog.resultado:
            return False

        self.senha_mestra = dialog.resultado
        self.salt = gerar_salt()
        self.registros = []
        self.salvar_cofre()

        messagebox.showinfo(
            "Cofre criado",
            "Seu cofre local foi criado com sucesso."
        )

        return True

    def salvar_cofre(self):
        try:
            conteudo = criptografar_dados(
                self.registros,
                self.senha_mestra,
                self.salt
            )

            with open(VAULT_FILE, "w", encoding="utf-8") as arquivo:
                arquivo.write(conteudo)

        except Exception as erro:
            messagebox.showerror(
                "Erro ao salvar",
                f"Não foi possível salvar o cofre.\n\nDetalhes: {erro}"
            )

    # ==========================
    # Interface
    # ==========================

    def criar_interface(self):
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

        titulo = ctk.CTkLabel(
            self,
            text="Gerenciador de Senhas Local",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        titulo.grid(row=0, column=0, columnspan=2, pady=15)

        self.criar_formulario()
        self.criar_area_tabela()

    def criar_formulario(self):
        frame = ctk.CTkFrame(self)
        frame.grid(row=1, column=0, sticky="nsw", padx=(15, 8), pady=(0, 15))
        frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            frame,
            text="Dados da conta",
            font=ctk.CTkFont(size=18, weight="bold")
        ).grid(row=0, column=0, padx=15, pady=(15, 10), sticky="w")

        self.entry_site = ctk.CTkEntry(frame, placeholder_text="Site/Sistema *", width=300)
        self.entry_site.grid(row=1, column=0, padx=15, pady=6, sticky="ew")

        self.entry_login = ctk.CTkEntry(frame, placeholder_text="Login/E-mail *")
        self.entry_login.grid(row=2, column=0, padx=15, pady=6, sticky="ew")

        senha_frame = ctk.CTkFrame(frame, fg_color="transparent")
        senha_frame.grid(row=3, column=0, padx=15, pady=6, sticky="ew")
        senha_frame.grid_columnconfigure(0, weight=1)

        self.entry_senha = ctk.CTkEntry(senha_frame, placeholder_text="Senha *", show="*")
        self.entry_senha.grid(row=0, column=0, sticky="ew")

        self.btn_ver_senha = ctk.CTkButton(
            senha_frame,
            text="Ver",
            width=55,
            command=self.alternar_visibilidade_senha
        )
        self.btn_ver_senha.grid(row=0, column=1, padx=(6, 0))

        self.entry_categoria = ctk.CTkEntry(frame, placeholder_text="Categoria")
        self.entry_categoria.grid(row=4, column=0, padx=15, pady=6, sticky="ew")

        self.entry_url = ctk.CTkEntry(frame, placeholder_text="URL")
        self.entry_url.grid(row=5, column=0, padx=15, pady=6, sticky="ew")

        self.text_observacoes = ctk.CTkTextbox(frame, height=100)
        self.text_observacoes.grid(row=6, column=0, padx=15, pady=6, sticky="ew")
        self.text_observacoes.insert("1.0", "")

        ctk.CTkButton(
            frame,
            text="Gerar senha forte",
            command=self.gerar_senha_forte
        ).grid(row=7, column=0, padx=15, pady=(10, 6), sticky="ew")

        ctk.CTkButton(
            frame,
            text="Salvar / Atualizar",
            command=self.salvar_registro
        ).grid(row=8, column=0, padx=15, pady=6, sticky="ew")

        ctk.CTkButton(
            frame,
            text="Limpar campos",
            fg_color="gray",
            command=self.limpar_formulario
        ).grid(row=9, column=0, padx=15, pady=6, sticky="ew")

        ctk.CTkButton(
            frame,
            text="Excluir selecionado",
            fg_color="#a83232",
            hover_color="#7d2525",
            command=self.excluir_registro
        ).grid(row=10, column=0, padx=15, pady=(6, 15), sticky="ew")

    def criar_area_tabela(self):
        frame = ctk.CTkFrame(self)
        frame.grid(row=1, column=1, sticky="nsew", padx=(8, 15), pady=(0, 15))
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(2, weight=1)

        busca_frame = ctk.CTkFrame(frame, fg_color="transparent")
        busca_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=(15, 8))
        busca_frame.grid_columnconfigure(0, weight=1)

        self.entry_busca = ctk.CTkEntry(
            busca_frame,
            placeholder_text="Buscar por site, login ou categoria..."
        )
        self.entry_busca.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self.entry_busca.bind("<KeyRelease>", lambda event: self.atualizar_tabela())

        ctk.CTkButton(
            busca_frame,
            text="Limpar busca",
            width=110,
            command=self.limpar_busca
        ).grid(row=0, column=1)

        acoes_frame = ctk.CTkFrame(frame, fg_color="transparent")
        acoes_frame.grid(row=1, column=0, sticky="ew", padx=15, pady=(0, 8))

        ctk.CTkButton(
            acoes_frame,
            text="Copiar login",
            command=self.copiar_login
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            acoes_frame,
            text="Copiar senha",
            command=self.copiar_senha
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            acoes_frame,
            text="Importar CSV/Excel",
            command=self.importar_dados
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            acoes_frame,
            text="Exportar CSV/Excel",
            command=self.exportar_dados
        ).pack(side="left", padx=(0, 8))

        tabela_frame = ctk.CTkFrame(frame)
        tabela_frame.grid(row=2, column=0, sticky="nsew", padx=15, pady=(0, 15))
        tabela_frame.grid_columnconfigure(0, weight=1)
        tabela_frame.grid_rowconfigure(0, weight=1)

        colunas = (
            "site",
            "login",
            "senha",
            "categoria",
            "url",
            "criado",
            "alterado"
        )

        self.tabela = ttk.Treeview(
            tabela_frame,
            columns=colunas,
            show="headings",
            selectmode="browse"
        )

        self.tabela.heading("site", text="Site/Sistema")
        self.tabela.heading("login", text="Login/E-mail")
        self.tabela.heading("senha", text="Senha")
        self.tabela.heading("categoria", text="Categoria")
        self.tabela.heading("url", text="URL")
        self.tabela.heading("criado", text="Criado em")
        self.tabela.heading("alterado", text="Alterado em")

        self.tabela.column("site", width=160)
        self.tabela.column("login", width=180)
        self.tabela.column("senha", width=90, anchor="center")
        self.tabela.column("categoria", width=120)
        self.tabela.column("url", width=180)
        self.tabela.column("criado", width=130)
        self.tabela.column("alterado", width=130)

        scrollbar_y = ttk.Scrollbar(tabela_frame, orient="vertical", command=self.tabela.yview)
        scrollbar_x = ttk.Scrollbar(tabela_frame, orient="horizontal", command=self.tabela.xview)

        self.tabela.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        self.tabela.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        self.tabela.bind("<<TreeviewSelect>>", self.ao_selecionar_registro)
        self.tabela.bind("<Double-1>", self.ao_selecionar_registro)

    # ==========================
    # Utilidades
    # ==========================

    def data_atual(self):
        return datetime.now().strftime("%d/%m/%Y %H:%M")

    def encontrar_registro_por_id(self, registro_id):
        for registro in self.registros:
            if registro.get("id") == registro_id:
                return registro
        return None

    def obter_registro_selecionado(self):
        selecionado = self.tabela.selection()

        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um registro na tabela.")
            return None

        registro_id = selecionado[0]
        registro = self.encontrar_registro_por_id(registro_id)

        if not registro:
            messagebox.showerror("Erro", "Registro não encontrado.")
            return None

        return registro

    def copiar_para_area_transferencia(self, texto, mensagem):
        self.clipboard_clear()
        self.clipboard_append(texto)
        self.update()

        messagebox.showinfo("Copiado", mensagem)

    # ==========================
    # Formulário
    # ==========================

    def alternar_visibilidade_senha(self):
        if self.entry_senha.cget("show") == "*":
            self.entry_senha.configure(show="")
            self.btn_ver_senha.configure(text="Ocultar")
        else:
            self.entry_senha.configure(show="*")
            self.btn_ver_senha.configure(text="Ver")

    def limpar_formulario(self):
        self.registro_selecionado_id = None

        self.entry_site.delete(0, "end")
        self.entry_login.delete(0, "end")
        self.entry_senha.delete(0, "end")
        self.entry_categoria.delete(0, "end")
        self.entry_url.delete(0, "end")
        self.text_observacoes.delete("1.0", "end")

        self.tabela.selection_remove(self.tabela.selection())

    def salvar_registro(self):
        site = self.entry_site.get().strip()
        login = self.entry_login.get().strip()
        senha = self.entry_senha.get().strip()
        categoria = self.entry_categoria.get().strip()
        url = self.entry_url.get().strip()
        observacoes = self.text_observacoes.get("1.0", "end").strip()

        if not site or not login or not senha:
            messagebox.showwarning(
                "Campos obrigatórios",
                "Preencha Site/Sistema, Login/E-mail e Senha."
            )
            return

        agora = self.data_atual()

        if self.registro_selecionado_id:
            registro = self.encontrar_registro_por_id(self.registro_selecionado_id)

            if not registro:
                messagebox.showerror("Erro", "Registro selecionado não encontrado.")
                return

            registro["site"] = site
            registro["login"] = login
            registro["senha"] = senha
            registro["categoria"] = categoria
            registro["url"] = url
            registro["observacoes"] = observacoes
            registro["alterado_em"] = agora

            mensagem = "Registro atualizado com sucesso."

        else:
            novo_registro = {
                "id": str(uuid.uuid4()),
                "site": site,
                "login": login,
                "senha": senha,
                "categoria": categoria,
                "url": url,
                "observacoes": observacoes,
                "criado_em": agora,
                "alterado_em": agora
            }

            self.registros.append(novo_registro)
            mensagem = "Registro cadastrado com sucesso."

        self.salvar_cofre()
        self.atualizar_tabela()
        self.limpar_formulario()

        messagebox.showinfo("Sucesso", mensagem)

    def excluir_registro(self):
        registro = self.obter_registro_selecionado()

        if not registro:
            return

        confirmar = messagebox.askyesno(
            "Confirmar exclusão",
            f"Deseja excluir o registro de '{registro.get('site', '')}'?"
        )

        if not confirmar:
            return

        self.registros = [
            item for item in self.registros
            if item.get("id") != registro.get("id")
        ]

        self.salvar_cofre()
        self.atualizar_tabela()
        self.limpar_formulario()

        messagebox.showinfo("Excluído", "Registro excluído com sucesso.")

    def ao_selecionar_registro(self, event=None):
        selecionado = self.tabela.selection()

        if not selecionado:
            return

        registro_id = selecionado[0]
        registro = self.encontrar_registro_por_id(registro_id)

        if not registro:
            return

        self.registro_selecionado_id = registro_id

        self.entry_site.delete(0, "end")
        self.entry_login.delete(0, "end")
        self.entry_senha.delete(0, "end")
        self.entry_categoria.delete(0, "end")
        self.entry_url.delete(0, "end")
        self.text_observacoes.delete("1.0", "end")

        self.entry_site.insert(0, registro.get("site", ""))
        self.entry_login.insert(0, registro.get("login", ""))
        self.entry_senha.insert(0, registro.get("senha", ""))
        self.entry_categoria.insert(0, registro.get("categoria", ""))
        self.entry_url.insert(0, registro.get("url", ""))
        self.text_observacoes.insert("1.0", registro.get("observacoes", ""))

    # ==========================
    # Tabela e busca
    # ==========================

    def atualizar_tabela(self):
        for item in self.tabela.get_children():
            self.tabela.delete(item)

        termo = self.entry_busca.get().strip().lower() if hasattr(self, "entry_busca") else ""

        registros_ordenados = sorted(
            self.registros,
            key=lambda x: x.get("site", "").lower()
        )

        for registro in registros_ordenados:
            site = registro.get("site", "")
            login = registro.get("login", "")
            categoria = registro.get("categoria", "")

            texto_busca = f"{site} {login} {categoria}".lower()

            if termo and termo not in texto_busca:
                continue

            self.tabela.insert(
                "",
                "end",
                iid=registro.get("id"),
                values=(
                    site,
                    login,
                    "********",
                    categoria,
                    registro.get("url", ""),
                    registro.get("criado_em", ""),
                    registro.get("alterado_em", "")
                )
            )

    def limpar_busca(self):
        self.entry_busca.delete(0, "end")
        self.atualizar_tabela()

    # ==========================
    # Copiar dados
    # ==========================

    def copiar_login(self):
        registro = self.obter_registro_selecionado()

        if not registro:
            return

        self.copiar_para_area_transferencia(
            registro.get("login", ""),
            "Login copiado para a área de transferência."
        )

    def copiar_senha(self):
        registro = self.obter_registro_selecionado()

        if not registro:
            return

        self.copiar_para_area_transferencia(
            registro.get("senha", ""),
            "Senha copiada para a área de transferência."
        )

    # ==========================
    # Gerador de senha
    # ==========================

    def gerar_senha_forte(self):
        tamanho = 18

        letras_minusculas = string.ascii_lowercase
        letras_maiusculas = string.ascii_uppercase
        numeros = string.digits
        simbolos = "!@#$%&*()-_=+[]{};:,.?/"

        senha = [
            secrets.choice(letras_minusculas),
            secrets.choice(letras_maiusculas),
            secrets.choice(numeros),
            secrets.choice(simbolos)
        ]

        todos = letras_minusculas + letras_maiusculas + numeros + simbolos

        while len(senha) < tamanho:
            senha.append(secrets.choice(todos))

        secrets.SystemRandom().shuffle(senha)
        senha_final = "".join(senha)

        self.entry_senha.delete(0, "end")
        self.entry_senha.insert(0, senha_final)

        messagebox.showinfo("Senha gerada", "Senha forte gerada no campo de senha.")

    # ==========================
    # Importação
    # ==========================

    def importar_dados(self):
        caminho = filedialog.askopenfilename(
            title="Importar dados",
            filetypes=[
                ("Arquivos CSV ou Excel", "*.csv *.xlsx"),
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx")
            ]
        )

        if not caminho:
            return

        try:
            if caminho.lower().endswith(".csv"):
                novos = self.ler_csv(caminho)
            elif caminho.lower().endswith(".xlsx"):
                novos = self.ler_excel(caminho)
            else:
                messagebox.showwarning("Formato inválido", "Selecione um arquivo CSV ou Excel.")
                return

            if not novos:
                messagebox.showwarning("Atenção", "Nenhum registro válido foi encontrado.")
                return

            confirmar = messagebox.askyesno(
                "Confirmar importação",
                f"Foram encontrados {len(novos)} registros válidos.\n\nDeseja importar?"
            )

            if not confirmar:
                return

            self.registros.extend(novos)
            self.salvar_cofre()
            self.atualizar_tabela()

            messagebox.showinfo("Importação concluída", f"{len(novos)} registros importados.")

        except Exception as erro:
            messagebox.showerror(
                "Erro ao importar",
                f"Não foi possível importar o arquivo.\n\nDetalhes: {erro}"
            )

    def normalizar_cabecalho(self, texto):
        return str(texto).strip().lower().replace("/", "").replace("-", "").replace("_", "").replace(" ", "")

    def mapear_linha(self, linha):
        """
        Aceita colunas:
        Site, Login, Senha, Categoria, URL, Observações
        """
        def pegar(*nomes):
            for nome in nomes:
                if nome in linha:
                    return str(linha.get(nome, "") or "").strip()
            return ""

        site = pegar("site", "sitesistema", "sistema")
        login = pegar("login", "email", "loginemail", "usuario", "usuário")
        senha = pegar("senha", "password")
        categoria = pegar("categoria")
        url = pegar("url", "link")
        observacoes = pegar("observacoes", "observações", "obs")

        if not site or not login or not senha:
            return None

        agora = self.data_atual()

        return {
            "id": str(uuid.uuid4()),
            "site": site,
            "login": login,
            "senha": senha,
            "categoria": categoria,
            "url": url,
            "observacoes": observacoes,
            "criado_em": agora,
            "alterado_em": agora
        }

    def ler_csv(self, caminho):
        registros = []

        with open(caminho, "r", encoding="utf-8-sig", newline="") as arquivo:
            leitor = csv.DictReader(arquivo)

            for linha_original in leitor:
                linha = {
                    self.normalizar_cabecalho(chave): valor
                    for chave, valor in linha_original.items()
                }

                registro = self.mapear_linha(linha)

                if registro:
                    registros.append(registro)

        return registros

    def ler_excel(self, caminho):
        registros = []

        workbook = load_workbook(caminho)
        sheet = workbook.active

        cabecalhos = []

        for celula in sheet[1]:
            cabecalhos.append(self.normalizar_cabecalho(celula.value))

        for row in sheet.iter_rows(min_row=2, values_only=True):
            linha = {}

            for indice, valor in enumerate(row):
                if indice < len(cabecalhos):
                    linha[cabecalhos[indice]] = valor

            registro = self.mapear_linha(linha)

            if registro:
                registros.append(registro)

        return registros

    # ==========================
    # Exportação
    # ==========================

    def exportar_dados(self):
        if not self.registros:
            messagebox.showwarning("Atenção", "Não há registros para exportar.")
            return

        confirmar = messagebox.askyesno(
            "Atenção",
            "O arquivo exportado NÃO será criptografado.\n"
            "Qualquer pessoa que abrir o CSV/Excel poderá ver as senhas.\n\n"
            "Deseja continuar?"
        )

        if not confirmar:
            return

        caminho = filedialog.asksaveasfilename(
            title="Exportar dados",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("CSV", "*.csv")
            ]
        )

        if not caminho:
            return

        try:
            if caminho.lower().endswith(".csv"):
                self.exportar_csv(caminho)
            elif caminho.lower().endswith(".xlsx"):
                self.exportar_excel(caminho)
            else:
                caminho += ".xlsx"
                self.exportar_excel(caminho)

            messagebox.showinfo(
                "Exportação concluída",
                "Dados exportados com sucesso.\n\n"
                "Lembre-se: esse arquivo exportado não está criptografado."
            )

        except Exception as erro:
            messagebox.showerror(
                "Erro ao exportar",
                f"Não foi possível exportar os dados.\n\nDetalhes: {erro}"
            )

    def exportar_csv(self, caminho):
        campos = [
            "Site",
            "Login",
            "Senha",
            "Categoria",
            "URL",
            "Observações",
            "Data de criação",
            "Data da última alteração"
        ]

        with open(caminho, "w", encoding="utf-8-sig", newline="") as arquivo:
            escritor = csv.writer(arquivo)
            escritor.writerow(campos)

            for registro in self.registros:
                escritor.writerow([
                    registro.get("site", ""),
                    registro.get("login", ""),
                    registro.get("senha", ""),
                    registro.get("categoria", ""),
                    registro.get("url", ""),
                    registro.get("observacoes", ""),
                    registro.get("criado_em", ""),
                    registro.get("alterado_em", "")
                ])

    def exportar_excel(self, caminho):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Senhas"

        cabecalhos = [
            "Site",
            "Login",
            "Senha",
            "Categoria",
            "URL",
            "Observações",
            "Data de criação",
            "Data da última alteração"
        ]

        sheet.append(cabecalhos)

        for registro in self.registros:
            sheet.append([
                registro.get("site", ""),
                registro.get("login", ""),
                registro.get("senha", ""),
                registro.get("categoria", ""),
                registro.get("url", ""),
                registro.get("observacoes", ""),
                registro.get("criado_em", ""),
                registro.get("alterado_em", "")
            ])

        workbook.save(caminho)


# ==========================
# Execução
# ==========================

if __name__ == "__main__":
    app = GerenciadorSenhas()
    app.mainloop()