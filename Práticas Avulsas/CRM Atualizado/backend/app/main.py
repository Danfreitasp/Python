from __future__ import annotations

import csv
import io
import json
import re
import shutil
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from .database import DATA_DIR, DATABASE_PATH, db, prepare_database, rows

app = FastAPI(title="CRM Consig Next", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["http://localhost:5173"], allow_methods=["*"], allow_headers=["*"])

CLOSED = {"Pago", "Perdido / Cancelado", "Perdido", "Cancelado"}
PROPOSAL_FIELDS = {
    "nome", "cpf", "nb_matricula", "tipo_cliente", "banco_atual", "banco_digitado", "produto", "promotora",
    "beneficio_bloqueado", "parcela_atual", "nova_parcela", "troco", "comissao", "comissao_percentual", "margem_apos",
    "status", "responsavel", "telefone", "proxima_acao", "data_retorno", "observacoes", "endereco", "dados_bancarios",
    "valor_caiu_promotora", "valor_sacado", "numero_proposta", "numero_port_vinculada", "numero_refin_vinculada",
}


class ProposalInput(BaseModel):
    values: dict[str, Any] = Field(default_factory=dict)


class StatusInput(BaseModel):
    status: str
    observacao: str = "Status atualizado pelo CRM Next"


class NoteInput(BaseModel):
    texto: str


class TaskInput(BaseModel):
    titulo: str
    descricao: str = ""
    data_tarefa: str = Field(default_factory=lambda: date.today().isoformat())
    horario: str = ""
    prioridade: str = "normal"
    status: str = "pendente"
    categoria: str = "Retorno"
    proposta_id: int | None = None


class TemplateInput(BaseModel):
    nome: str
    texto: str
    ordem: int = 0


class StageInput(BaseModel):
    nome: str
    grupo: str = "geral"
    ordem: int = 0
    ativo: bool = True


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def proposal_or_404(proposal_id: int) -> dict:
    with db() as conn:
        row = conn.execute("SELECT * FROM propostas WHERE id=?", (proposal_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Proposta não encontrada")
    return dict(row)


def history(proposal_id: int, previous: str | None, current: str | None, observation: str) -> None:
    with db() as conn:
        conn.execute("INSERT INTO historico (proposta_id,data_hora,status_anterior,status_novo,observacao) VALUES (?,?,?,?,?)",
                     (proposal_id, now(), previous, current, observation))


def save_client(conn: Any, values: dict[str, Any]) -> None:
    cpf, registration = str(values.get("cpf") or ""), str(values.get("nb_matricula") or "")
    if not cpf or not registration:
        return
    found = conn.execute("SELECT id FROM clientes WHERE cpf=? AND nb_matricula=?", (cpf, registration)).fetchone()
    fields = (values.get("nome", ""), values.get("telefone", ""), values.get("tipo_cliente", ""), values.get("endereco", ""), values.get("dados_bancarios", ""), now())
    if found:
        conn.execute("UPDATE clientes SET nome=?,telefone=?,tipo_cliente=?,endereco=?,dados_bancarios=?,data_atualizacao=? WHERE id=?", (*fields, found["id"]))
    else:
        conn.execute("INSERT INTO clientes (nome,cpf,nb_matricula,telefone,tipo_cliente,endereco,dados_bancarios,data_criacao,data_atualizacao) VALUES (?,?,?,?,?,?,?,?,?)",
                     (values.get("nome", ""), cpf, registration, values.get("telefone", ""), values.get("tipo_cliente", ""), values.get("endereco", ""), values.get("dados_bancarios", ""), now(), now()))


@app.on_event("startup")
def startup() -> None:
    prepare_database()


@app.get("/api/health")
def health() -> dict:
    return {"ok": True, "database": str(DATABASE_PATH), "source_isolated": True}


@app.get("/api/meta")
def meta() -> dict:
    with db() as conn:
        stages = rows(conn.execute("SELECT * FROM status_etapas ORDER BY grupo,ordem,nome"))
        notifications = rows(conn.execute("SELECT * FROM notificacoes_importantes ORDER BY data_hora DESC LIMIT 8"))
    return {"stages": stages, "closed_statuses": list(CLOSED), "notifications": notifications}


@app.get("/api/proposals")
def list_proposals(q: str = "", status: str = "", mes: str = "", limit: int = Query(300, le=1000), offset: int = 0, **filters: str) -> dict:
    clauses, parameters = [], []
    if q:
        clauses.append("(nome LIKE ? OR cpf LIKE ? OR telefone LIKE ? OR numero_proposta LIKE ?)")
        parameters.extend([f"%{q}%"] * 4)
    for field in ("status", "banco_atual", "banco_digitado", "tipo_cliente", "produto", "promotora", "beneficio_bloqueado", "valor_caiu_promotora", "valor_sacado"):
        value = status if field == "status" else filters.get(field, "")
        if value:
            clauses.append(f"{field}=?"); parameters.append(value)
    if mes:
        clauses.append("substr(data_criacao,1,7)=?"); parameters.append(mes)
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    with db() as conn:
        total = conn.execute("SELECT count(*) FROM propostas" + where, parameters).fetchone()[0]
        result = rows(conn.execute("SELECT * FROM propostas" + where + " ORDER BY data_atualizacao DESC LIMIT ? OFFSET ?", (*parameters, limit, offset)))
    return {"items": result, "total": total}


@app.post("/api/proposals", status_code=201)
def create_proposal(body: ProposalInput) -> dict:
    values = {key: body.values.get(key) for key in PROPOSAL_FIELDS}
    values["nome"] = str(values.get("nome") or "").strip()
    if not values["nome"]:
        raise HTTPException(422, "Nome é obrigatório")
    values["status"] = values.get("status") or "Aguardando inserção"
    values["data_criacao"] = values["data_atualizacao"] = now()
    columns = list(values)
    with db() as conn:
        cursor = conn.execute(f"INSERT INTO propostas ({','.join(columns)}) VALUES ({','.join('?' for _ in columns)})", [values[col] for col in columns])
        proposal_id = cursor.lastrowid
        save_client(conn, values)
    history(proposal_id, None, values["status"], "Proposta criada")
    return proposal_or_404(proposal_id)


@app.get("/api/proposals/{proposal_id}")
def proposal_detail(proposal_id: int) -> dict:
    proposal = proposal_or_404(proposal_id)
    with db() as conn:
        proposal["notes"] = rows(conn.execute("SELECT * FROM anotacoes WHERE proposta_id=? ORDER BY data_hora DESC", (proposal_id,)))
        proposal["attachments"] = rows(conn.execute("SELECT * FROM anexos WHERE proposta_id=? ORDER BY data_upload DESC", (proposal_id,)))
        proposal["history"] = rows(conn.execute("SELECT * FROM historico WHERE proposta_id=? ORDER BY data_hora DESC", (proposal_id,)))
        proposal["tasks"] = rows(conn.execute("SELECT * FROM tarefas WHERE proposta_id=? ORDER BY data_tarefa", (proposal_id,)))
    return proposal


@app.put("/api/proposals/{proposal_id}")
def update_proposal(proposal_id: int, body: ProposalInput) -> dict:
    old = proposal_or_404(proposal_id)
    values = {key: value for key, value in body.values.items() if key in PROPOSAL_FIELDS}
    if not values:
        return old
    if "status" in values and values["status"] != old["status"]:
        return change_status(proposal_id, StatusInput(status=str(values.pop("status"))))
    values["data_atualizacao"] = now()
    with db() as conn:
        conn.execute("UPDATE propostas SET " + ",".join(f"{key}=?" for key in values) + " WHERE id=?", (*values.values(), proposal_id))
        merged = {**old, **values}
        save_client(conn, merged)
    return proposal_or_404(proposal_id)


@app.delete("/api/proposals/{proposal_id}")
def delete_proposal(proposal_id: int) -> None:
    proposal_or_404(proposal_id)
    with db() as conn:
        conn.execute("DELETE FROM propostas WHERE id=?", (proposal_id,))


@app.post("/api/proposals/{proposal_id}/status")
def change_status(proposal_id: int, body: StatusInput) -> dict:
    old = proposal_or_404(proposal_id)
    closing = now() if body.status in CLOSED and old["status"] not in CLOSED else (None if body.status not in CLOSED else old.get("data_encerramento"))
    with db() as conn:
        conn.execute("UPDATE propostas SET status=?,data_atualizacao=?,data_encerramento=? WHERE id=?", (body.status, now(), closing, proposal_id))
    history(proposal_id, old["status"], body.status, body.observacao)
    return proposal_or_404(proposal_id)


@app.post("/api/proposals/{proposal_id}/notes", status_code=201)
def add_note(proposal_id: int, body: NoteInput) -> dict:
    proposal_or_404(proposal_id)
    with db() as conn:
        cursor = conn.execute("INSERT INTO anotacoes(proposta_id,data_hora,texto) VALUES (?,?,?)", (proposal_id, now(), body.texto))
    return {"id": cursor.lastrowid, "texto": body.texto}


@app.post("/api/proposals/{proposal_id}/attachments", status_code=201)
async def upload_attachment(proposal_id: int, files: list[UploadFile] = File(...)) -> dict:
    proposal = proposal_or_404(proposal_id)
    folder = DATA_DIR / "uploads" / f"{proposal_id}-{re.sub(r'[^a-zA-Z0-9_-]', '-', proposal['nome'])}"
    folder.mkdir(parents=True, exist_ok=True)
    saved = []
    with db() as conn:
        for file in files:
            filename = f"{uuid.uuid4().hex}_{Path(file.filename or 'arquivo').name}"
            target = folder / filename
            target.write_bytes(await file.read())
            cursor = conn.execute("INSERT INTO anexos(proposta_id,nome_original,nome_arquivo,caminho,data_upload) VALUES (?,?,?,?,?)", (proposal_id, file.filename, filename, str(target), now()))
            saved.append({"id": cursor.lastrowid, "name": file.filename})
    return {"items": saved}


@app.get("/api/attachments/{attachment_id}")
def attachment_info(attachment_id: int) -> dict:
    with db() as conn:
        item = conn.execute("SELECT * FROM anexos WHERE id=?", (attachment_id,)).fetchone()
    if not item: raise HTTPException(404, "Anexo não encontrado")
    return dict(item)


@app.get("/api/attachments/{attachment_id}/download")
def download_attachment(attachment_id: int) -> FileResponse:
    item = attachment_info(attachment_id)
    if not Path(item["caminho"]).exists():
        raise HTTPException(404, "Arquivo não encontrado no armazenamento local")
    return FileResponse(item["caminho"], filename=item["nome_original"])


@app.delete("/api/attachments/{attachment_id}")
def delete_attachment(attachment_id: int) -> dict:
    item = attachment_info(attachment_id)
    with db() as conn:
        conn.execute("DELETE FROM anexos WHERE id=?", (attachment_id,))
    Path(item["caminho"]).unlink(missing_ok=True)
    return {"deleted": attachment_id}


@app.post("/api/proposals/{proposal_id}/verification")
def set_verification(proposal_id: int, verified: bool = True) -> dict:
    proposal_or_404(proposal_id)
    value = now() if verified else None
    with db() as conn:
        conn.execute("UPDATE propostas SET data_verificacao=?,data_atualizacao=? WHERE id=?", (value, now(), proposal_id))
    history(proposal_id, None, None, "Verificação marcada" if verified else "Verificação removida")
    return proposal_or_404(proposal_id)


@app.post("/api/proposals/{proposal_id}/contacted-today")
def contacted_today(proposal_id: int) -> dict:
    proposal_or_404(proposal_id)
    history(proposal_id, None, None, "Cliente contatado hoje")
    return {"ok": True}


@app.post("/api/proposals/{proposal_id}/linked-refin", status_code=201)
def create_linked_refin(proposal_id: int) -> dict:
    source = proposal_or_404(proposal_id)
    values = {key: source.get(key) for key in PROPOSAL_FIELDS}
    values.update({"produto": "Refin", "status": "Aguardando inserção", "numero_port_vinculada": source.get("numero_proposta") or str(proposal_id), "numero_proposta": ""})
    return create_proposal(ProposalInput(values=values))


@app.get("/api/clients/by-cpf/{cpf}")
def clients_by_cpf(cpf: str) -> list[dict]:
    normalized = re.sub(r"\D", "", cpf)
    with db() as conn:
        return rows(conn.execute("SELECT * FROM clientes WHERE replace(replace(replace(cpf,'.',''),'-',''),' ','')=? ORDER BY data_atualizacao DESC", (normalized,)))


@app.get("/api/board")
def board(closed: bool = False, mes: str = "") -> dict:
    with db() as conn:
        stages = rows(conn.execute("SELECT * FROM status_etapas WHERE ativo=1 ORDER BY ordem,nome"))
        condition = "status IN (" + ",".join("?" for _ in CLOSED) + ")" if closed else "status NOT IN (" + ",".join("?" for _ in CLOSED) + ")"
        params: list[Any] = list(CLOSED)
        if closed and mes:
            condition += " AND substr(COALESCE(data_encerramento,data_atualizacao,data_criacao),1,7)=?"; params.append(mes)
        proposals = rows(conn.execute(f"SELECT * FROM propostas WHERE {condition} ORDER BY data_atualizacao DESC", params))
    columns: dict[str, list[dict]] = {stage["nome"]: [] for stage in stages}
    for item in proposals: columns.setdefault(item["status"], []).append(item)
    return {"stages": stages, "columns": columns}


@app.get("/api/tasks")
def tasks(status: str = "", date_from: str = "", date_to: str = "") -> list[dict]:
    clauses, params = [], []
    for field, value, op in (("status", status, "="), ("data_tarefa", date_from, ">="), ("data_tarefa", date_to, "<=")):
        if value: clauses.append(f"{field}{op}?"); params.append(value)
    with db() as conn:
        return rows(conn.execute("SELECT t.*,p.nome proposta_nome FROM tarefas t LEFT JOIN propostas p ON p.id=t.proposta_id" + (" WHERE " + " AND ".join(clauses) if clauses else "") + " ORDER BY data_tarefa,horario", params))


@app.post("/api/tasks/{task_id}/{action}")
def task_action(task_id: int, action: str) -> dict:
    mapping = {"conclude": "concluida", "postpone": "adiada", "cancel": "cancelada"}
    if action not in mapping: raise HTTPException(422, "Ação de tarefa inválida")
    with db() as conn:
        found = conn.execute("SELECT * FROM tarefas WHERE id=?", (task_id,)).fetchone()
        if not found: raise HTTPException(404, "Tarefa não encontrada")
        done = now() if action == "conclude" else found["concluido_em"]
        conn.execute("UPDATE tarefas SET status=?,concluido_em=?,atualizado_em=? WHERE id=?", (mapping[action], done, now(), task_id))
        return dict(conn.execute("SELECT * FROM tarefas WHERE id=?", (task_id,)).fetchone())


@app.post("/api/tasks", status_code=201)
def create_task(body: TaskInput) -> dict:
    values = body.model_dump(); timestamp = now()
    with db() as conn:
        cursor = conn.execute("INSERT INTO tarefas(titulo,descricao,data_tarefa,horario,prioridade,status,categoria,proposta_id,criado_em,atualizado_em) VALUES (?,?,?,?,?,?,?,?,?,?)", (*values.values(), timestamp, timestamp))
        item = conn.execute("SELECT * FROM tarefas WHERE id=?", (cursor.lastrowid,)).fetchone()
    return dict(item)


@app.patch("/api/tasks/{task_id}")
def update_task(task_id: int, body: TaskInput) -> dict:
    with db() as conn:
        found = conn.execute("SELECT * FROM tarefas WHERE id=?", (task_id,)).fetchone()
        if not found: raise HTTPException(404, "Tarefa não encontrada")
        values = body.model_dump(); values["atualizado_em"] = now()
        if values["status"] == "concluida": values["concluido_em"] = now()
        conn.execute("UPDATE tarefas SET " + ",".join(f"{k}=?" for k in values) + " WHERE id=?", (*values.values(), task_id))
        return dict(conn.execute("SELECT * FROM tarefas WHERE id=?", (task_id,)).fetchone())


@app.get("/api/templates/{kind}")
def templates(kind: str) -> list[dict]:
    table = "modelos_gerador_mensagens" if kind == "generator" else "modelos_mensagens"
    with db() as conn: return rows(conn.execute(f"SELECT * FROM {table} ORDER BY ordem,nome"))


@app.post("/api/templates/{kind}")
def save_template(kind: str, body: TemplateInput) -> dict:
    table = "modelos_gerador_mensagens" if kind == "generator" else "modelos_mensagens"
    with db() as conn:
        conn.execute(f"INSERT INTO {table}(nome,texto,ordem,data_atualizacao) VALUES (?,?,?,?) ON CONFLICT(nome) DO UPDATE SET texto=excluded.texto,ordem=excluded.ordem,data_atualizacao=excluded.data_atualizacao", (body.nome,body.texto,body.ordem,now()))
    return body.model_dump()


@app.delete("/api/templates/{kind}/{name}")
def delete_template(kind: str, name: str) -> None:
    table = "modelos_gerador_mensagens" if kind == "generator" else "modelos_mensagens"
    with db() as conn: conn.execute(f"DELETE FROM {table} WHERE nome=?", (name,))


@app.get("/api/stages")
def stages() -> list[dict]:
    with db() as conn: return rows(conn.execute("SELECT * FROM status_etapas ORDER BY ordem,nome"))


@app.post("/api/stages", status_code=201)
def create_stage(body: StageInput) -> dict:
    with db() as conn:
        cursor = conn.execute("INSERT INTO status_etapas(nome,grupo,ordem,ativo) VALUES(?,?,?,?)", (body.nome,body.grupo,body.ordem,int(body.ativo)))
        return dict(conn.execute("SELECT * FROM status_etapas WHERE id=?", (cursor.lastrowid,)).fetchone())


@app.put("/api/stages/{stage_id}")
def update_stage(stage_id: int, body: StageInput) -> dict:
    with db() as conn:
        conn.execute("UPDATE status_etapas SET nome=?,grupo=?,ordem=?,ativo=? WHERE id=?", (body.nome,body.grupo,body.ordem,int(body.ativo),stage_id))
        item=conn.execute("SELECT * FROM status_etapas WHERE id=?", (stage_id,)).fetchone()
    if not item: raise HTTPException(404, "Etapa não encontrada")
    return dict(item)


@app.get("/api/dashboard")
def dashboard(mes: str = "") -> dict:
    with db() as conn:
        monthly = "substr(data_criacao,1,7)=?" if mes else "1=1"; params = [mes] if mes else []
        total = conn.execute(f"SELECT count(*) FROM propostas WHERE {monthly}", params).fetchone()[0]
        by_status = rows(conn.execute(f"SELECT status,count(*) total,sum(troco) troco,sum(comissao) comissao FROM propostas WHERE {monthly} GROUP BY status ORDER BY total DESC", params))
        closed = rows(conn.execute("SELECT status,count(*) total,sum(comissao) comissao FROM propostas WHERE status IN ('Pago','Perdido / Cancelado')" + (" AND substr(COALESCE(data_encerramento,data_atualizacao,data_criacao),1,7)=?" if mes else "") + " GROUP BY status", params))
    return {"mes": mes, "total": total, "by_status": by_status, "closed": closed}


@app.post("/api/notifications/read")
def mark_notifications_read() -> dict:
    with db() as conn:
        conn.execute("INSERT INTO notificacoes_leitura(id,lido_ate) VALUES(1,?) ON CONFLICT(id) DO UPDATE SET lido_ate=excluded.lido_ate", (now(),))
    return {"ok": True}


@app.get("/api/export/{format}")
def export_proposals(format: str) -> StreamingResponse:
    with db() as conn: data = rows(conn.execute("SELECT * FROM propostas ORDER BY id"))
    if format == "csv":
        output = io.StringIO(); writer = csv.DictWriter(output, fieldnames=list(data[0]) if data else ["id"]); writer.writeheader(); writer.writerows(data)
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=propostas.csv"})
    book=Workbook(); sheet=book.active; sheet.title="Propostas"
    if data: sheet.append(list(data[0])); [sheet.append(list(row.values())) for row in data]
    stream=io.BytesIO(); book.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment; filename=propostas.xlsx"})


@app.post("/api/import")
async def import_proposals(file: UploadFile = File(...)) -> dict:
    raw = await file.read(); imported = 0
    if (file.filename or "").lower().endswith(".csv"):
        records = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    else:
        sheet = load_workbook(io.BytesIO(raw), data_only=True).active; headers=[str(x.value or "").strip() for x in next(sheet.iter_rows())]; records=[dict(zip(headers,[x.value for x in row])) for row in sheet.iter_rows(min_row=2)]
    for record in records:
        allowed={key: record.get(key) for key in PROPOSAL_FIELDS if record.get(key) is not None}
        if allowed.get("nome"):
            create_proposal(ProposalInput(values=allowed)); imported += 1
    return {"imported": imported}


@app.post("/api/contacts/convert")
async def convert_contacts(file: UploadFile = File(...)) -> StreamingResponse:
    raw=await file.read(); text=raw.decode("utf-8-sig", errors="replace")
    reader=csv.DictReader(io.StringIO(text)); output=io.StringIO(); writer=csv.DictWriter(output,fieldnames=["nome","telefone","cpf"]); writer.writeheader()
    for row in reader:
        normalized={str(key).lower(): value for key,value in row.items()}; writer.writerow({"nome":normalized.get("nome", ""),"telefone":re.sub(r"\D", "", normalized.get("telefone", normalized.get("celular", "")) or ""),"cpf":normalized.get("cpf", "")})
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=contatos_convertidos.csv"})


@app.post("/api/simulator/inss")
def simulate_inss(values: dict[str, Any]) -> dict:
    coefficient=float(str(values.get("coeficiente", 0.024088)).replace(",",".")); value=float(str(values.get("valor_base", 0)).replace(",",".") or 0); margin=float(str(values.get("margem", 0)).replace(",",".") or 0)
    installment=value*coefficient if value else margin; estimated=margin/coefficient if margin and not value else value
    return {"coeficiente":coefficient,"valor_estimado":round(estimated,2),"parcela_estimada":round(installment,2),"mensagem":f"Simulação INSS: valor estimado R$ {estimated:,.2f}; parcela R$ {installment:,.2f}."}
