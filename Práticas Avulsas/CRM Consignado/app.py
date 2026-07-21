from __future__ import annotations

import csv
import calendar
import io
import json
import os
import re
import sqlite3
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from flask import (
    Flask,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
DATABASE = BASE_DIR / "database.db"
DATA_DIR = BASE_DIR / "data"
MODELOS_PATH = DATA_DIR / "modelos_mensagens.json"
BACKUP_DIR = BASE_DIR / "backups"
MAX_BACKUPS = 30
_backup_checked_date: str | None = None

# Pasta padrão dos documentos dos clientes. A configuração salva no CRM tem
# prioridade; CRM_ANEXOS_DIR continua disponível como fallback da instalação.
ANEXOS_BASE_DIR_PADRAO = Path(
    os.environ.get(
        "CRM_ANEXOS_DIR",
        r"C:\Users\tatia\OneDrive\DOCUMENTOS FACILITA - VILA VELHA\DANIEL",
    )
)

app = Flask(__name__)
app.config["SECRET_KEY"] = "troque-esta-chave-em-producao-local"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

DEFAULT_STATUS_ETAPAS = [
    {"nome": "Novo lead", "grupo": "geral", "ordem": 1},
    {"nome": "Em atendimento", "grupo": "geral", "ordem": 2},
    {"nome": "Simulação enviada", "grupo": "geral", "ordem": 3},
    {"nome": "Aguardando documentos", "grupo": "geral", "ordem": 4},
    {"nome": "Aguardando desbloqueio", "grupo": "geral", "ordem": 5},
    {"nome": "Aguardando inserção", "grupo": "geral", "ordem": 6},
    {"nome": "Aguardando interação", "grupo": "geral", "ordem": 7},
    {"nome": "Atuação", "grupo": "geral", "ordem": 8},
    {"nome": "Em análise", "grupo": "geral", "ordem": 9},
    {"nome": "Aguardando CIP", "grupo": "geral", "ordem": 10},
    {"nome": "Aguardando Averbação", "grupo": "geral", "ordem": 11},
    {"nome": "Averbado", "grupo": "geral", "ordem": 12},
    {"nome": "Aguardando Pagamento", "grupo": "geral", "ordem": 13},
    {"nome": "Aguardando Reapresentação", "grupo": "geral", "ordem": 14},
    {"nome": "Pago", "grupo": "geral", "ordem": 15},
    {"nome": "Perdido / Cancelado", "grupo": "geral", "ordem": 16},
]

# Fallback usado antes da criação do banco. A partir da v4 o funil é único.
STATUS_LIST = [s["nome"] for s in DEFAULT_STATUS_ETAPAS]
STATUS_VENDEDOR = STATUS_LIST
STATUS_ADMINISTRATIVO = STATUS_LIST
STATUS_ENCERRADOS = ["Pago", "Perdido / Cancelado"]
STATUS_COMISSAO_PREVISTA = (
    "Aguardando CIP",
    "Refin da Port",
    "Aguardando Averbação",
    "Averbado",
    "Aguardando Reapresentação",
)
INSS_PRAZO_PADRAO = "108_carencia"

TIPOS_CLIENTE = ["INSS", "SIAPE"]
PRODUTOS = ["Portabilidade", "Portabilidade com Refinanciamento", "Refinanciamento", "Novo", "Cartão", "Saque Complementar", "Outro"]
TAREFAS_HOJE_PRIORIDADE = ("paradas", "cip", "averbacao", "pagamento", "reapresentacao", "bloqueado", "acompanhamento")
TAREFAS_HOJE_INFO = {
    "paradas": {"titulo": "Sem interação recente", "vazia": "Nenhuma proposta sem interação recente."},
    "cip": {"titulo": "Aguardando CIP", "vazia": "Nenhuma proposta aguardando CIP."},
    "averbacao": {"titulo": "Aguardando Averbação", "vazia": "Nenhuma proposta aguardando averbação."},
    "pagamento": {"titulo": "Aguardando Pagamento", "vazia": "Nenhuma proposta aguardando pagamento."},
    "reapresentacao": {"titulo": "Reapresentação", "vazia": "Nenhuma proposta em reapresentação."},
    "bloqueado": {"titulo": "Benefício bloqueado", "vazia": "Nenhuma proposta com benefício bloqueado."},
    "acompanhamento": {"titulo": "Demais em acompanhamento", "vazia": "Nenhuma proposta em acompanhamento geral."},
}
DIAS_PARADA_OPERACIONAL = 3
TAREFA_STATUS = ("pendente", "concluida", "adiada", "cancelada")
TAREFA_PRIORIDADES = ("baixa", "normal", "alta")
TAREFA_CATEGORIAS = ("Ligação", "WhatsApp", "Conferência", "Retorno", "Pagamento", "Documentação", "Administrativo", "Outro")
AGENDA_ANTECEDENCIAS_MINUTOS = (5, 10, 15, 30, 60)
AGENDA_ANTECEDENCIA_PADRAO = 10
DASHBOARD_CAMPO_TIPOS = ("numero", "moeda", "percentual")
DASHBOARD_CAMPO_MODALIDADES = ("manual", "agregado", "formula")
DASHBOARD_CAMPO_CORES = ("azul", "verde", "laranja", "vermelho", "roxo", "neutro")
DASHBOARD_AGREGACOES = ("contagem", "soma", "media")
DASHBOARD_FORMULA_OPERACOES = {
    "somar": "Somar (+)",
    "subtrair": "Subtrair (−)",
    "multiplicar": "Multiplicar (×)",
    "dividir": "Dividir (÷)",
    "percentual_de": "Percentual de",
    "variacao_percentual": "Variação percentual",
    "media": "Média",
    "maior": "Maior valor",
    "menor": "Menor valor",
    "diferenca_absoluta": "Diferença absoluta",
}
DASHBOARD_FORMULA_OPERANDO_TIPOS = {
    "indicador": "Outro indicador",
    "valor_fixo": "Valor fixo",
    "percentual_fixo": "Percentual fixo",
}
DASHBOARD_AGREGACAO_CAMPOS = {
    "troco": "Troco",
    "comissao": "Comissão",
    "parcela_atual": "Parcela atual",
    "nova_parcela": "Nova parcela",
    "comissao_percentual": "Comissão percentual",
}
DASHBOARD_FILTRO_CAMPOS = {
    "": "Sem filtro adicional",
    "status": "Status",
    "produto": "Produto",
    "banco_digitado": "Banco digitado",
    "promotora": "Promotora",
    "valor_caiu_promotora": "Valor caiu na promotora",
    "valor_sacado": "Valor sacado",
}
DASHBOARD_INDICADORES_BASE = {
    "total": {"nome": "Propostas criadas", "tipo": "numero"},
    "pagas": {"nome": "Propostas pagas", "tipo": "numero"},
    "perdidas": {"nome": "Propostas perdidas", "tipo": "numero"},
    "troco_previsto": {"nome": "Troco previsto", "tipo": "moeda"},
    "troco_pago": {"nome": "Valor pago no mês", "tipo": "moeda"},
    "comissao_prevista": {"nome": "Comissão prevista", "tipo": "moeda"},
    "comissao_paga": {"nome": "Comissão paga no mês", "tipo": "moeda"},
    "valor_a_sacar": {"nome": "Valor a sacar", "tipo": "moeda"},
    "falta_cair_promotora": {"nome": "Falta cair na promotora", "tipo": "moeda"},
    "valor_ja_sacado": {"nome": "Já sacado", "tipo": "moeda"},
    "saldo_em_conta": {"nome": "Saldo em conta", "tipo": "moeda"},
    "valor_a_receber": {"nome": "Valor a receber", "tipo": "moeda"},
    "valor_previsto": {"nome": "Valor previsto", "tipo": "moeda"},
}
NOTIFICACOES_IMPORTANTES_OBSERVACOES = (
    "Proposta criada%",
    "Proposta importada%",
)
DESTINOS_INTERNOS_PREFIXOS = (
    "/propostas",
    "/funil",
    "/hoje",
    "/agenda",
    "/tarefas",
    "/encerradas",
    "/dashboard",
    "/nova",
    "/simulador-inss",
    "/gerador-mensagens",
    "/converter-contatos",
    "/mensagens",
    "/configuracoes",
    "/configuracoes/status",
    "/proposta/",
)

# Coeficientes extraídos da aba INSS do arquivo "SIMULADOR 05.06.2026.xlsx".
# Novo por valor: parcela = valor * coeficiente.
# Novo por margem: valor liberado = margem / coeficiente.
INSS_NOVO_COEFICIENTES = {
    "36": {"label": "36x", "coeficiente": 0.040573, "idade": "Até 78 anos"},
    "48": {"label": "48x", "coeficiente": 0.033523, "idade": "Até 77 anos"},
    "60": {"label": "60x", "coeficiente": 0.029420, "idade": "Até 76 anos"},
    "72": {"label": "72x", "coeficiente": 0.026789, "idade": "Até 75 anos"},
    "84": {"label": "84x", "coeficiente": 0.024995, "idade": "Até 74 anos"},
    "96": {"label": "96x", "coeficiente": 0.023720, "idade": "Até 73 anos"},
    "108": {"label": "108x", "coeficiente": 0.022227, "idade": "Até 72 anos"},
    "108_carencia": {"label": "108x com carência 90d", "coeficiente": 0.024088, "idade": "Até 71 anos"},
}

# Cartão INSS por margem, conforme aba INSS do simulador.
INSS_CARTAO_COEFICIENTES = {
    "ate_74": {"label": "Até 74 anos", "coeficiente": 0.04465},
    "75": {"label": "75 anos", "coeficiente": 0.05450},
    "76": {"label": "76 anos", "coeficiente": 0.06654},
}

CAMPOS_PROPOSTA = [
    "nome",
    "cpf",
    "nascimento",
    "nb_matricula",
    "especie",
    "numero_proposta",
    "numero_port_vinculada",
    "numero_refin_vinculada",
    "tipo_cliente",
    "banco_atual",
    "banco_destino",
    "banco_digitado",
    "produto",
    "promotora",
    "beneficio_bloqueado",
    "valor_caiu_promotora",
    "valor_sacado",
    "data_verificacao",
    "parcela_atual",
    "nova_parcela",
    "troco",
    "comissao_percentual",
    "comissao",
    "margem_apos",
    "status",
    "responsavel",
    "telefone",
    "endereco",
    "dados_bancarios",
    "data_encerramento",
    "proxima_acao",
    "data_retorno",
    "observacoes",
]

COLUNAS_IMPORTACAO = CAMPOS_PROPOSTA.copy()


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception: Exception | None = None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = get_db()
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS propostas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            nome TEXT NOT NULL,
            cpf TEXT,
            nascimento TEXT,
            nb_matricula TEXT,
            especie TEXT,
            numero_proposta TEXT,
            numero_port_vinculada TEXT,
            numero_refin_vinculada TEXT,
            tipo_cliente TEXT,
            banco_atual TEXT,
            banco_destino TEXT,
            banco_digitado TEXT,
            produto TEXT,
            promotora TEXT,
            beneficio_bloqueado TEXT DEFAULT 'NÃO',
            valor_caiu_promotora TEXT DEFAULT 'NÃO',
            valor_sacado TEXT DEFAULT 'NÃO',
            data_verificacao TEXT,
            parcela_atual REAL DEFAULT 0,
            nova_parcela REAL DEFAULT 0,
            troco REAL DEFAULT 0,
            comissao_percentual REAL DEFAULT 0,
            comissao REAL DEFAULT 0,
            margem_apos TEXT,
            status TEXT NOT NULL DEFAULT 'Novo lead',
            responsavel TEXT,
            telefone TEXT,
            endereco TEXT,
            dados_bancarios TEXT,
            data_criacao TEXT NOT NULL,
            data_atualizacao TEXT NOT NULL,
            data_encerramento TEXT,
            proxima_acao TEXT,
            data_retorno TEXT,
            observacoes TEXT
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS historico (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proposta_id INTEGER NOT NULL,
            data_hora TEXT NOT NULL,
            status_anterior TEXT,
            status_novo TEXT,
            observacao TEXT,
            FOREIGN KEY (proposta_id) REFERENCES propostas(id) ON DELETE CASCADE
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS notificacoes_importantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proposta_id INTEGER,
            proposta_nome TEXT,
            proposta_numero TEXT,
            tipo TEXT NOT NULL,
            titulo TEXT NOT NULL,
            mensagem TEXT NOT NULL,
            data_hora TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_notificacoes_importantes_data
        ON notificacoes_importantes(data_hora DESC, id DESC)
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS notificacoes_leitura (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            lido_ate TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS configuracoes (
            chave TEXT PRIMARY KEY,
            valor TEXT NOT NULL,
            atualizado_em TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS anotacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proposta_id INTEGER NOT NULL,
            data_hora TEXT NOT NULL,
            texto TEXT NOT NULL,
            FOREIGN KEY (proposta_id) REFERENCES propostas(id) ON DELETE CASCADE
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proposta_id INTEGER NOT NULL,
            nome_original TEXT NOT NULL,
            nome_arquivo TEXT NOT NULL,
            caminho TEXT NOT NULL,
            data_upload TEXT NOT NULL,
            FOREIGN KEY (proposta_id) REFERENCES propostas(id) ON DELETE CASCADE
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS status_etapas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            grupo TEXT NOT NULL DEFAULT 'administrativo',
            ordem INTEGER NOT NULL DEFAULT 0,
            ativo INTEGER NOT NULL DEFAULT 1
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS modelos_mensagens (
            nome TEXT PRIMARY KEY,
            texto TEXT NOT NULL,
            ordem INTEGER NOT NULL DEFAULT 0,
            data_atualizacao TEXT
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS modelos_gerador_mensagens (
            nome TEXT PRIMARY KEY,
            texto TEXT NOT NULL,
            ordem INTEGER NOT NULL DEFAULT 0,
            data_atualizacao TEXT
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cpf TEXT NOT NULL,
            nb_matricula TEXT NOT NULL DEFAULT '',
            telefone TEXT,
            tipo_cliente TEXT,
            endereco TEXT,
            dados_bancarios TEXT,
            data_criacao TEXT NOT NULL,
            data_atualizacao TEXT NOT NULL,
            UNIQUE(cpf, nb_matricula)
        )
        """
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tarefas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            descricao TEXT,
            data_tarefa TEXT NOT NULL,
            horario TEXT,
            prioridade TEXT NOT NULL DEFAULT 'normal',
            status TEXT NOT NULL DEFAULT 'pendente',
            categoria TEXT NOT NULL DEFAULT 'Retorno',
            proposta_id INTEGER,
            criado_em TEXT NOT NULL,
            concluido_em TEXT,
            atualizado_em TEXT NOT NULL,
            notificar INTEGER NOT NULL DEFAULT 0,
            notificado_em TEXT,
            FOREIGN KEY (proposta_id) REFERENCES propostas(id) ON DELETE SET NULL
        )
        """
    )

    # Migrações incrementais: adiciona colunas em bancos criados nas versões anteriores.
    colunas_propostas = {row["name"] for row in db.execute("PRAGMA table_info(propostas)").fetchall()}
    if "cliente_id" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN cliente_id INTEGER")
    if "numero_proposta" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN numero_proposta TEXT")
    if "numero_port_vinculada" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN numero_port_vinculada TEXT")
    if "numero_refin_vinculada" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN numero_refin_vinculada TEXT")
    if "comissao" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN comissao REAL DEFAULT 0")
    if "comissao_percentual" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN comissao_percentual REAL DEFAULT 0")
    if "promotora" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN promotora TEXT")
    if "beneficio_bloqueado" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN beneficio_bloqueado TEXT DEFAULT 'NÃO'")
    if "banco_digitado" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN banco_digitado TEXT")
    if "valor_caiu_promotora" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN valor_caiu_promotora TEXT DEFAULT 'NÃO'")
    if "valor_sacado" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN valor_sacado TEXT DEFAULT 'NÃO'")
    if "data_verificacao" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN data_verificacao TEXT")
    if "data_encerramento" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN data_encerramento TEXT")
    if "endereco" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN endereco TEXT")
    if "dados_bancarios" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN dados_bancarios TEXT")
    if "especie" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN especie TEXT")
    if "nascimento" not in colunas_propostas:
        db.execute("ALTER TABLE propostas ADD COLUMN nascimento TEXT")

    # Completa dados de cliente que ficaram vazios em refins vinculados criados
    # antes de nascimento e espécie fazerem parte desse fluxo.
    db.execute(
        """
        UPDATE propostas
        SET
            nascimento = COALESCE(
                NULLIF(TRIM(nascimento), ''),
                (
                    SELECT NULLIF(TRIM(origem.nascimento), '')
                    FROM propostas origem
                    WHERE UPPER(TRIM(origem.numero_proposta)) = UPPER(TRIM(propostas.numero_port_vinculada))
                    LIMIT 1
                ),
                ''
            ),
            especie = COALESCE(
                NULLIF(TRIM(especie), ''),
                (
                    SELECT NULLIF(TRIM(origem.especie), '')
                    FROM propostas origem
                    WHERE UPPER(TRIM(origem.numero_proposta)) = UPPER(TRIM(propostas.numero_port_vinculada))
                    LIMIT 1
                ),
                ''
            ),
            tipo_cliente = COALESCE(
                NULLIF(TRIM(tipo_cliente), ''),
                (
                    SELECT COALESCE(
                        NULLIF(TRIM(origem.tipo_cliente), ''),
                        NULLIF(TRIM(cliente.tipo_cliente), ''),
                        CASE
                            WHEN COALESCE(TRIM(origem.especie), '') <> '' THEN 'INSS'
                            ELSE NULL
                        END
                    )
                    FROM propostas origem
                    LEFT JOIN clientes cliente ON cliente.id = origem.cliente_id
                    WHERE UPPER(TRIM(origem.numero_proposta)) = UPPER(TRIM(propostas.numero_port_vinculada))
                    LIMIT 1
                ),
                ''
            )
        WHERE UPPER(TRIM(COALESCE(produto, ''))) = 'REFINANCIAMENTO'
          AND COALESCE(TRIM(numero_port_vinculada), '') <> ''
          AND (
              COALESCE(TRIM(nascimento), '') = ''
              OR COALESCE(TRIM(especie), '') = ''
              OR COALESCE(TRIM(tipo_cliente), '') = ''
          )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS dashboard_campos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'numero',
            modalidade TEXT NOT NULL DEFAULT 'manual',
            configuracao_json TEXT NOT NULL DEFAULT '{}',
            cor TEXT NOT NULL DEFAULT 'azul',
            ordem INTEGER NOT NULL DEFAULT 0,
            ativo INTEGER NOT NULL DEFAULT 1,
            criado_em TEXT NOT NULL,
            atualizado_em TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS dashboard_valores_manuais (
            campo_id INTEGER NOT NULL,
            mes TEXT NOT NULL,
            valor REAL NOT NULL DEFAULT 0,
            atualizado_em TEXT NOT NULL,
            PRIMARY KEY (campo_id, mes),
            FOREIGN KEY (campo_id) REFERENCES dashboard_campos(id) ON DELETE CASCADE
        )
        """
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_dashboard_campos_ordem ON dashboard_campos(ativo, ordem, id)")
    db.execute(
        """
        UPDATE propostas
        SET data_verificacao = SUBSTR(data_criacao, 1, 10)
        WHERE COALESCE(TRIM(data_verificacao), '') = ''
          AND COALESCE(TRIM(data_criacao), '') <> ''
        """
    )

    colunas_tarefas = {row["name"] for row in db.execute("PRAGMA table_info(tarefas)").fetchall()}
    if "titulo" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN titulo TEXT NOT NULL DEFAULT 'Tarefa'")
    if "descricao" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN descricao TEXT")
    if "data_tarefa" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN data_tarefa TEXT NOT NULL DEFAULT ''")
    if "horario" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN horario TEXT")
    if "prioridade" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN prioridade TEXT NOT NULL DEFAULT 'normal'")
    if "status" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN status TEXT NOT NULL DEFAULT 'pendente'")
    if "categoria" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN categoria TEXT NOT NULL DEFAULT 'Retorno'")
    if "proposta_id" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN proposta_id INTEGER")
    if "criado_em" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN criado_em TEXT NOT NULL DEFAULT ''")
    if "concluido_em" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN concluido_em TEXT")
    if "atualizado_em" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN atualizado_em TEXT NOT NULL DEFAULT ''")
    if "notificar" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN notificar INTEGER NOT NULL DEFAULT 0")
    if "notificado_em" not in colunas_tarefas:
        db.execute("ALTER TABLE tarefas ADD COLUMN notificado_em TEXT")

    db.execute("CREATE INDEX IF NOT EXISTS idx_tarefas_data_status ON tarefas(data_tarefa, status)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_tarefas_proposta ON tarefas(proposta_id)")

    qtd_status = db.execute("SELECT COUNT(*) AS total FROM status_etapas").fetchone()["total"]
    if qtd_status == 0:
        db.executemany(
            "INSERT INTO status_etapas (nome, grupo, ordem, ativo) VALUES (?, ?, ?, 1)",
            [(item["nome"], item["grupo"], item["ordem"]) for item in DEFAULT_STATUS_ETAPAS],
        )

    # Migrações automáticas para bases criadas nas versões anteriores.
    # Assim, propostas antigas continuam aparecendo no funil novo.
    migracoes_status = {
        "Aguardando digitação": "Aguardando inserção",
        "Inserção": "Aguardando interação",
        "Contrato em andamento": "Atuação",
        "Perdido": "Perdido / Cancelado",
        "Cancelado": "Perdido / Cancelado",
    }
    for antigo, novo in migracoes_status.items():
        db.execute("UPDATE propostas SET status = ? WHERE status = ?", (novo, antigo))
        db.execute("UPDATE historico SET status_anterior = ? WHERE status_anterior = ?", (novo, antigo))
        db.execute("UPDATE historico SET status_novo = ? WHERE status_novo = ?", (novo, antigo))

    db.execute(
        """
        UPDATE propostas
        SET data_encerramento = COALESCE(
            (
                SELECT MAX(h.data_hora)
                FROM historico h
                WHERE h.proposta_id = propostas.id
                  AND h.status_novo IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
            ),
            data_atualizacao,
            data_criacao
        )
        WHERE status IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
          AND COALESCE(TRIM(data_encerramento), '') = ''
        """
    )
    db.execute(
        """
        UPDATE propostas
        SET data_encerramento = NULL
        WHERE status NOT IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
          AND COALESCE(TRIM(data_encerramento), '') <> ''
        """
    )

    for removido in ("Inserção", "Perdido", "Cancelado"):
        db.execute("DELETE FROM status_etapas WHERE nome = ?", (removido,))

    # A v4 usa um único funil compartilhado, sem separação por vendedor/administrativo.
    db.execute("UPDATE status_etapas SET grupo = 'geral'")

    # Garante que a etapa unificada exista mesmo em bancos antigos.
    existentes = {row["nome"] for row in db.execute("SELECT nome FROM status_etapas").fetchall()}
    for item in DEFAULT_STATUS_ETAPAS:
        if item["nome"] not in existentes:
            db.execute(
                "INSERT INTO status_etapas (nome, grupo, ordem, ativo) VALUES (?, ?, ?, 1)",
                (item["nome"], item["grupo"], item["ordem"]),
            )

    # v9: migra observações fixas para um diário/log de anotações.
    propostas_com_observacao = db.execute(
        """
        SELECT id, observacoes, data_criacao, data_atualizacao
        FROM propostas
        WHERE COALESCE(TRIM(observacoes), '') <> ''
        """
    ).fetchall()
    for proposta in propostas_com_observacao:
        ja_tem = db.execute(
            "SELECT id FROM anotacoes WHERE proposta_id = ? LIMIT 1",
            (proposta["id"],),
        ).fetchone()
        if not ja_tem:
            db.execute(
                "INSERT INTO anotacoes (proposta_id, data_hora, texto) VALUES (?, ?, ?)",
                (
                    proposta["id"],
                    proposta["data_atualizacao"] or proposta["data_criacao"] or agora_iso(),
                    proposta["observacoes"],
                ),
            )
    sincronizar_modelos_banco(db)
    migrar_clientes_a_partir_de_propostas()
    db.commit()


def modelos_padrao() -> dict[str, str]:
    return {
        "Mensagem inicial": "Olá, {nome}. Tudo bem? Tenho uma simulação de {produto} para seu contrato do banco {banco_atual}, com possibilidade de redução de parcela e liberação de valor. Posso te passar os detalhes?",
        "Solicitação de documentos": "Olá, {nome}. Para seguir com sua proposta de {produto}, preciso que me envie os documentos necessários para análise. Assim que receber, dou andamento por aqui.",
        "Proposta em andamento": "Olá, {nome}. Sua proposta de {produto} está em andamento no banco {banco_digitado}. Status atual: {status}. Assim que houver atualização, te aviso.",
        "Aguardando interação": "Olá, {nome}. Sua proposta já foi digitada e agora precisamos que você conclua a assinatura digital com selfie. Assim que concluir, me avise para seguirmos acompanhando.",
        "Em análise": "Olá, {nome}. Sua proposta está em análise documental pelo banco {banco_digitado}. Sigo acompanhando e te aviso assim que houver atualização.",
        "Aguardando CIP": "Olá, {nome}. Sua proposta está aguardando retorno da CIP. Essa etapa depende da comunicação entre os bancos, mas sigo acompanhando de perto.",
        "Aguardando averbação": "Olá, {nome}. Sua proposta está aguardando averbação. Essa é uma das etapas finais antes da liberação do pagamento.",
        "Aguardando pagamento": "Olá, {nome}. Sua proposta já avançou e está aguardando pagamento pelo banco {banco_digitado}. Assim que constar pago, te aviso.",
        "Aguardando reapresentação": "Olá, {nome}. O pagamento retornou e precisamos ajustar os dados bancários para reapresentar. Pode me confirmar a conta para pagamento?",
        "Proposta paga": "Olá, {nome}. Sua proposta consta como paga. Verifique sua conta e me avise caso precise de mais alguma informação.",
        "Falta de margem": "Olá, {nome}. No momento, a análise indicou falta de margem para seguir com a proposta. Vou manter seu cadastro acompanhado para uma nova oportunidade.",
        "Recontato": "Olá, {nome}. Passando para saber se ainda tem interesse na simulação de {produto}. Posso atualizar as condições para você sem compromisso.",
    }


def garantir_modelos() -> None:
    """Mantém um arquivo de modelos como fallback/backup.

    A partir da v22, os modelos editados ficam salvos no database.db para não
    serem perdidos quando a pasta data/ for substituída em uma atualização.
    """
    DATA_DIR.mkdir(exist_ok=True)
    if MODELOS_PATH.exists():
        return
    MODELOS_PATH.write_text(json.dumps(modelos_padrao(), ensure_ascii=False, indent=2), encoding="utf-8")


def carregar_modelos_arquivo() -> dict[str, str]:
    garantir_modelos()
    try:
        dados = json.loads(MODELOS_PATH.read_text(encoding="utf-8"))
        if isinstance(dados, dict):
            return {str(k): str(v) for k, v in dados.items()}
    except Exception:
        pass
    return modelos_padrao()


def sincronizar_modelos_banco(db: sqlite3.Connection) -> None:
    """Cria/preserva modelos no banco.

    Regras:
    - Se o banco ainda não tem modelos, importa do arquivo data/modelos_mensagens.json.
    - Se o banco já tem modelos, mantém os textos editados pelo usuário.
    - Novos modelos padrão podem ser adicionados em versões futuras sem apagar os existentes.
    """
    existentes = db.execute("SELECT nome, texto FROM modelos_mensagens ORDER BY ordem, nome").fetchall()
    if not existentes:
        origem = carregar_modelos_arquivo()
        for ordem, (nome, texto) in enumerate(origem.items(), start=1):
            db.execute(
                """
                INSERT OR REPLACE INTO modelos_mensagens (nome, texto, ordem, data_atualizacao)
                VALUES (?, ?, ?, ?)
                """,
                (nome, texto, ordem, agora_iso()),
            )
        return

    nomes_existentes = {row["nome"] for row in existentes}
    maior_ordem = db.execute("SELECT COALESCE(MAX(ordem), 0) AS maior FROM modelos_mensagens").fetchone()["maior"]
    for nome, texto in modelos_padrao().items():
        if nome not in nomes_existentes:
            maior_ordem += 1
            db.execute(
                """
                INSERT INTO modelos_mensagens (nome, texto, ordem, data_atualizacao)
                VALUES (?, ?, ?, ?)
                """,
                (nome, texto, maior_ordem, agora_iso()),
            )



def modelos_gerador_padrao() -> dict[str, str]:
    return {
        "Portabilidade com redução de parcela": """Olá, {nome},

Meu nome é {atendente} e entrei em contato para falar sobre a portabilidade do seu contrato, com redução da taxa de juros e liberação de troco.

Além disso, atualmente o INSS está permitindo uma carência de até 90 dias para o início do desconto da parcela após a realização da portabilidade.

Tenho uma proposta para seu contrato do Banco {banco}:

✅ Parcela reduzida de {parcela_antiga} para {parcela_nova}
✅ Regularizar sua margem negativa de {economia}
✅ Liberação de troco no valor de {troco}
✅ Carência de 90 dias, ficando até 3 meses sem o desconto dessa parcela
✅ Sua conta de recebimento permanece a mesma, pois somos conveniados ao INSS

Podemos dar prosseguimento e garantir essa redução na sua parcela?""",
        "Portabilidade apenas com troco": """Olá, {nome},

Meu nome é {atendente} e encontrei uma possibilidade de portabilidade para seu contrato do Banco {banco}.

Nessa opção, há possibilidade de liberação de troco no valor de {troco}, mantendo a operação sujeita à análise e aprovação do banco.

Posso seguir com a simulação para confirmar as condições?""",
        "Redução sem troco": """Olá, {nome}, tudo bem?

Meu nome é {atendente}. Fiz uma análise do seu contrato do Banco {banco} e encontrei uma possibilidade de redução de parcela.

✅ Parcela atual: {parcela_antiga}
✅ Nova parcela: {parcela_nova}
✅ Economia mensal estimada: {economia}

Posso seguir com a análise para confirmar essa condição?""",
        "Recontato de proposta": """Olá, {nome}. Tudo bem?

Meu nome é {atendente}. Estou passando para saber se ainda tem interesse na proposta do Banco {banco}.

Condição simulada:
✅ Parcela de {parcela_antiga} para {parcela_nova}
✅ Economia mensal de {economia}
✅ Valor liberado: {troco}

Posso atualizar essa simulação para você?""",
    }


def sincronizar_modelos_gerador_banco(db: sqlite3.Connection) -> None:
    existentes = db.execute("SELECT nome FROM modelos_gerador_mensagens ORDER BY ordem, nome").fetchall()
    if not existentes:
        for ordem, (nome, texto) in enumerate(modelos_gerador_padrao().items(), start=1):
            db.execute(
                """
                INSERT OR REPLACE INTO modelos_gerador_mensagens (nome, texto, ordem, data_atualizacao)
                VALUES (?, ?, ?, ?)
                """,
                (nome, texto, ordem, agora_iso()),
            )
        return

    nomes_existentes = {row["nome"] for row in existentes}
    maior_ordem = db.execute("SELECT COALESCE(MAX(ordem), 0) AS maior FROM modelos_gerador_mensagens").fetchone()["maior"]
    for nome, texto in modelos_gerador_padrao().items():
        if nome not in nomes_existentes:
            maior_ordem += 1
            db.execute(
                """
                INSERT INTO modelos_gerador_mensagens (nome, texto, ordem, data_atualizacao)
                VALUES (?, ?, ?, ?)
                """,
                (nome, texto, maior_ordem, agora_iso()),
            )


def carregar_modelos_gerador() -> dict[str, str]:
    db = get_db()
    sincronizar_modelos_gerador_banco(db)
    db.commit()
    rows = db.execute("SELECT nome, texto FROM modelos_gerador_mensagens ORDER BY ordem, nome").fetchall()
    return {row["nome"]: row["texto"] for row in rows}

def criar_backup_automatico() -> None:
    """Cria um backup diário do database.db na pasta backups.

    O backup usa a API nativa do SQLite para reduzir risco de cópia inconsistente.
    Para não gerar centenas de arquivos, cria no máximo um backup por dia e mantém
    apenas os últimos MAX_BACKUPS arquivos.
    """
    global _backup_checked_date
    hoje = date.today().strftime("%Y-%m-%d")
    if _backup_checked_date == hoje:
        return
    _backup_checked_date = hoje

    if not DATABASE.exists():
        return

    try:
        BACKUP_DIR.mkdir(exist_ok=True)
        ja_existe_hoje = any(BACKUP_DIR.glob(f"database_{hoje}_*.db"))
        if ja_existe_hoje:
            return

        destino = BACKUP_DIR / f"database_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.db"
        origem_conn = sqlite3.connect(DATABASE)
        destino_conn = sqlite3.connect(destino)
        try:
            origem_conn.backup(destino_conn)
        finally:
            destino_conn.close()
            origem_conn.close()

        backups = sorted(BACKUP_DIR.glob("database_*.db"), key=lambda item: item.stat().st_mtime, reverse=True)
        for antigo in backups[MAX_BACKUPS:]:
            try:
                antigo.unlink()
            except OSError:
                pass
    except Exception as exc:
        # Não bloqueia o CRM caso o backup falhe; apenas informa no terminal.
        print(f"Aviso: não foi possível criar backup automático: {exc}")

@app.before_request
def preparar_app() -> None:
    init_db()
    garantir_modelos()
    criar_backup_automatico()


def agora_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoje_iso() -> str:
    return date.today().strftime("%Y-%m-%d")


def limpar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def obter_configuracao(chave: str, padrao: str = "") -> str:
    registro = get_db().execute(
        "SELECT valor FROM configuracoes WHERE chave = ?",
        (chave,),
    ).fetchone()
    return limpar_texto(registro["valor"]) if registro else padrao


def salvar_configuracao(chave: str, valor: Any) -> None:
    get_db().execute(
        """INSERT INTO configuracoes (chave, valor, atualizado_em) VALUES (?, ?, ?)
           ON CONFLICT(chave) DO UPDATE SET valor = excluded.valor, atualizado_em = excluded.atualizado_em""",
        (chave, str(valor), agora_iso()),
    )


def agenda_antecedencia_minutos() -> int:
    try:
        valor = int(obter_configuracao("agenda_antecedencia_minutos", str(AGENDA_ANTECEDENCIA_PADRAO)))
    except (TypeError, ValueError):
        valor = AGENDA_ANTECEDENCIA_PADRAO
    return valor if valor in AGENDA_ANTECEDENCIAS_MINUTOS else AGENDA_ANTECEDENCIA_PADRAO


def agenda_notificar_padrao() -> bool:
    return obter_configuracao("agenda_notificar_padrao", "0") == "1"


def agenda_alerta_info() -> dict[str, Any]:
    antecedencia = agenda_antecedencia_minutos()
    limite = (datetime.now() + timedelta(minutes=antecedencia)).strftime("%Y-%m-%d %H:%M:%S")
    registro = get_db().execute(
        """SELECT COUNT(*) AS total
           FROM tarefas
           WHERE notificar = 1
             AND status IN ('pendente', 'adiada')
             AND COALESCE(TRIM(horario), '') <> ''
             AND datetime(data_tarefa || ' ' || horario) <= datetime(?)""",
        (limite,),
    ).fetchone()
    total = int(registro["total"] or 0)
    return {"ativo": total > 0, "total": total, "antecedencia_minutos": antecedencia}


def pasta_base_anexos() -> Path:
    configurada = obter_configuracao("anexos_base_dir", "")
    if configurada:
        return Path(os.path.expandvars(os.path.expanduser(configurada)))
    return ANEXOS_BASE_DIR_PADRAO


def validar_pasta_base_anexos(valor: Any) -> tuple[Path | None, str]:
    texto = limpar_texto(valor)
    if not texto:
        return None, "Informe a pasta base dos documentos."
    caminho = Path(os.path.expandvars(os.path.expanduser(texto)))
    if not caminho.is_absolute():
        return None, "Informe um caminho absoluto, por exemplo C:\\Documentos CRM."
    if caminho.exists() and not caminho.is_dir():
        return None, "O caminho informado existe, mas não é uma pasta."
    try:
        caminho.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        return None, f"Não foi possível acessar ou criar a pasta informada: {exc}"
    return caminho, ""


def url_interna_segura(valor: Any, fallback: str | None = None) -> str:
    destino = limpar_texto(valor)
    if destino:
        partes = urlsplit(destino)
        caminho_conhecido = partes.path == "/" or any(partes.path.startswith(prefixo) for prefixo in DESTINOS_INTERNOS_PREFIXOS)
        if (
            caminho_conhecido
            and partes.path.startswith("/")
            and not partes.path.startswith("//")
            and not partes.scheme
            and not partes.netloc
        ):
            return destino
    return fallback or "/funil"


def url_interna_com_parametros(valor: Any, fallback: str | None = None, **parametros: Any) -> str:
    destino = url_interna_segura(valor, fallback)
    partes = urlsplit(destino)
    query = dict(parse_qsl(partes.query, keep_blank_values=True))
    for chave, valor_parametro in parametros.items():
        if valor_parametro is not None and limpar_texto(valor_parametro) != "":
            query[chave] = str(valor_parametro)
    return urlunsplit(("", "", partes.path, urlencode(query), partes.fragment))


def url_origem_atual() -> str:
    return url_interna_segura(request.full_path.rstrip("?"), "/funil")


def url_retorno_padrao() -> str:
    return url_interna_segura(request.values.get("origem") or request.values.get("next"), "/funil")


def origem_eh_hoje() -> bool:
    return urlsplit(url_retorno_padrao()).path == "/hoje"


def nome_pasta_cliente(nome: str) -> str:
    """Gera nome seguro para pasta local do cliente, preservando leitura humana."""
    base = remover_acentos(limpar_texto(nome)).upper()
    base = re.sub(r"[^A-Z0-9 _.-]", "", base)
    base = re.sub(r"\s+", " ", base).strip(" .")
    return base or "CLIENTE SEM NOME"


def pasta_cliente(proposta: sqlite3.Row | dict[str, Any]) -> Path:
    return pasta_base_anexos() / nome_pasta_cliente(proposta["nome"])


def arquivo_destino_unico(pasta: Path, nome_arquivo: str) -> Path:
    destino = pasta / nome_arquivo
    if not destino.exists():
        return destino
    stem = destino.stem
    suffix = destino.suffix
    contador = 2
    while True:
        candidato = pasta / f"{stem}_{contador}{suffix}"
        if not candidato.exists():
            return candidato
        contador += 1


def salvar_anexos_upload(proposta_id: int, proposta: sqlite3.Row | dict[str, Any], arquivos: list[Any]) -> int:
    """Salva arquivos enviados na pasta do cliente e registra no banco.

    Usado tanto na criação da proposta quanto na tela de detalhes.
    Retorna a quantidade de arquivos salvos com sucesso.
    """
    arquivos_validos = [arquivo for arquivo in arquivos if arquivo and arquivo.filename]
    if not arquivos_validos:
        return 0

    pasta = pasta_cliente(proposta)
    try:
        pasta.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        flash(f"Não foi possível criar a pasta de anexos: {exc}", "erro")
        return 0

    db = get_db()
    salvos = 0
    for arquivo in arquivos_validos:
        nome_original = arquivo.filename
        seguro = secure_filename(nome_original) or "arquivo"
        destino = arquivo_destino_unico(pasta, seguro)
        try:
            arquivo.save(destino)
        except OSError as exc:
            flash(f"Erro ao salvar {nome_original}: {exc}", "erro")
            continue
        db.execute(
            """
            INSERT INTO anexos (proposta_id, nome_original, nome_arquivo, caminho, data_upload)
            VALUES (?, ?, ?, ?, ?)
            """,
            (proposta_id, nome_original, destino.name, str(destino), agora_iso()),
        )
        salvos += 1

    db.commit()
    return salvos


def tamanho_arquivo(caminho: str) -> str:
    try:
        size = Path(caminho).stat().st_size
    except OSError:
        return ""
    unidades = ["B", "KB", "MB", "GB"]
    valor = float(size)
    for unidade in unidades:
        if valor < 1024 or unidade == unidades[-1]:
            if unidade == "B":
                return f"{int(valor)} {unidade}"
            return f"{valor:.1f} {unidade}".replace(".", ",")
        valor /= 1024


def formatar_cpf(valor: str) -> str:
    digitos = re.sub(r"\D", "", valor or "")
    if len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    return valor.strip()


def parse_moeda(valor: Any) -> float:
    if valor is None:
        return 0.0
    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = texto.replace("R$", "").replace(" ", "")
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(\.\d{3})+", texto):
        texto = texto.replace(".", "")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def parse_percentual(valor: Any) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("%", "").replace(" ", "")
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def br_moeda(valor: Any) -> str:
    try:
        numero = float(valor or 0)
    except (TypeError, ValueError):
        numero = 0.0
    texto = f"R$ {numero:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def br_percentual(valor: Any) -> str:
    try:
        numero = float(valor or 0)
    except (TypeError, ValueError):
        numero = 0.0
    texto = f"{numero:.2f}".replace(".", ",")
    if texto.endswith(",00"):
        texto = texto[:-3]
    return f"{texto}%"


def br_data(valor: Any) -> str:
    texto = limpar_texto(valor)
    if not texto:
        return ""
    for formato in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto, formato).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return texto


def br_data_hora(valor: Any) -> str:
    texto = limpar_texto(valor)
    if not texto:
        return ""
    try:
        return datetime.strptime(texto, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return texto




def remover_acentos(texto: str) -> str:
    mapa = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ", "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC")
    return texto.translate(mapa)


def normalizar_bloqueado(valor: Any) -> str:
    texto = remover_acentos(limpar_texto(valor)).upper()
    if not texto:
        return "NÃO"
    if texto in {"SIM", "S", "BLOQUEADO", "BLOCK", "TRUE", "1", "YES"}:
        return "SIM"
    if texto in {"NAO", "N", "DESBLOQUEADO", "LIBERADO", "FALSE", "0", "NO"}:
        return "NÃO"
    return "SIM" if "BLOQUE" in texto and "DES" not in texto else "NÃO"


def normalizar_sim_nao(valor: Any) -> str:
    texto = remover_acentos(limpar_texto(valor)).upper()
    if texto in {"SIM", "S", "TRUE", "1", "YES", "PAGO", "SACADO", "CAIU", "OK"}:
        return "SIM"
    return "NÃO"


def parse_data_iso(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")
    texto = limpar_texto(valor)
    if not texto:
        return ""
    for formato in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return texto


def extrair_data_do_status(status_original: str, data_base: Any = None) -> str:
    texto = limpar_texto(status_original)
    match = re.search(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?", texto)
    if not match:
        return ""
    dia = int(match.group(1))
    mes = int(match.group(2))
    ano_txt = match.group(3)
    if ano_txt:
        ano = int(ano_txt)
        if ano < 100:
            ano += 2000
    else:
        base_iso = parse_data_iso(data_base)
        try:
            ano = datetime.strptime(base_iso[:10], "%Y-%m-%d").year
        except ValueError:
            ano = date.today().year
    try:
        return date(ano, mes, dia).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def normalizar_status_importacao(valor: Any) -> str:
    original = limpar_texto(valor)
    if not original:
        return "Novo lead"
    texto = remover_acentos(original).upper()
    texto_sem_data = re.sub(r"\b\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\b", "", texto).strip()
    texto_sem_data = re.sub(r"\s+", " ", texto_sem_data)

    aliases = {
        "NOVO": "Novo lead",
        "NOVO LEAD": "Novo lead",
        "EM ATENDIMENTO": "Em atendimento",
        "SIMULACAO ENVIADA": "Simulação enviada",
        "AG DOCUMENTOS": "Aguardando documentos",
        "AG. DOCUMENTOS": "Aguardando documentos",
        "AG DESBLOQUEIO": "Aguardando desbloqueio",
        "AG. DESBLOQUEIO": "Aguardando desbloqueio",
        "AG INSERCAO": "Aguardando inserção",
        "AG. INSERCAO": "Aguardando inserção",
        "AG INTERACAO": "Aguardando interação",
        "AG. INTERACAO": "Aguardando interação",
        "AGUARDANDO INTERACAO": "Aguardando interação",
        "ATUACAO": "Atuação",
        "EM ANALISE": "Em análise",
        "AG CIP": "Aguardando CIP",
        "AG. CIP": "Aguardando CIP",
        "AG AVERBACAO": "Aguardando Averbação",
        "AG. AVERBACAO": "Aguardando Averbação",
        "AVERBADO": "Averbado",
        "AG PAGAMENTO": "Aguardando Pagamento",
        "AG. PAGAMENTO": "Aguardando Pagamento",
        "AG REAPRESENTACAO": "Aguardando Reapresentação",
        "AG. REAPRESENTACAO": "Aguardando Reapresentação",
        "PAGO": "Pago",
        "REPROVADO": "Perdido / Cancelado",
        "PERDIDO": "Perdido / Cancelado",
        "CANCELADO": "Perdido / Cancelado",
    }
    if texto_sem_data in aliases:
        return aliases[texto_sem_data]
    if texto_sem_data.startswith("AG CIP") or texto_sem_data.startswith("AG. CIP"):
        return "Aguardando CIP"
    if texto_sem_data.startswith("AG INTERACAO") or texto_sem_data.startswith("AG. INTERACAO"):
        return "Aguardando interação"
    if texto_sem_data.startswith("AG AVERBACAO") or texto_sem_data.startswith("AG. AVERBACAO"):
        return "Aguardando Averbação"
    if texto_sem_data.startswith("REPROV"):
        return "Perdido / Cancelado"

    # Se o texto já corresponder a uma etapa cadastrada, preserva o nome oficial.
    for status in nomes_status():
        if remover_acentos(status).upper() == texto_sem_data:
            return status
    return original


def mes_atual() -> str:
    return date.today().strftime("%Y-%m")


def status_encerrado(status: str) -> bool:
    return limpar_texto(status) in STATUS_ENCERRADOS


def data_encerramento_para_status(status_anterior: str | None, status_novo: str | None, data_atual: Any = None) -> str | None:
    if status_encerrado(status_novo or ""):
        return limpar_texto(data_atual) or agora_iso()
    return None


def produto_portabilidade(produto: Any) -> bool:
    texto = remover_acentos(limpar_texto(produto)).casefold()
    return texto in {"portabilidade", "portabilidade com refinanciamento"}


def banco_digitado_exibicao(proposta: Any) -> str:
    try:
        dados = dict(proposta)
    except Exception:
        dados = proposta or {}
    return limpar_texto(dados.get("banco_digitado"))


def status_ativos() -> list[str]:
    return [status for status in nomes_status() if status not in STATUS_ENCERRADOS]


def resumo_coluna_funil(status: str) -> dict[str, Any]:
    row = get_db().execute(
        """
        SELECT COUNT(*) AS quantidade, COALESCE(SUM(comissao), 0) AS comissao
        FROM propostas
        WHERE status = ?
        """,
        (status,),
    ).fetchone()
    comissao = float(row["comissao"] or 0) if row else 0.0
    return {
        "quantidade": int(row["quantidade"] or 0) if row else 0,
        "comissao": br_moeda(comissao),
    }


def resposta_status_json(
    sucesso: bool,
    mensagem: str,
    proposta_id: int | None = None,
    etapa_origem: str | None = None,
    etapa_destino: str | None = None,
    status_http: int = 200,
):
    payload: dict[str, Any] = {
        "success": sucesso,
        "ok": sucesso,
        "message": mensagem,
    }
    if not sucesso:
        payload["erro"] = mensagem
    if proposta_id is not None:
        payload["proposta_id"] = proposta_id
    if etapa_origem is not None:
        payload["etapa_origem"] = etapa_origem
    if etapa_destino is not None:
        payload["etapa_destino"] = etapa_destino
    if etapa_origem is not None and etapa_destino is not None:
        payload["colunas"] = {
            "origem": resumo_coluna_funil(etapa_origem),
            "destino": resumo_coluna_funil(etapa_destino),
        }
    return jsonify(payload), status_http


def mes_nome_para_numero(nome: str) -> int | None:
    texto = remover_acentos(limpar_texto(nome)).upper()
    mapa = {
        "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "MARÇO": 3,
        "ABRIL": 4, "MAIO": 5, "JUNHO": 6, "JULHO": 7,
        "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10,
        "NOVEMBRO": 11, "DEZEMBRO": 12,
    }
    return mapa.get(texto)


def nome_mes_pt(mes: str) -> str:
    nomes = [
        "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
        "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
    ]
    try:
        numero = int((mes or "").split("-")[1])
        return nomes[numero - 1]
    except Exception:
        return ""


def linha_no_mes(row: dict[str, Any], mes_importacao: str) -> bool:
    if not mes_importacao:
        return True
    data_iso = parse_data_iso(row.get("data_criacao"))
    if not data_iso:
        # Sem data: importa para não descartar manualmente uma linha válida.
        return True
    return data_iso[:7] == mes_importacao


def normalizar_produto_importacao(valor: Any) -> str:
    original = limpar_texto(valor)
    texto = remover_acentos(original).upper()
    aliases = {
        "PORT": "Portabilidade",
        "PORTABILIDADE": "Portabilidade",
        "PORT COM REFIN": "Portabilidade com Refinanciamento",
        "PORT + REFIN": "Portabilidade com Refinanciamento",
        "PORTABILIDADE COM REFIN": "Portabilidade com Refinanciamento",
        "PORTABILIDADE COM REFINANCIAMENTO": "Portabilidade com Refinanciamento",
        "REFIN": "Refinanciamento",
        "REFIN DA PORT": "Refinanciamento",
        "REFINANCIAMENTO": "Refinanciamento",
        "NOVO": "Novo",
        "CARTAO": "Cartão",
        "SAQUE COMPLEMENTAR": "Saque Complementar",
    }
    return aliases.get(texto, original or "Outro")


def texto_planilha(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return limpar_texto(valor)

def registrar_historico(proposta_id: int, anterior: str | None, novo: str | None, observacao: str) -> None:
    db = get_db()
    db.execute(
        """
        INSERT INTO historico (proposta_id, data_hora, status_anterior, status_novo, observacao)
        VALUES (?, ?, ?, ?, ?)
        """,
        (proposta_id, agora_iso(), anterior, novo, observacao),
    )
    db.commit()


def registrar_notificacao_importante(
    *,
    proposta_id: int | None,
    proposta_nome: str | None,
    proposta_numero: str | None,
    tipo: str,
    titulo: str,
    mensagem: str,
    data_hora: str | None = None,
) -> None:
    get_db().execute(
        """
        INSERT INTO notificacoes_importantes
            (proposta_id, proposta_nome, proposta_numero, tipo, titulo, mensagem, data_hora)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (proposta_id, proposta_nome, proposta_numero, tipo, titulo, mensagem, data_hora or agora_iso()),
    )
    get_db().commit()


def tipo_notificacao_historico(status_anterior: str | None, status_novo: str | None, observacao: str | None) -> tuple[str, str, str]:
    anterior = limpar_texto(status_anterior)
    novo = limpar_texto(status_novo)
    texto_observacao = limpar_texto(observacao)
    mudou_status = bool(anterior and novo and anterior != novo)

    if texto_observacao.startswith("Proposta criada") or texto_observacao.startswith("Proposta importada"):
        detalhe = f"{texto_observacao}. Status inicial: {novo}." if novo else texto_observacao
        return "proposta_criada", "Proposta criada", detalhe or "Nova proposta adicionada ao CRM."
    if novo == "Pago" and mudou_status:
        return "paga", "Proposta marcada como paga", f"De {anterior} para {novo}."
    if novo in {"Perdido / Cancelado", "Perdido", "Cancelado"} and mudou_status:
        return "reprovada", "Proposta reprovada/perdida", f"De {anterior} para {novo}."
    if novo == "Aguardando Reapresentação" and mudou_status:
        return "reapresentacao", "Aguardando reapresentação", f"De {anterior} para {novo}."
    if mudou_status:
        return "etapa", "Etapa alterada", f"De {anterior} para {novo}."
    return "alteracao", "Alteração importante", texto_observacao or "Proposta atualizada."


def notificacoes_lidas_ate() -> str:
    row = get_db().execute("SELECT lido_ate FROM notificacoes_leitura WHERE id = 1").fetchone()
    return row["lido_ate"] if row else ""


def marcar_todas_notificacoes_lidas() -> None:
    db = get_db()
    agora = agora_iso()
    atualizado = db.execute("UPDATE notificacoes_leitura SET lido_ate = ? WHERE id = 1", (agora,))
    if atualizado.rowcount == 0:
        db.execute("INSERT INTO notificacoes_leitura (id, lido_ate) VALUES (1, ?)", (agora,))
    db.commit()


def carregar_notificacoes_importantes(limite: int = 8) -> list[dict[str, Any]]:
    filtros_observacao = " OR ".join("h.observacao LIKE ?" for _ in NOTIFICACOES_IMPORTANTES_OBSERVACOES)
    db = get_db()
    lido_ate = notificacoes_lidas_ate()
    historicos = db.execute(
        f"""
        SELECT
            h.id,
            h.proposta_id,
            h.data_hora,
            h.status_anterior,
            h.status_novo,
            h.observacao,
            p.nome AS proposta_nome,
            p.numero_proposta AS proposta_numero
        FROM historico h
        LEFT JOIN propostas p ON p.id = h.proposta_id
        WHERE (
            COALESCE(h.status_anterior, '') <> ''
            AND COALESCE(h.status_novo, '') <> ''
            AND h.status_anterior <> h.status_novo
        )
        OR ({filtros_observacao})
        ORDER BY h.data_hora DESC, h.id DESC
        LIMIT ?
        """,
        (*NOTIFICACOES_IMPORTANTES_OBSERVACOES, limite * 2),
    ).fetchall()
    notificacoes_salvas = db.execute(
        """
        SELECT
            id,
            proposta_id,
            proposta_nome,
            proposta_numero,
            tipo,
            titulo,
            mensagem,
            data_hora
        FROM notificacoes_importantes
        ORDER BY data_hora DESC, id DESC
        LIMIT ?
        """,
        (limite * 2,),
    ).fetchall()
    lembretes_agenda = db.execute(
        """
        SELECT t.id, t.titulo, t.descricao, t.data_tarefa, t.horario, t.notificado_em,
               p.nome AS proposta_nome, p.numero_proposta AS proposta_numero
        FROM tarefas t
        LEFT JOIN propostas p ON p.id = t.proposta_id
        WHERE t.notificar = 1
          AND t.status IN ('pendente', 'adiada')
          AND COALESCE(t.horario, '') <> ''
          AND datetime(t.data_tarefa || ' ' || t.horario) <= datetime(?)
        ORDER BY t.data_tarefa DESC, t.horario DESC, t.id DESC
        LIMIT ?
        """,
        (agora_iso(), limite),
    ).fetchall()

    notificacoes: list[dict[str, Any]] = []
    for h in historicos:
        tipo, titulo, mensagem = tipo_notificacao_historico(
            h["status_anterior"],
            h["status_novo"],
            h["observacao"],
        )
        notificacoes.append(
            {
                "id": f"historico-{h['id']}",
                "proposta_id": h["proposta_id"],
                "proposta_nome": h["proposta_nome"],
                "proposta_numero": h["proposta_numero"],
                "tipo": tipo,
                "titulo": titulo,
                "mensagem": mensagem,
                "data_hora": h["data_hora"],
                "lida": bool(lido_ate and h["data_hora"] <= lido_ate),
                "url": url_for("detalhe_proposta", proposta_id=h["proposta_id"], origem=url_origem_atual()),
            }
        )
    for n in notificacoes_salvas:
        url = url_for("detalhe_proposta", proposta_id=n["proposta_id"], origem=url_origem_atual()) if n["proposta_id"] else url_for("index")
        notificacoes.append(
            {
                "id": f"notificacao-{n['id']}",
                "proposta_id": n["proposta_id"],
                "proposta_nome": n["proposta_nome"],
                "proposta_numero": n["proposta_numero"],
                "tipo": n["tipo"],
                "titulo": n["titulo"],
                "mensagem": n["mensagem"],
                "data_hora": n["data_hora"],
                "lida": bool(lido_ate and n["data_hora"] <= lido_ate),
                "url": url,
            }
        )

    for tarefa in lembretes_agenda:
        data_hora = tarefa["notificado_em"] or agora_iso()
        detalhes = f"Horário: {tarefa['horario']}."
        if tarefa["descricao"]:
            detalhes += f" {tarefa['descricao']}"
        notificacoes.append(
            {
                "id": f"lembrete-{tarefa['id']}",
                "proposta_id": None,
                "proposta_nome": tarefa["proposta_nome"] or "Tarefa sem proposta vinculada",
                "proposta_numero": tarefa["proposta_numero"],
                "tipo": "lembrete",
                "titulo": "Lembrete de agenda",
                "mensagem": f"{tarefa['titulo']}. {detalhes}",
                "data_hora": data_hora,
                "lida": bool(tarefa["notificado_em"] and lido_ate and data_hora <= lido_ate),
                "url": url_for("editar_tarefa", tarefa_id=tarefa["id"], origem=url_origem_atual()),
            }
        )

    notificacoes.sort(key=lambda item: (item["data_hora"], str(item["id"])), reverse=True)
    return notificacoes[:limite]


def registrar_anotacao(proposta_id: int, texto: str, data_hora: str | None = None) -> None:
    texto = limpar_texto(texto)
    if not texto:
        return
    db = get_db()
    db.execute(
        """
        INSERT INTO anotacoes (proposta_id, data_hora, texto)
        VALUES (?, ?, ?)
        """,
        (proposta_id, data_hora or agora_iso(), texto),
    )
    db.execute(
        "UPDATE propostas SET observacoes = ?, data_atualizacao = ? WHERE id = ?",
        (texto, agora_iso(), proposta_id),
    )
    db.commit()


def carregar_status_etapas(grupo: str | None = None) -> list[sqlite3.Row]:
    # O parâmetro grupo é mantido apenas para compatibilidade com versões anteriores.
    # Na v4 todas as etapas aparecem no mesmo funil compartilhado.
    db = get_db()
    return db.execute("SELECT * FROM status_etapas WHERE ativo = 1 ORDER BY ordem ASC, id ASC").fetchall()


def nomes_status(grupo: str | None = None) -> list[str]:
    nomes = [row["nome"] for row in carregar_status_etapas(grupo)]
    return nomes or STATUS_LIST


def status_padrao() -> str:
    lista = nomes_status()
    return lista[0] if lista else "Novo lead"


def status_valido(status: str) -> bool:
    return status in nomes_status()

def status_entrada_proposta() -> str:
    status = "Aguardando inserção"
    db = get_db()
    existente = db.execute("SELECT id, ativo FROM status_etapas WHERE nome = ?", (status,)).fetchone()
    if existente:
        if not existente["ativo"]:
            db.execute("UPDATE status_etapas SET ativo = 1 WHERE id = ?", (existente["id"],))
            db.commit()
        return status

    proxima_ordem = db.execute("SELECT COALESCE(MAX(ordem), 0) + 1 AS ordem FROM status_etapas").fetchone()["ordem"]
    db.execute(
        "INSERT INTO status_etapas (nome, grupo, ordem, ativo) VALUES (?, 'geral', ?, 1)",
        (status, proxima_ordem),
    )
    db.commit()
    return status

def dados_formulario(proposta_atual: sqlite3.Row | dict[str, Any] | None = None) -> dict[str, Any]:
    status = limpar_texto(request.form.get("status")) or "Novo lead"
    if not status_valido(status):
        status = status_padrao()
    dados = {
        "nome": limpar_texto(request.form.get("nome")), "cpf": formatar_cpf(limpar_texto(request.form.get("cpf"))),
        "nascimento": limpar_texto(request.form.get("nascimento")),
        "nb_matricula": limpar_texto(request.form.get("nb_matricula")), "especie": limpar_texto(request.form.get("especie")), "numero_proposta": limpar_texto(request.form.get("numero_proposta")),
        "numero_port_vinculada": limpar_texto(request.form.get("numero_port_vinculada")), "numero_refin_vinculada": limpar_texto(request.form.get("numero_refin_vinculada")),
        "tipo_cliente": limpar_texto(request.form.get("tipo_cliente")), "banco_atual": limpar_texto(request.form.get("banco_atual")),
        "banco_destino": limpar_texto(request.form.get("banco_destino")), "banco_digitado": limpar_texto(request.form.get("banco_digitado")),
        "produto": limpar_texto(request.form.get("produto")), "promotora": limpar_texto(request.form.get("promotora")),
        "beneficio_bloqueado": normalizar_bloqueado(request.form.get("beneficio_bloqueado")), "valor_caiu_promotora": normalizar_sim_nao(request.form.get("valor_caiu_promotora")),
        "valor_sacado": normalizar_sim_nao(request.form.get("valor_sacado")), "parcela_atual": parse_moeda(request.form.get("parcela_atual")),
        "nova_parcela": parse_moeda(request.form.get("nova_parcela")), "troco": parse_moeda(request.form.get("troco")),
        "comissao_percentual": parse_percentual(request.form.get("comissao_percentual")), "comissao": parse_moeda(request.form.get("comissao")),
        "margem_apos": limpar_texto(request.form.get("margem_apos")), "status": status, "responsavel": "",
        "telefone": limpar_texto(request.form.get("telefone")), "endereco": limpar_texto(request.form.get("endereco")),
        "dados_bancarios": limpar_texto(request.form.get("dados_bancarios")), "proxima_acao": "",
        "data_retorno": limpar_texto(request.form.get("data_retorno")), "observacoes": limpar_texto(request.form.get("observacoes")),
    }
    if not produto_tem_campos_portabilidade(dados["produto"]):
        dados["banco_atual"] = ""
        dados["nova_parcela"] = 0
        dados["margem_apos"] = ""
    if not produto_tem_campos_vinculo(dados["produto"]):
        atual = dict(proposta_atual) if proposta_atual else {}
        refin_vinculado = proposta_eh_refin_vinculado(atual)
        if refin_vinculado:
            dados["numero_port_vinculada"] = limpar_texto(atual.get("numero_port_vinculada"))
            dados["numero_refin_vinculada"] = ""
        else:
            dados["numero_port_vinculada"] = ""
            dados["numero_refin_vinculada"] = ""
    return dados




def dados_nova_proposta() -> dict[str, Any]:
    return {
        "nome": limpar_texto(request.form.get("nome")),
        "cpf": formatar_cpf(limpar_texto(request.form.get("cpf"))),
        "nascimento": limpar_texto(request.form.get("nascimento")),
        "nb_matricula": limpar_texto(request.form.get("nb_matricula")),
        "especie": limpar_texto(request.form.get("especie")),
        "numero_proposta": "",
        "numero_port_vinculada": "",
        "numero_refin_vinculada": "",
        "tipo_cliente": "",
        "banco_atual": "",
        "banco_destino": "",
        "banco_digitado": "",
        "produto": limpar_texto(request.form.get("produto")),
        "promotora": "",
        "beneficio_bloqueado": "NÃO",
        "valor_caiu_promotora": "NÃO",
        "valor_sacado": "NÃO",
        "parcela_atual": 0,
        "nova_parcela": 0,
        "troco": 0,
        "comissao_percentual": 0,
        "comissao": 0,
        "margem_apos": "",
        "status": status_entrada_proposta(),
        "responsavel": "",
        "telefone": limpar_texto(request.form.get("telefone")),
        "endereco": limpar_texto(request.form.get("endereco")),
        "dados_bancarios": limpar_texto(request.form.get("dados_bancarios")),
        "proxima_acao": "",
        "data_retorno": "",
        "observacoes": limpar_texto(request.form.get("observacoes")),
    }


def proposta_vazia(status: str | None = None) -> dict[str, Any]:
    proposta = {campo: "" for campo in CAMPOS_PROPOSTA}
    proposta["status"] = status or status_entrada_proposta()
    proposta["beneficio_bloqueado"] = "NÃO"
    proposta["valor_caiu_promotora"] = "NÃO"
    proposta["valor_sacado"] = "NÃO"
    return proposta


def validar_nova_proposta(dados: dict[str, Any]) -> list[str]:
    obrigatorios = [
        ("nome", "Nome"),
        ("cpf", "CPF"),
        ("nb_matricula", "Benefício (NB/Matrícula)"),
        ("produto", "Produto"),
        ("telefone", "Telefone"),
    ]
    return [rotulo for campo, rotulo in obrigatorios if not limpar_texto(dados.get(campo))]

def chave_cliente(cpf: Any, nb_matricula: Any) -> tuple[str, str]:
    """Retorna a chave lógica do cliente.

    A chave é CPF + matrícula. Quando a matrícula está vazia, usamos string vazia
    para permitir reaproveitamento em cadastros antigos, mas o ideal é preencher a matrícula.
    """
    return formatar_cpf(limpar_texto(cpf)), limpar_texto(nb_matricula)


def salvar_cliente_dos_dados(dados: dict[str, Any] | sqlite3.Row) -> int | None:
    """Cria ou atualiza o cadastro do cliente a partir dos dados da proposta.

    O benefício bloqueado não é salvo no cadastro do cliente porque muda de uma
    proposta para outra. A chave usada é CPF + NB/Matrícula.
    """
    item = dict(dados)
    nome = limpar_texto(item.get("nome"))
    cpf, nb = chave_cliente(item.get("cpf"), item.get("nb_matricula"))
    if not nome or not cpf:
        return None

    agora = agora_iso()
    db = get_db()
    existente = db.execute(
        "SELECT id FROM clientes WHERE cpf = ? AND COALESCE(nb_matricula, '') = ? LIMIT 1",
        (cpf, nb),
    ).fetchone()
    if existente:
        db.execute(
            """
            UPDATE clientes SET
                nome = ?, telefone = ?, tipo_cliente = ?, endereco = ?, dados_bancarios = ?, data_atualizacao = ?
            WHERE id = ?
            """,
            (
                nome,
                limpar_texto(item.get("telefone")),
                limpar_texto(item.get("tipo_cliente")),
                limpar_texto(item.get("endereco")),
                limpar_texto(item.get("dados_bancarios")),
                agora,
                existente["id"],
            ),
        )
        return int(existente["id"])

    cursor = db.execute(
        """
        INSERT INTO clientes (
            nome, cpf, nb_matricula, telefone, tipo_cliente, endereco, dados_bancarios, data_criacao, data_atualizacao
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            nome, cpf, nb,
            limpar_texto(item.get("telefone")),
            limpar_texto(item.get("tipo_cliente")),
            limpar_texto(item.get("endereco")),
            limpar_texto(item.get("dados_bancarios")),
            agora, agora,
        ),
    )
    return int(cursor.lastrowid)


def migrar_clientes_a_partir_de_propostas() -> None:
    db = get_db()
    propostas = db.execute(
        """
        SELECT * FROM propostas
        WHERE COALESCE(TRIM(cpf), '') <> ''
          AND (cliente_id IS NULL OR cliente_id = '')
        ORDER BY id ASC
        """
    ).fetchall()
    alterou = False
    for proposta in propostas:
        cliente_id = salvar_cliente_dos_dados(proposta)
        if cliente_id:
            db.execute("UPDATE propostas SET cliente_id = ? WHERE id = ?", (cliente_id, proposta["id"]))
            alterou = True
    if alterou:
        db.commit()

def buscar_proposta(proposta_id: int) -> sqlite3.Row | None:
    return get_db().execute("SELECT * FROM propostas WHERE id = ?", (proposta_id,)).fetchone()


def normalizar_numero_proposta(valor: Any) -> str:
    return limpar_texto(valor).upper()


def buscar_propostas_vinculadas(proposta: sqlite3.Row) -> list[sqlite3.Row]:
    """Busca propostas vinculadas manualmente por número de proposta.

    A vinculação é flexível: basta preencher o número próprio em uma proposta e,
    na outra, preencher esse número no campo de portabilidade/refin vinculado.
    """
    dados = dict(proposta)
    numeros = {
        normalizar_numero_proposta(dados.get("numero_proposta")),
        normalizar_numero_proposta(dados.get("numero_port_vinculada")),
        normalizar_numero_proposta(dados.get("numero_refin_vinculada")),
    }
    numeros.discard("")
    if not numeros:
        return []

    cpf = limpar_texto(dados.get("cpf"))
    placeholders = ",".join("?" for _ in numeros)
    params: list[Any] = [proposta["id"], *sorted(numeros)]
    where_cpf = ""
    if cpf:
        where_cpf = " OR cpf = ?"
        params.append(cpf)

    # Busca por referência cruzada ou mesmo CPF. Depois filtra em Python para
    # reduzir falsos positivos em clientes com mais de uma proposta independente.
    rows = get_db().execute(
        f"""
        SELECT * FROM propostas
        WHERE id <> ?
          AND (
                UPPER(COALESCE(numero_proposta, '')) IN ({placeholders})
             OR UPPER(COALESCE(numero_port_vinculada, '')) IN ({placeholders})
             OR UPPER(COALESCE(numero_refin_vinculada, '')) IN ({placeholders})
             {where_cpf}
          )
        ORDER BY data_criacao ASC, id ASC
        """,
        params[:1] + sorted(numeros) + sorted(numeros) + sorted(numeros) + ([cpf] if cpf else []),
    ).fetchall()

    vinculadas = []
    for row in rows:
        row_nums = {
            normalizar_numero_proposta(row["numero_proposta"]),
            normalizar_numero_proposta(row["numero_port_vinculada"]),
            normalizar_numero_proposta(row["numero_refin_vinculada"]),
        }
        row_nums.discard("")
        if numeros.intersection(row_nums) or (cpf and row["cpf"] == cpf and (dados.get("numero_port_vinculada") or dados.get("numero_refin_vinculada") or row["numero_port_vinculada"] or row["numero_refin_vinculada"])):
            vinculadas.append(row)
    return vinculadas


def produto_eh_portabilidade_com_refin(proposta: sqlite3.Row | dict[str, Any]) -> bool:
    """Retorna True somente para a operação que exige um refin vinculado."""
    try:
        produto = proposta["produto"]
    except (KeyError, IndexError, TypeError):
        produto = ""
    return remover_acentos(limpar_texto(produto)).casefold() == "portabilidade com refinanciamento"


def proposta_eh_refin_vinculado(proposta: sqlite3.Row | dict[str, Any]) -> bool:
    """Identifica o refin criado a partir de uma Portabilidade com Refinanciamento."""
    try:
        dados = dict(proposta)
    except (TypeError, ValueError):
        dados = {}
    return (
        remover_acentos(limpar_texto(dados.get("produto"))).casefold() == "refinanciamento"
        and bool(limpar_texto(dados.get("numero_port_vinculada")))
    )


def incrementar_numero_proposta(valor: Any) -> str:
    """Gera o número do refin como número da port + 1.

    Mantém zeros à esquerda quando o número é totalmente numérico.
    Para formatos com prefixo, incrementa os dígitos finais.
    """
    texto = limpar_texto(valor)
    if not texto:
        return ""
    if texto.isdigit():
        return str(int(texto) + 1).zfill(len(texto))

    match = re.search(r"(\d+)$", texto)
    if not match:
        return ""
    inicio, fim = match.span(1)
    numero = match.group(1)
    incrementado = str(int(numero) + 1).zfill(len(numero))
    return f"{texto[:inicio]}{incrementado}{texto[fim:]}"


def pode_criar_refin_vinculado(proposta: sqlite3.Row | dict[str, Any]) -> bool:
    """Exibe o botão apenas em portabilidades com refinanciamento."""
    try:
        numero = proposta["numero_proposta"]
    except (KeyError, IndexError, TypeError):
        numero = ""
    return produto_eh_portabilidade_com_refin(proposta) and bool(incrementar_numero_proposta(numero))


def produto_tem_campos_portabilidade(produto: Any) -> bool:
    return remover_acentos(limpar_texto(produto)).casefold() in {
        "portabilidade",
        "portabilidade com refinanciamento",
    }


def produto_tem_campos_vinculo(produto: Any) -> bool:
    return remover_acentos(limpar_texto(produto)).casefold() == "portabilidade com refinanciamento"



def filtros_sql() -> tuple[str, list[Any], dict[str, str]]:
    filtros = {
        "nome": limpar_texto(request.args.get("nome")),
        "cpf": limpar_texto(request.args.get("cpf")),
        "mes": limpar_texto(request.args.get("mes")),
        "status": limpar_texto(request.args.get("status")),
        "banco_atual": limpar_texto(request.args.get("banco_atual")),
        "banco_digitado": limpar_texto(request.args.get("banco_digitado")),
        "tipo_cliente": limpar_texto(request.args.get("tipo_cliente")),
        "produto": limpar_texto(request.args.get("produto")),
        "promotora": limpar_texto(request.args.get("promotora")),
        "beneficio_bloqueado": limpar_texto(request.args.get("beneficio_bloqueado")),
        "valor_caiu_promotora": limpar_texto(request.args.get("valor_caiu_promotora")),
        "valor_sacado": limpar_texto(request.args.get("valor_sacado")),
    }

    where = []
    params: list[Any] = []
    if filtros["nome"]:
        where.append("nome LIKE ?")
        params.append(f"%{filtros['nome']}%")
    if filtros["cpf"]:
        where.append("cpf LIKE ?")
        params.append(f"%{filtros['cpf']}%")
    if filtros["mes"]:
        where.append("substr(data_criacao, 1, 7) = ?")
        params.append(filtros["mes"])
    if filtros["status"]:
        where.append("status = ?")
        params.append(filtros["status"])
    if filtros["banco_atual"]:
        where.append("banco_atual LIKE ?")
        params.append(f"%{filtros['banco_atual']}%")
    if filtros["banco_digitado"]:
        where.append("COALESCE(banco_digitado, '') LIKE ?")
        params.append(f"%{filtros['banco_digitado']}%")
    if filtros["tipo_cliente"]:
        where.append("tipo_cliente = ?")
        params.append(filtros["tipo_cliente"])
    if filtros["produto"]:
        where.append("produto = ?")
        params.append(filtros["produto"])
    if filtros["promotora"]:
        where.append("promotora LIKE ?")
        params.append(f"%{filtros['promotora']}%")
    if filtros["beneficio_bloqueado"]:
        where.append("beneficio_bloqueado = ?")
        params.append(filtros["beneficio_bloqueado"])
    if filtros["valor_caiu_promotora"]:
        where.append("COALESCE(valor_caiu_promotora, 'NÃO') = ?")
        params.append(filtros["valor_caiu_promotora"])
    if filtros["valor_sacado"]:
        where.append("COALESCE(valor_sacado, 'NÃO') = ?")
        params.append(filtros["valor_sacado"])

    sql = " WHERE " + " AND ".join(where) if where else ""
    return sql, params, filtros


def carregar_modelos() -> dict[str, str]:
    try:
        db = get_db()
        # Garante a tabela para casos em que a função seja chamada antes do before_request.
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS modelos_mensagens (
                nome TEXT PRIMARY KEY,
                texto TEXT NOT NULL,
                ordem INTEGER NOT NULL DEFAULT 0,
                data_atualizacao TEXT
            )
            """
        )
        sincronizar_modelos_banco(db)
        db.commit()
        rows = db.execute("SELECT nome, texto FROM modelos_mensagens ORDER BY ordem, nome").fetchall()
        return {row["nome"]: row["texto"] for row in rows}
    except RuntimeError:
        # Fallback para execução fora do contexto Flask.
        return carregar_modelos_arquivo()



def exportar_modelos_para_json() -> None:
    """Mantém o arquivo JSON como backup legível dos modelos salvos no banco."""
    try:
        rows = get_db().execute("SELECT nome, texto FROM modelos_mensagens ORDER BY ordem, nome").fetchall()
        dados = {row["nome"]: row["texto"] for row in rows}
        DATA_DIR.mkdir(exist_ok=True)
        MODELOS_PATH.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        # O banco é a fonte principal. Falha no backup JSON não deve impedir o CRM.
        pass

def proposta_para_dict(row: sqlite3.Row) -> dict[str, str]:
    dados = dict(row)
    dados["parcela_atual"] = br_moeda(dados.get("parcela_atual"))
    dados["nova_parcela"] = br_moeda(dados.get("nova_parcela"))
    dados["troco"] = br_moeda(dados.get("troco"))
    dados["valor"] = dados["troco"]
    dados["comissao_percentual"] = br_percentual(dados.get("comissao_percentual"))
    dados["comissao"] = br_moeda(dados.get("comissao"))
    return {k: "" if v is None else str(v) for k, v in dados.items()}


def montar_mensagens(proposta: sqlite3.Row) -> dict[str, str]:
    modelos = carregar_modelos()
    dados = proposta_para_dict(proposta)
    mensagens = {}
    for nome, modelo in modelos.items():
        try:
            mensagens[nome] = modelo.format(**dados)
        except KeyError:
            mensagens[nome] = modelo
    return mensagens


def hoje_iso() -> str:
    return date.today().isoformat()


def verificado_hoje(proposta: sqlite3.Row | dict[str, Any]) -> bool:
    try:
        valor = proposta["data_verificacao"]
    except (KeyError, IndexError, TypeError):
        valor = None
    return bool(valor and str(valor)[:10] == hoje_iso())


def status_verificacao_texto(proposta: sqlite3.Row | dict[str, Any]) -> str:
    return "Verificado hoje" if verificado_hoje(proposta) else "Não verificado hoje"


@app.context_processor
def helpers() -> dict[str, Any]:
    notificacoes = carregar_notificacoes_importantes()
    total_nao_lidas = sum(1 for notificacao in notificacoes if not notificacao["lida"])
    alerta_agenda = agenda_alerta_info()
    return {
        "STATUS_LIST": nomes_status(),
        "STATUS_VENDEDOR": nomes_status("vendedor"),
        "STATUS_ADMINISTRATIVO": nomes_status("administrativo"),
        "TIPOS_CLIENTE": TIPOS_CLIENTE,
        "PRODUTOS": PRODUTOS,
        "TAREFA_STATUS": TAREFA_STATUS,
        "TAREFA_PRIORIDADES": TAREFA_PRIORIDADES,
        "TAREFA_CATEGORIAS": TAREFA_CATEGORIAS,
        "br_moeda": br_moeda,
        "br_percentual": br_percentual,
        "br_data": br_data,
        "br_data_hora": br_data_hora,
        "mes_atual": mes_atual,
        "status_encerrado": status_encerrado,
        "banco_digitado_exibicao": banco_digitado_exibicao,
        "tamanho_arquivo": tamanho_arquivo,
        "hoje_iso": hoje_iso,
        "verificado_hoje": verificado_hoje,
        "status_verificacao_texto": status_verificacao_texto,
        "pode_criar_refin_vinculado": pode_criar_refin_vinculado,
        "produto_tem_campos_portabilidade": produto_tem_campos_portabilidade,
        "produto_tem_campos_vinculo": produto_tem_campos_vinculo,
        "proposta_eh_refin_vinculado": proposta_eh_refin_vinculado,
        "incrementar_numero_proposta": incrementar_numero_proposta,
        "url_origem_atual": url_origem_atual,
        "url_retorno_padrao": url_retorno_padrao,
        "origem_eh_hoje": origem_eh_hoje,
        "notificacoes_importantes": notificacoes,
        "notificacoes_importantes_total": total_nao_lidas,
        "agenda_alerta_ativo": alerta_agenda["ativo"],
        "agenda_alerta_total": alerta_agenda["total"],
        "agenda_antecedencia_minutos": alerta_agenda["antecedencia_minutos"],
        "anexos_base_dir": str(pasta_base_anexos()),
    }


@app.route("/notificacoes/marcar-lidas", methods=["POST"])
def marcar_notificacoes_lidas():
    marcar_todas_notificacoes_lidas()
    destino = url_interna_segura(request.form.get("next") or request.referrer, "/propostas")
    return redirect(destino)


@app.route("/api/notificacoes")
def api_notificacoes():
    notificacoes = carregar_notificacoes_importantes()
    total_nao_lidas = sum(1 for notificacao in notificacoes if not notificacao["lida"])
    origem = url_interna_segura(request.args.get("origem"), "/propostas")
    return jsonify({
        "total": total_nao_lidas,
        "html": render_template(
            "_notificacoes_panel.html",
            notificacoes_importantes=notificacoes,
            notificacoes_importantes_total=total_nao_lidas,
            notificacoes_origem=origem,
        ),
    })



def prazos_simulador_inss(prazos_json: str | None = None) -> dict[str, dict[str, Any]]:
    prazos = {codigo: dict(info) for codigo, info in INSS_NOVO_COEFICIENTES.items()}
    if not prazos_json:
        return prazos
    try:
        enviados = json.loads(prazos_json)
    except json.JSONDecodeError:
        return prazos
    if not isinstance(enviados, dict):
        return prazos

    for codigo, info in enviados.items():
        codigo_limpo = re.sub(r"[^a-zA-Z0-9_-]", "", str(codigo or ""))[:40]
        if not codigo_limpo or not isinstance(info, dict):
            continue
        label = limpar_texto(info.get("label"))[:80]
        if not label:
            continue
        try:
            coeficiente = float(info.get("coeficiente") or 0)
        except (TypeError, ValueError):
            continue
        if coeficiente <= 0 or coeficiente > 1:
            continue
        idade = limpar_texto(info.get("idade"))[:80]
        prazos[codigo_limpo] = {
            "label": label,
            "coeficiente": coeficiente,
            "idade": idade,
        }
    return prazos


def dados_simulador_inss() -> dict[str, Any]:
    tipo_operacao = limpar_texto(request.form.get("tipo_operacao")) or "novo_valor"
    if tipo_operacao not in {"novo_valor", "novo_margem"}:
        tipo_operacao = "novo_valor"
    prazo = limpar_texto(request.form.get("prazo")) or INSS_PRAZO_PADRAO
    faixa_cartao = limpar_texto(request.form.get("faixa_cartao")) or "ate_74"
    valor_base = parse_moeda(request.form.get("valor_base"))
    margem = parse_moeda(request.form.get("margem"))
    return {
        "nome": limpar_texto(request.form.get("nome")),
        "cpf": formatar_cpf(limpar_texto(request.form.get("cpf"))),
        "telefone": limpar_texto(request.form.get("telefone")),
        "nb_matricula": limpar_texto(request.form.get("nb_matricula")),
        "banco_digitado": limpar_texto(request.form.get("banco_digitado")),
        "promotora": limpar_texto(request.form.get("promotora")),
        "tipo_operacao": tipo_operacao,
        "prazo": prazo,
        "faixa_cartao": faixa_cartao,
        "prazos_json": request.form.get("prazos_json") or "",
        "valor_base": valor_base,
        "margem": margem,
        "observacoes": limpar_texto(request.form.get("observacoes")),
    }


def calcular_simulador_inss(dados: dict[str, Any]) -> dict[str, Any]:
    tipo = dados.get("tipo_operacao") or "novo_valor"
    prazo = dados.get("prazo") or INSS_PRAZO_PADRAO
    faixa_cartao = dados.get("faixa_cartao") or "ate_74"
    prazos = prazos_simulador_inss(dados.get("prazos_json"))
    coef_info = prazos.get(prazo) or prazos[INSS_PRAZO_PADRAO]
    coeficiente = float(coef_info["coeficiente"])
    valor = float(dados.get("valor_base") or 0)
    parcela = 0.0
    valor_estimado = 0.0
    produto = "Novo"
    descricao = "Empréstimo novo INSS"
    prazo_label = coef_info["label"]

    if tipo == "novo_margem":
        parcela = float(dados.get("margem") or 0)
        valor_estimado = parcela / coeficiente if coeficiente else 0.0
        descricao = "Novo INSS por margem"
    else:
        valor_estimado = valor
        parcela = valor_estimado * coeficiente
        descricao = "Novo INSS por valor"

    return {
        "produto": produto,
        "descricao": descricao,
        "prazo_label": prazo_label,
        "coeficiente": coeficiente,
        "valor_estimado": round(valor_estimado, 2),
        "parcela_estimativa": round(parcela, 2),
        "mensagem": montar_mensagem_simulador_inss(dados, produto, descricao, valor_estimado, parcela, prazo_label),
    }


def montar_mensagem_simulador_inss(dados: dict[str, Any], produto: str, descricao: str, valor: float, parcela: float, prazo_label: str) -> str:
    return (
        f"Simulação INSS - {descricao}\n\n"
        f"Valor estimado: {br_moeda(valor)}\n"
        f"Parcela estimada: {br_moeda(parcela)}\n"
        f"Prazo: {prazo_label}\n\n"
        "Valores sujeitos à análise e confirmação do banco."
    )




@app.route("/converter-contatos", methods=["GET", "POST"])
def converter_contatos():
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        if not arquivo or not arquivo.filename:
            flash("Selecione um arquivo CSV ou XLSX para converter.", "erro")
            return redirect(url_for("converter_contatos"))
        try:
            linhas, total_lidas = linhas_contatos_convertidas(arquivo)
        except Exception as exc:
            flash(f"Erro ao converter arquivo: {exc}", "erro")
            return redirect(url_for("converter_contatos"))

        if not linhas:
            flash("Nenhum contato válido encontrado. Verifique se TELEFONE1 possui números preenchidos.", "erro")
            return redirect(url_for("converter_contatos"))

        memoria = gerar_excel_contatos_convertidos(linhas)
        nome_saida = f"contatos_convertidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            memoria,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=nome_saida,
        )

    return render_template("converter_contatos.html")


@app.route("/simulador-inss", methods=["GET", "POST"])
def simulador_inss():
    dados = {
        "nome": "", "cpf": "", "telefone": "", "nb_matricula": "", "banco_digitado": "", "promotora": "",
        "tipo_operacao": "novo_valor", "prazo": INSS_PRAZO_PADRAO, "faixa_cartao": "ate_74", "valor_base": 0,
        "margem": 0, "observacoes": "", "prazos_json": "",
    }
    resultado = None
    if request.method == "POST":
        dados = dados_simulador_inss()
        resultado = calcular_simulador_inss(dados)
    prazos = prazos_simulador_inss(dados.get("prazos_json"))
    return render_template(
        "simulador_inss.html",
        dados=dados,
        resultado=resultado,
        prazos=prazos
    )


@app.route("/simulador-inss/criar-proposta", methods=["POST"])
def simulador_inss_criar_proposta():
    dados_sim = dados_simulador_inss()
    resultado = calcular_simulador_inss(dados_sim)
    if not dados_sim["nome"]:
        flash("Informe o nome do cliente antes de criar a proposta.", "erro")
        return redirect(url_for("simulador_inss"))

    agora = agora_iso()
    dados_prop = {
        "nome": dados_sim["nome"],
        "cpf": dados_sim["cpf"],
        "nb_matricula": dados_sim["nb_matricula"],
        "numero_proposta": "",
        "numero_port_vinculada": "",
        "numero_refin_vinculada": "",
        "tipo_cliente": "INSS",
        "banco_atual": "",
        "banco_destino": "",
        "banco_digitado": dados_sim["banco_digitado"],
        "produto": resultado["produto"],
        "promotora": dados_sim["promotora"],
        "beneficio_bloqueado": "NÃO",
        "valor_caiu_promotora": "NÃO",
        "valor_sacado": "NÃO",
        "parcela_atual": 0,
        "nova_parcela": resultado["parcela_estimativa"],
        "troco": resultado["valor_estimado"],
        "comissao_percentual": 0,
        "comissao": 0,
        "margem_apos": "",
        "status": "Em atendimento" if status_valido("Em atendimento") else status_padrao(),
        "responsavel": "",
        "telefone": dados_sim["telefone"],
        "endereco": "",
        "dados_bancarios": "",
        "proxima_acao": "",
        "data_retorno": "",
        "observacoes": dados_sim["observacoes"] or f"Proposta criada a partir do Simulador INSS: {resultado['descricao']} ({resultado['prazo_label']}).",
    }

    db = get_db()
    cursor = db.execute(
        """
        INSERT INTO propostas (
            cliente_id, nome, cpf, nb_matricula, numero_proposta, numero_port_vinculada, numero_refin_vinculada, tipo_cliente, banco_atual, banco_destino, banco_digitado, produto,
            promotora, beneficio_bloqueado, valor_caiu_promotora, valor_sacado, data_verificacao, parcela_atual, nova_parcela, troco, comissao_percentual, comissao, margem_apos, status, responsavel,
            telefone, endereco, dados_bancarios, data_criacao, data_atualizacao, proxima_acao, data_retorno, observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            salvar_cliente_dos_dados(dados_prop), dados_prop["nome"], dados_prop["cpf"], dados_prop["nb_matricula"], dados_prop["numero_proposta"],
            dados_prop["numero_port_vinculada"], dados_prop["numero_refin_vinculada"], dados_prop["tipo_cliente"], dados_prop["banco_atual"], dados_prop["banco_destino"], dados_prop["banco_digitado"], dados_prop["produto"],
            dados_prop["promotora"], dados_prop["beneficio_bloqueado"], dados_prop["valor_caiu_promotora"], dados_prop["valor_sacado"], hoje_iso(), dados_prop["parcela_atual"], dados_prop["nova_parcela"], dados_prop["troco"],
            dados_prop["comissao_percentual"], dados_prop["comissao"], dados_prop["margem_apos"], dados_prop["status"], dados_prop["responsavel"], dados_prop["telefone"], dados_prop["endereco"], dados_prop["dados_bancarios"],
            agora, agora, dados_prop["proxima_acao"], dados_prop["data_retorno"], dados_prop["observacoes"],
        ),
    )
    proposta_id = cursor.lastrowid
    db.commit()
    registrar_historico(proposta_id, None, dados_prop["status"], "Proposta criada a partir do Simulador INSS.")
    registrar_anotacao(proposta_id, dados_prop["observacoes"], agora)
    flash("Proposta criada a partir da simulação INSS.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))


@app.route("/")
@app.route("/propostas")
def index():
    sql, params, filtros = filtros_sql()
    propostas = get_db().execute(
        f"SELECT * FROM propostas {sql} ORDER BY data_atualizacao DESC, id DESC", params
    ).fetchall()
    return render_template("index.html", propostas=propostas, filtros=filtros)


@app.route("/nova", methods=["GET", "POST"])
def nova_proposta():
    if request.method == "POST":
        dados = dados_nova_proposta()
        pendentes = validar_nova_proposta(dados)
        if pendentes:
            flash("Preencha os campos obrigatórios: " + ", ".join(pendentes) + ".", "erro")
            return render_template("nova_proposta.html", proposta=dados)

        agora = agora_iso()
        db = get_db()
        cursor = db.execute(
            """
            INSERT INTO propostas (
                cliente_id, nome, cpf, nascimento, nb_matricula, especie, numero_proposta, numero_port_vinculada, numero_refin_vinculada, tipo_cliente, banco_atual, banco_destino, banco_digitado, produto,
                promotora, beneficio_bloqueado, valor_caiu_promotora, valor_sacado, data_verificacao, parcela_atual, nova_parcela, troco, comissao_percentual, comissao, margem_apos, status, responsavel,
                telefone, endereco, dados_bancarios, data_criacao, data_atualizacao, proxima_acao, data_retorno, observacoes
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?
            )
            """,
            (
                salvar_cliente_dos_dados(dados), dados["nome"], dados["cpf"], dados["nascimento"], dados["nb_matricula"], dados["especie"], dados["numero_proposta"],
                dados["numero_port_vinculada"], dados["numero_refin_vinculada"], dados["tipo_cliente"],
                dados["banco_atual"], dados["banco_destino"], dados["banco_digitado"], dados["produto"],
                dados["promotora"], dados["beneficio_bloqueado"], dados["valor_caiu_promotora"], dados["valor_sacado"], hoje_iso(), dados["parcela_atual"], dados["nova_parcela"], dados["troco"], dados["comissao_percentual"], dados["comissao"], dados["margem_apos"],
                dados["status"], dados["responsavel"], dados["telefone"], dados["endereco"], dados["dados_bancarios"], agora, agora,
                dados["proxima_acao"], dados["data_retorno"], dados["observacoes"],
            ),
        )
        db.commit()
        registrar_historico(cursor.lastrowid, None, dados["status"], "Proposta criada em Aguardando inserção")
        if dados.get("observacoes"):
            registrar_anotacao(cursor.lastrowid, dados["observacoes"], agora)

        proposta_criada = buscar_proposta(cursor.lastrowid)
        salvos = salvar_anexos_upload(
            cursor.lastrowid,
            proposta_criada or dados,
            request.files.getlist("arquivos"),
        )
        if salvos:
            registrar_historico(cursor.lastrowid, dados["status"], dados["status"], f"{salvos} anexo(s) enviado(s) na criação")
        flash("Proposta criada com sucesso.", "ok")
        return redirect(url_for("detalhe_proposta", proposta_id=cursor.lastrowid))

    return render_template("nova_proposta.html", proposta=proposta_vazia())


@app.route("/proposta/<int:proposta_id>/criar-refin-vinculado", methods=["POST"])
def criar_refin_vinculado(proposta_id: int):
    port = buscar_proposta(proposta_id)
    if not port:
        flash("Proposta de portabilidade não encontrada.", "erro")
        return redirect(url_for("index"))

    if not produto_eh_portabilidade_com_refin(port):
        flash("O refinanciamento vinculado só pode ser criado a partir de uma proposta de Portabilidade com Refinanciamento.", "erro")
        return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))

    numero_port = limpar_texto(port["numero_proposta"])
    numero_refin = incrementar_numero_proposta(numero_port)
    if not numero_refin:
        flash("Informe um número de proposta válido na portabilidade para gerar o refinanciamento +1.", "erro")
        return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))

    db = get_db()
    existente = db.execute(
        """
        SELECT * FROM propostas
        WHERE id <> ?
          AND UPPER(COALESCE(numero_proposta, '')) = UPPER(?)
        LIMIT 1
        """,
        (proposta_id, numero_refin),
    ).fetchone()

    agora = agora_iso()
    if existente:
        # Se o refin já existe, apenas garante o vínculo cruzado.
        db.execute(
            "UPDATE propostas SET numero_refin_vinculada = ?, data_atualizacao = ? WHERE id = ?",
            (numero_refin, agora, proposta_id),
        )
        db.execute(
            "UPDATE propostas SET numero_port_vinculada = ?, data_atualizacao = ? WHERE id = ? AND COALESCE(TRIM(numero_port_vinculada), '') = ''",
            (numero_port, agora, existente["id"]),
        )
        db.commit()
        registrar_historico(
            proposta_id,
            port["status"],
            port["status"],
            f"Refinanciamento vinculado existente localizado: nº {numero_refin}.",
        )
        flash("Refinanciamento existente vinculado à portabilidade.", "ok")
        return redirect(url_for("detalhe_proposta", proposta_id=existente["id"]))

    banco_refin = limpar_texto(port["banco_digitado"]) or banco_digitado_exibicao(port) or port["banco_atual"] or ""
    status_inicial = status_padrao()
    observacao = f"Refinanciamento criado a partir da portabilidade nº {numero_port}."

    cliente_id_refin = (
        port["cliente_id"]
        if "cliente_id" in port.keys() and port["cliente_id"]
        else salvar_cliente_dos_dados(port)
    )
    tipo_cliente_refin = limpar_texto(port["tipo_cliente"])
    if not tipo_cliente_refin and cliente_id_refin:
        cliente = db.execute("SELECT tipo_cliente FROM clientes WHERE id = ?", (cliente_id_refin,)).fetchone()
        if cliente:
            tipo_cliente_refin = limpar_texto(cliente["tipo_cliente"])
    if not tipo_cliente_refin and limpar_texto(port["especie"]):
        # Espécie numérica é um dado próprio do benefício previdenciário do INSS.
        tipo_cliente_refin = "INSS"

    cursor = db.execute(
        """
        INSERT INTO propostas (
            cliente_id, nome, cpf, nascimento, nb_matricula, especie, numero_proposta, numero_port_vinculada, numero_refin_vinculada, tipo_cliente, banco_atual, banco_destino, banco_digitado, produto,
            promotora, beneficio_bloqueado, valor_caiu_promotora, valor_sacado, data_verificacao, parcela_atual, nova_parcela, troco, comissao_percentual, comissao, margem_apos, status, responsavel,
            telefone, endereco, dados_bancarios, data_criacao, data_atualizacao, proxima_acao, data_retorno, observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            cliente_id_refin, port["nome"], port["cpf"], port["nascimento"], port["nb_matricula"], port["especie"], numero_refin, numero_port, "",
            tipo_cliente_refin, "", "", banco_refin, "Refinanciamento",
            port["promotora"], port["beneficio_bloqueado"] or "NÃO", "NÃO", "NÃO", hoje_iso(),
            0, 0, 0, 0, 0, "", status_inicial, port["responsavel"], port["telefone"],
            port["endereco"] if "endereco" in port.keys() else "",
            port["dados_bancarios"] if "dados_bancarios" in port.keys() else "",
            agora, agora, "", "", observacao,
        ),
    )
    refin_id = cursor.lastrowid
    db.execute(
        "UPDATE propostas SET numero_refin_vinculada = ?, data_atualizacao = ? WHERE id = ?",
        (numero_refin, agora, proposta_id),
    )
    db.commit()

    registrar_historico(refin_id, None, status_inicial, f"Refinanciamento vinculado criado a partir da portabilidade nº {numero_port}.")
    registrar_anotacao(refin_id, observacao, agora)
    registrar_historico(proposta_id, port["status"], port["status"], f"Refinanciamento vinculado criado: nº {numero_refin}.")

    flash(f"Refinanciamento vinculado criado com nº {numero_refin}.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=refin_id))


@app.route("/proposta/<int:proposta_id>")
def detalhe_proposta(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    if not proposta:
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))
    historico = get_db().execute(
        "SELECT * FROM historico WHERE proposta_id = ? ORDER BY data_hora ASC", (proposta_id,)
    ).fetchall()
    anotacoes = get_db().execute(
        "SELECT * FROM anotacoes WHERE proposta_id = ? ORDER BY data_hora DESC, id DESC", (proposta_id,)
    ).fetchall()
    anexos = get_db().execute(
        "SELECT * FROM anexos WHERE proposta_id = ? ORDER BY data_upload DESC, id DESC",
        (proposta_id,),
    ).fetchall()
    vinculadas = buscar_propostas_vinculadas(proposta)
    mensagens = montar_mensagens(proposta)
    return render_template(
        "detalhe_proposta.html",
        proposta=proposta,
        historico=historico,
        anotacoes=anotacoes,
        anexos=anexos,
        vinculadas=vinculadas,
        pasta_anexos=pasta_cliente(proposta),
        mensagens=mensagens,
        modelos_mensagens=carregar_modelos(),
    )


@app.route("/api/propostas/buscar")
def api_buscar_propostas():
    termo_original = limpar_texto(request.args.get("q"))
    termo = termo_original.lower()
    termo_digitos = re.sub(r"\D", "", termo_original)
    if len(termo_original) < 2:
        return jsonify([])

    like = f"%{termo_original}%"
    like_digitos = f"%{termo_digitos}%" if termo_digitos else like
    propostas = get_db().execute(
        """
        SELECT id, nome, cpf, telefone, status, produto, banco_digitado, banco_atual, banco_destino,
               numero_proposta, numero_port_vinculada, numero_refin_vinculada
        FROM propostas
        WHERE nome LIKE ?
           OR cpf LIKE ?
           OR telefone LIKE ?
           OR REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(cpf, ''), '.', ''), '-', ''), ' ', ''), '/', '') LIKE ?
           OR REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(telefone, ''), ' ', ''), '-', ''), '(', ''), ')', '') LIKE ?
           OR COALESCE(numero_proposta, '') LIKE ?
           OR COALESCE(numero_port_vinculada, '') LIKE ?
           OR COALESCE(numero_refin_vinculada, '') LIKE ?
        ORDER BY data_atualizacao DESC, id DESC
        LIMIT 10
        """,
        (like, like, like, like_digitos, like_digitos, like, like, like),
    ).fetchall()

    def match_info(p: sqlite3.Row) -> tuple[str, str]:
        campos = [
            ("Telefone", p["telefone"] or "", True),
            ("CPF", p["cpf"] or "", True),
            ("Nº proposta", p["numero_proposta"] or "", False),
            ("Nº port vinculada", p["numero_port_vinculada"] or "", False),
            ("Nº refin vinculado", p["numero_refin_vinculada"] or "", False),
            ("Nome", p["nome"] or "", False),
        ]
        for rotulo, valor, comparar_digitos in campos:
            valor_texto = str(valor)
            if comparar_digitos and termo_digitos:
                valor_digitos = re.sub(r"\D", "", valor_texto)
                if termo_digitos in valor_digitos:
                    return rotulo, valor_texto
            if termo and termo in valor_texto.lower():
                return rotulo, valor_texto
        return "Resultado", termo_original

    resultados = []
    for p in propostas:
        campo, valor = match_info(p)
        resultados.append({
            "id": p["id"],
            "nome": p["nome"] or "",
            "cpf": p["cpf"] or "",
            "telefone": p["telefone"] or "",
            "status": p["status"] or "",
            "produto": p["produto"] or "",
            "banco": banco_digitado_exibicao(p) or "",
            "match_campo": campo,
            "match_valor": valor,
            "url": url_for("detalhe_proposta", proposta_id=p["id"]),
        })
    return jsonify(resultados)



@app.route("/api/clientes/por-cpf")
def api_clientes_por_cpf():
    cpf = formatar_cpf(limpar_texto(request.args.get("cpf")))
    if len(re.sub(r"\D", "", cpf)) < 11:
        return jsonify([])
    clientes = get_db().execute(
        """
        SELECT id, nome, cpf, nb_matricula, telefone, tipo_cliente, endereco, dados_bancarios, data_atualizacao
        FROM clientes
        WHERE cpf = ?
        ORDER BY
            CASE WHEN COALESCE(NULLIF(nb_matricula, ''), '') = '' THEN 1 ELSE 0 END,
            nb_matricula ASC,
            data_atualizacao DESC,
            id DESC
        """,
        (cpf,),
    ).fetchall()

    # Evita sugestões duplicadas para o mesmo CPF + matrícula.
    # Também evita mostrar "Sem matrícula" quando já existe uma matrícula real para o CPF,
    # porque isso confundiria o cadastro de nova proposta.
    por_matricula = {}
    possui_matricula_real = any((c["nb_matricula"] or "").strip() for c in clientes)

    for c in clientes:
        nb = (c["nb_matricula"] or "").strip()
        if not nb and possui_matricula_real:
            continue
        chave = nb or "__sem_matricula__"
        if chave not in por_matricula:
            por_matricula[chave] = c

    resposta = []
    for idx, c in enumerate(por_matricula.values(), start=1):
        nb = (c["nb_matricula"] or "").strip()
        label = f"Matrícula {idx}: {nb}" if nb else "Sem matrícula cadastrada"
        resposta.append({
            "id": c["id"],
            "label": label,
            "nome": c["nome"] or "",
            "cpf": c["cpf"] or "",
            "nb_matricula": c["nb_matricula"] or "",
            "telefone": c["telefone"] or "",
            "tipo_cliente": c["tipo_cliente"] or "",
            "endereco": c["endereco"] or "",
            "dados_bancarios": c["dados_bancarios"] or "",
        })
    return jsonify(resposta)

@app.route("/proposta/<int:proposta_id>/anexos", methods=["POST"])
def enviar_anexos(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    if not proposta:
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))

    arquivos = request.files.getlist("arquivos")
    if not any(arquivo and arquivo.filename for arquivo in arquivos):
        flash("Selecione pelo menos um arquivo para anexar.", "erro")
        return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))

    salvos = salvar_anexos_upload(proposta_id, proposta, arquivos)
    if salvos:
        registrar_historico(proposta_id, proposta["status"], proposta["status"], f"{salvos} anexo(s) enviado(s)")
        flash(f"{salvos} arquivo(s) anexado(s) na pasta do cliente.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))


@app.route("/anexo/<int:anexo_id>/baixar")
def baixar_anexo(anexo_id: int):
    anexo = get_db().execute("SELECT * FROM anexos WHERE id = ?", (anexo_id,)).fetchone()
    if not anexo:
        flash("Anexo não encontrado.", "erro")
        return redirect(url_for("index"))
    caminho = Path(anexo["caminho"])
    if not caminho.exists():
        flash("O arquivo não foi encontrado na pasta do cliente.", "erro")
        return redirect(request.referrer or url_for("index"))
    return send_file(caminho, as_attachment=True, download_name=anexo["nome_original"])


@app.route("/anexo/<int:anexo_id>/excluir", methods=["POST"])
def excluir_anexo(anexo_id: int):
    db = get_db()
    anexo = db.execute("SELECT * FROM anexos WHERE id = ?", (anexo_id,)).fetchone()
    if not anexo:
        flash("Anexo não encontrado.", "erro")
        return redirect(url_for("index"))
    proposta_id = anexo["proposta_id"]
    caminho = Path(anexo["caminho"])
    try:
        if caminho.exists():
            caminho.unlink()
    except OSError as exc:
        flash(f"Registro removido, mas não foi possível apagar o arquivo: {exc}", "erro")
    db.execute("DELETE FROM anexos WHERE id = ?", (anexo_id,))
    db.commit()
    flash("Anexo removido.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))


@app.route("/proposta/<int:proposta_id>/anotacao", methods=["POST"])
def adicionar_anotacao(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    if not proposta:
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))
    texto = limpar_texto(request.form.get("texto"))
    if not texto:
        flash("Digite uma anotação antes de salvar.", "erro")
        return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))
    registrar_anotacao(proposta_id, texto)
    registrar_historico(proposta_id, proposta["status"], proposta["status"], "Nova anotação adicionada")
    flash("Anotação adicionada ao histórico da proposta.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))

@app.route("/proposta/<int:proposta_id>/verificacao", methods=["POST"])
def atualizar_verificacao(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    is_fetch = request.headers.get("X-Requested-With") == "fetch"
    if not proposta:
        if is_fetch:
            return jsonify({"success": False, "message": "Proposta não encontrada.", "proposta_id": proposta_id}), 404
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))
    origem = url_retorno_padrao()
    verificado = normalizar_sim_nao(request.form.get("verificado"))
    data_verificacao = hoje_iso() if verificado == "SIM" else None
    get_db().execute(
        "UPDATE propostas SET data_verificacao = ?, data_atualizacao = ? WHERE id = ?",
        (data_verificacao, agora_iso(), proposta_id),
    )
    get_db().commit()
    registrar_historico(
        proposta_id,
        proposta["status"],
        proposta["status"],
        "Proposta marcada como verificada hoje" if verificado == "SIM" else "Verificação diária removida",
    )
    if is_fetch:
        return jsonify({
            "success": True,
            "message": "Verificação diária atualizada.",
            "proposta_id": proposta_id,
            "verificado": verificado == "SIM",
            "status_texto": "Verificado hoje" if verificado == "SIM" else "Não verificado hoje",
        })
    flash("Verificação diária atualizada.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=proposta_id, origem=origem))


@app.route("/proposta/<int:proposta_id>/contatado-hoje", methods=["POST"])
def marcar_contatado_hoje(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    is_fetch = request.headers.get("X-Requested-With") == "fetch"
    if not proposta:
        if is_fetch:
            return jsonify({"success": False, "message": "Proposta não encontrada.", "proposta_id": proposta_id}), 404
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))

    registrar_historico(proposta_id, proposta["status"], proposta["status"], "Contato realizado hoje")
    if is_fetch:
        return jsonify({
            "success": True,
            "message": "Contato registrado.",
            "proposta_id": proposta_id,
            "ultima_interacao": "Hoje",
        })
    flash("Contato registrado.", "ok")
    return redirect(url_interna_segura(request.form.get("next") or request.referrer, "/hoje"))


@app.route("/tarefas/nova", methods=["GET", "POST"])
def nova_tarefa():
    origem = url_retorno_padrao()
    proposta = None
    proposta_id_txt = limpar_texto(request.values.get("proposta_id"))
    if proposta_id_txt.isdigit():
        proposta = buscar_proposta(int(proposta_id_txt))

    if request.method == "POST":
        dados, erros = dados_formulario_tarefa()
        proposta = buscar_proposta(dados["proposta_id"]) if dados.get("proposta_id") else None
        if erros:
            for erro in erros:
                flash(erro, "erro")
            return render_template(
                "tarefa_form.html",
                tarefa=dados,
                proposta_vinculada=proposta,
                modo="nova",
                next_url=origem,
                titulo="Nova tarefa",
                subtitulo="Crie uma tarefa manual para a agenda.",
            )

        agora = agora_iso()
        cursor = get_db().execute(
            """
            INSERT INTO tarefas (
                titulo, descricao, data_tarefa, horario, prioridade, status, categoria,
                proposta_id, criado_em, concluido_em, atualizado_em, notificar, notificado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                dados["titulo"], dados["descricao"], dados["data_tarefa"], dados["horario"],
                dados["prioridade"], dados["status"], dados["categoria"], dados["proposta_id"],
                agora, agora if dados["status"] == "concluida" else None, agora,
                dados["notificar"], None,
            ),
        )
        get_db().commit()
        flash("Tarefa criada com sucesso.", "ok")
        return redirect(origem)

    tarefa = tarefa_vazia(proposta)
    data_informada = parse_data_iso(request.args.get("data"))
    if data_informada:
        tarefa["data_tarefa"] = data_informada
    horario_informado = normalizar_horario(request.args.get("horario"))
    if horario_informado:
        tarefa["horario"] = horario_informado
    return render_template(
        "tarefa_form.html",
        tarefa=tarefa,
        proposta_vinculada=proposta,
        modo="nova",
        next_url=origem,
        titulo="Nova tarefa",
        subtitulo="Crie uma tarefa manual para a agenda.",
    )


@app.route("/tarefas/<int:tarefa_id>/editar", methods=["GET", "POST"])
def editar_tarefa(tarefa_id: int):
    tarefa = buscar_tarefa(tarefa_id)
    if not tarefa:
        flash("Tarefa não encontrada.", "erro")
        return redirect(url_for("hoje"))

    origem = url_retorno_padrao()
    proposta = buscar_proposta(tarefa["proposta_id"]) if tarefa["proposta_id"] else None
    if request.method == "POST":
        dados, erros = dados_formulario_tarefa()
        proposta = buscar_proposta(dados["proposta_id"]) if dados.get("proposta_id") else None
        if erros:
            for erro in erros:
                flash(erro, "erro")
            return render_template(
                "tarefa_form.html",
                tarefa={**dict(tarefa), **dados},
                proposta_vinculada=proposta,
                modo="editar",
                next_url=origem,
                titulo="Editar tarefa",
                subtitulo="Atualize a tarefa manual.",
            )

        concluido_em = tarefa["concluido_em"]
        if dados["status"] == "concluida" and not concluido_em:
            concluido_em = agora_iso()
        if dados["status"] != "concluida":
            concluido_em = None
        get_db().execute(
            """
            UPDATE tarefas
            SET titulo = ?, descricao = ?, data_tarefa = ?, horario = ?, prioridade = ?,
                status = ?, categoria = ?, proposta_id = ?, concluido_em = ?, atualizado_em = ?,
                notificar = ?, notificado_em = ?
            WHERE id = ?
            """,
            (
                dados["titulo"], dados["descricao"], dados["data_tarefa"], dados["horario"],
                dados["prioridade"], dados["status"], dados["categoria"], dados["proposta_id"],
                concluido_em, agora_iso(), dados["notificar"], None, tarefa_id,
            ),
        )
        get_db().commit()
        flash("Tarefa atualizada.", "ok")
        return redirect(origem)

    return render_template(
        "tarefa_form.html",
        tarefa=tarefa,
        proposta_vinculada=proposta,
        modo="editar",
        next_url=origem,
        titulo="Editar tarefa",
        subtitulo="Atualize a tarefa manual.",
    )


@app.route("/tarefas/<int:tarefa_id>/concluir", methods=["POST"])
def concluir_tarefa(tarefa_id: int):
    tarefa = buscar_tarefa(tarefa_id)
    if not tarefa:
        if request.headers.get("X-Requested-With") == "fetch":
            return jsonify({"success": False, "message": "Tarefa não encontrada.", "tarefa_id": tarefa_id}), 404
        flash("Tarefa não encontrada.", "erro")
        return redirect(url_for("hoje"))
    agora = agora_iso()
    get_db().execute(
        "UPDATE tarefas SET status = 'concluida', concluido_em = ?, atualizado_em = ? WHERE id = ?",
        (agora, agora, tarefa_id),
    )
    get_db().commit()
    return resposta_tarefa(tarefa_id, "Tarefa concluída.")


@app.route("/tarefas/<int:tarefa_id>/adiar", methods=["POST"])
def adiar_tarefa(tarefa_id: int):
    tarefa = buscar_tarefa(tarefa_id)
    if not tarefa:
        if request.headers.get("X-Requested-With") == "fetch":
            return jsonify({"success": False, "message": "Tarefa não encontrada.", "tarefa_id": tarefa_id}), 404
        flash("Tarefa não encontrada.", "erro")
        return redirect(url_for("hoje"))
    nova_data = parse_data_iso(request.form.get("nova_data"))
    if not nova_data:
        if request.headers.get("X-Requested-With") == "fetch":
            return jsonify({"success": False, "message": "Informe uma nova data.", "tarefa_id": tarefa_id}), 400
        flash("Informe uma nova data.", "erro")
        return redirect(request.form.get("next") or url_for("hoje"))
    get_db().execute(
        """
        UPDATE tarefas
        SET data_tarefa = ?, status = 'adiada', concluido_em = NULL, notificado_em = NULL, atualizado_em = ?
        WHERE id = ?
        """,
        (nova_data, agora_iso(), tarefa_id),
    )
    get_db().commit()
    return resposta_tarefa(tarefa_id, "Tarefa adiada.")


@app.route("/tarefas/<int:tarefa_id>/cancelar", methods=["POST"])
def cancelar_tarefa(tarefa_id: int):
    tarefa = buscar_tarefa(tarefa_id)
    if not tarefa:
        if request.headers.get("X-Requested-With") == "fetch":
            return jsonify({"success": False, "message": "Tarefa não encontrada.", "tarefa_id": tarefa_id}), 404
        flash("Tarefa não encontrada.", "erro")
        return redirect(url_for("hoje"))
    get_db().execute(
        "UPDATE tarefas SET status = 'cancelada', concluido_em = NULL, atualizado_em = ? WHERE id = ?",
        (agora_iso(), tarefa_id),
    )
    get_db().commit()
    return resposta_tarefa(tarefa_id, "Tarefa cancelada.")


@app.route("/tarefas/<int:tarefa_id>/excluir", methods=["POST"])
def excluir_tarefa(tarefa_id: int):
    tarefa = buscar_tarefa(tarefa_id)
    if not tarefa:
        if request.headers.get("X-Requested-With") == "fetch":
            return jsonify({"success": False, "message": "Tarefa não encontrada.", "tarefa_id": tarefa_id}), 404
        flash("Tarefa não encontrada.", "erro")
        return redirect(url_for("hoje"))
    get_db().execute("DELETE FROM tarefas WHERE id = ?", (tarefa_id,))
    get_db().commit()
    if request.headers.get("X-Requested-With") == "fetch":
        alerta_agenda = agenda_alerta_info()
        return jsonify({
            "success": True,
            "message": "Tarefa excluída.",
            "tarefa_id": tarefa_id,
            "status": "excluida",
            "agenda_alerta_ativo": alerta_agenda["ativo"],
            "agenda_alerta_total": alerta_agenda["total"],
        })
    flash("Tarefa excluída.", "ok")
    return redirect(url_interna_segura(request.form.get("next") or request.referrer, "/hoje"))


@app.route("/mensagens")
def modelos_mensagens_page():
    return render_template(
        "modelos_mensagens.html",
        modelos_mensagens=carregar_modelos(),
        titulo="Modelos de mensagens",
        subtitulo="Cadastre, edite e exclua modelos usados nas propostas.",
    )


@app.route("/mensagens/modelos", methods=["POST"])
def atualizar_modelos_mensagens():
    """Compatibilidade com a tela antiga: atualiza todos os modelos enviados no formulário."""
    db = get_db()
    modelos_atuais = carregar_modelos()

    for ordem, nome in enumerate(modelos_atuais.keys(), start=1):
        campo = f"modelo__{nome}"
        texto = request.form.get(campo)
        texto_final = texto if texto is not None else modelos_atuais[nome]
        db.execute(
            """
            INSERT OR REPLACE INTO modelos_mensagens (nome, texto, ordem, data_atualizacao)
            VALUES (?, ?, ?, ?)
            """,
            (nome, texto_final, ordem, agora_iso()),
        )

    db.commit()
    exportar_modelos_para_json()
    flash("Mensagens padrão atualizadas.", "ok")
    destino = request.form.get("next") or request.referrer or url_for("index")
    return redirect(destino)


@app.route("/mensagens/modelos/adicionar", methods=["POST"])
def adicionar_modelo_mensagem():
    db = get_db()
    nome = limpar_texto(request.form.get("nome_modelo"))
    texto = (request.form.get("texto_modelo") or "").strip()

    if not nome or not texto:
        flash("Informe o nome e o texto do modelo.", "erro")
        return redirect(request.form.get("next") or request.referrer or url_for("index"))

    existente = db.execute("SELECT nome FROM modelos_mensagens WHERE LOWER(nome) = LOWER(?)", (nome,)).fetchone()
    if existente:
        flash("Já existe um modelo com esse nome.", "erro")
        return redirect(request.form.get("next") or request.referrer or url_for("index"))

    maior_ordem = db.execute("SELECT COALESCE(MAX(ordem), 0) AS maior FROM modelos_mensagens").fetchone()["maior"]
    db.execute(
        """
        INSERT INTO modelos_mensagens (nome, texto, ordem, data_atualizacao)
        VALUES (?, ?, ?, ?)
        """,
        (nome, texto, maior_ordem + 1, agora_iso()),
    )
    db.commit()
    exportar_modelos_para_json()
    flash("Modelo de mensagem adicionado.", "ok")
    return redirect(request.form.get("next") or request.referrer or url_for("index"))


@app.route("/mensagens/modelos/editar", methods=["POST"])
def editar_modelo_mensagem():
    db = get_db()
    nome_original = request.form.get("nome_original") or ""
    nome_novo = limpar_texto(request.form.get("nome_modelo"))
    texto = (request.form.get("texto_modelo") or "").strip()

    if not nome_original or not nome_novo or not texto:
        flash("Informe o nome e o texto do modelo para editar.", "erro")
        return redirect(request.form.get("next") or request.referrer or url_for("index"))

    atual = db.execute("SELECT nome, ordem FROM modelos_mensagens WHERE nome = ?", (nome_original,)).fetchone()
    if not atual:
        flash("Modelo não encontrado.", "erro")
        return redirect(request.form.get("next") or request.referrer or url_for("index"))

    if nome_novo != nome_original:
        duplicado = db.execute("SELECT nome FROM modelos_mensagens WHERE LOWER(nome) = LOWER(?)", (nome_novo,)).fetchone()
        if duplicado:
            flash("Já existe outro modelo com esse nome.", "erro")
            return redirect(request.form.get("next") or request.referrer or url_for("index"))

    db.execute(
        """
        UPDATE modelos_mensagens
        SET nome = ?, texto = ?, data_atualizacao = ?
        WHERE nome = ?
        """,
        (nome_novo, texto, agora_iso(), nome_original),
    )
    db.commit()
    exportar_modelos_para_json()
    flash("Modelo de mensagem atualizado.", "ok")
    return redirect(request.form.get("next") or request.referrer or url_for("index"))


@app.route("/mensagens/modelos/excluir", methods=["POST"])
def excluir_modelo_mensagem():
    db = get_db()
    nome = request.form.get("nome_modelo") or ""
    total = db.execute("SELECT COUNT(*) AS total FROM modelos_mensagens").fetchone()["total"]
    if total <= 1:
        flash("Mantenha pelo menos um modelo de mensagem cadastrado.", "erro")
        return redirect(request.form.get("next") or request.referrer or url_for("index"))

    db.execute("DELETE FROM modelos_mensagens WHERE nome = ?", (nome,))
    db.commit()
    exportar_modelos_para_json()
    flash("Modelo de mensagem excluído.", "ok")
    return redirect(request.form.get("next") or request.referrer or url_for("index"))



@app.route("/gerador-mensagens")
def gerador_mensagens_page():
    return render_template(
        "gerador_mensagens.html",
        modelos_gerador=carregar_modelos_gerador(),
        titulo="Gerador de mensagens",
        subtitulo="Preencha os valores, selecione um modelo e copie a mensagem pronta.",
    )


@app.route("/gerador-mensagens/modelos/adicionar", methods=["POST"])
def adicionar_modelo_gerador():
    db = get_db()
    nome = limpar_texto(request.form.get("nome_modelo"))
    texto = (request.form.get("texto_modelo") or "").strip()
    if not nome or not texto:
        flash("Informe o nome e o texto do modelo.", "erro")
        return redirect(url_for("gerador_mensagens_page"))
    existente = db.execute(
        "SELECT nome FROM modelos_gerador_mensagens WHERE LOWER(nome) = LOWER(?)",
        (nome,),
    ).fetchone()
    if existente:
        flash("Já existe um modelo com esse nome.", "erro")
        return redirect(url_for("gerador_mensagens_page"))
    maior_ordem = db.execute("SELECT COALESCE(MAX(ordem), 0) AS maior FROM modelos_gerador_mensagens").fetchone()["maior"]
    db.execute(
        """
        INSERT INTO modelos_gerador_mensagens (nome, texto, ordem, data_atualizacao)
        VALUES (?, ?, ?, ?)
        """,
        (nome, texto, maior_ordem + 1, agora_iso()),
    )
    db.commit()
    flash("Modelo do gerador adicionado.", "ok")
    return redirect(url_for("gerador_mensagens_page"))


@app.route("/gerador-mensagens/modelos/editar", methods=["POST"])
def editar_modelo_gerador():
    db = get_db()
    nome_original = request.form.get("nome_original") or ""
    nome_novo = limpar_texto(request.form.get("nome_modelo"))
    texto = (request.form.get("texto_modelo") or "").strip()
    if not nome_original or not nome_novo or not texto:
        flash("Informe o nome e o texto do modelo para editar.", "erro")
        return redirect(url_for("gerador_mensagens_page"))
    atual = db.execute("SELECT nome, ordem FROM modelos_gerador_mensagens WHERE nome = ?", (nome_original,)).fetchone()
    if not atual:
        flash("Modelo não encontrado.", "erro")
        return redirect(url_for("gerador_mensagens_page"))
    if nome_novo.lower() != nome_original.lower():
        duplicado = db.execute(
            "SELECT nome FROM modelos_gerador_mensagens WHERE LOWER(nome) = LOWER(?)",
            (nome_novo,),
        ).fetchone()
        if duplicado:
            flash("Já existe outro modelo com esse nome.", "erro")
            return redirect(url_for("gerador_mensagens_page"))
    db.execute(
        """
        UPDATE modelos_gerador_mensagens
        SET nome = ?, texto = ?, data_atualizacao = ?
        WHERE nome = ?
        """,
        (nome_novo, texto, agora_iso(), nome_original),
    )
    db.commit()
    flash("Modelo do gerador atualizado.", "ok")
    return redirect(url_for("gerador_mensagens_page"))


@app.route("/gerador-mensagens/modelos/excluir", methods=["POST"])
def excluir_modelo_gerador():
    db = get_db()
    nome = request.form.get("nome_modelo") or ""
    total = db.execute("SELECT COUNT(*) AS total FROM modelos_gerador_mensagens").fetchone()["total"]
    if total <= 1:
        flash("Mantenha pelo menos um modelo no gerador.", "erro")
        return redirect(url_for("gerador_mensagens_page"))
    db.execute("DELETE FROM modelos_gerador_mensagens WHERE nome = ?", (nome,))
    db.commit()
    flash("Modelo do gerador excluído.", "ok")
    return redirect(url_for("gerador_mensagens_page"))

@app.route("/proposta/<int:proposta_id>/financeiro", methods=["POST"])
def atualizar_financeiro(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    if not proposta:
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))
    caiu = normalizar_sim_nao(request.form.get("valor_caiu_promotora"))
    sacado = normalizar_sim_nao(request.form.get("valor_sacado"))
    get_db().execute(
        "UPDATE propostas SET valor_caiu_promotora = ?, valor_sacado = ?, data_atualizacao = ? WHERE id = ?",
        (caiu, sacado, agora_iso(), proposta_id),
    )
    get_db().commit()
    registrar_historico(
        proposta_id,
        proposta["status"],
        proposta["status"],
        f"Controle financeiro atualizado: caiu na promotora {caiu}; sacado {sacado}",
    )
    flash("Controle financeiro atualizado.", "ok")
    return redirect(url_for("detalhe_proposta", proposta_id=proposta_id))


@app.route("/proposta/<int:proposta_id>/editar", methods=["GET", "POST"])
def editar_proposta(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    if not proposta:
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))

    if request.method == "POST":
        dados = dados_formulario(proposta)
        origem = url_retorno_padrao()
        if not dados["nome"]:
            flash("Informe o nome do cliente.", "erro")
            return render_template("editar_proposta.html", proposta={**dict(proposta), **dados})

        status_anterior = proposta["status"]
        db = get_db()
        data_encerramento = data_encerramento_para_status(
            status_anterior,
            dados["status"],
            proposta["data_encerramento"] if "data_encerramento" in proposta.keys() else None,
        )
        atualizado_em = agora_iso()
        db.execute(
            """
            UPDATE propostas SET
                cliente_id = ?, nome = ?, cpf = ?, nascimento = ?, nb_matricula = ?, especie = ?, numero_proposta = ?, numero_port_vinculada = ?, numero_refin_vinculada = ?, tipo_cliente = ?, banco_atual = ?,
                banco_destino = ?, banco_digitado = ?, produto = ?, promotora = ?, beneficio_bloqueado = ?, valor_caiu_promotora = ?, valor_sacado = ?, parcela_atual = ?, nova_parcela = ?, troco = ?,
                comissao_percentual = ?, comissao = ?, margem_apos = ?, status = ?, responsavel = ?, telefone = ?, endereco = ?, dados_bancarios = ?, data_atualizacao = ?, data_encerramento = ?,
                proxima_acao = ?, data_retorno = ?, observacoes = ?
            WHERE id = ?
            """,
            (
                salvar_cliente_dos_dados(dados), dados["nome"], dados["cpf"], dados["nascimento"], dados["nb_matricula"], dados["especie"], dados["numero_proposta"],
                dados["numero_port_vinculada"], dados["numero_refin_vinculada"], dados["tipo_cliente"],
                dados["banco_atual"], dados["banco_destino"], dados["banco_digitado"], dados["produto"],
                dados["promotora"], dados["beneficio_bloqueado"], dados["valor_caiu_promotora"], dados["valor_sacado"], dados["parcela_atual"], dados["nova_parcela"], dados["troco"], dados["comissao_percentual"], dados["comissao"], dados["margem_apos"],
                dados["status"], dados["responsavel"], dados["telefone"], dados["endereco"], dados["dados_bancarios"], atualizado_em, data_encerramento,
                dados["proxima_acao"], dados["data_retorno"], dados["observacoes"], proposta_id,
            ),
        )
        db.commit()
        if status_anterior != dados["status"]:
            registrar_historico(proposta_id, status_anterior, dados["status"], "Status alterado na edição")
        else:
            registrar_historico(proposta_id, status_anterior, dados["status"], "Dados da proposta atualizados")
        observacao_antiga = limpar_texto(proposta["observacoes"])
        if dados.get("observacoes") and dados["observacoes"] != observacao_antiga:
            registrar_anotacao(proposta_id, dados["observacoes"])
        acao = request.form.get("acao")
        if acao == "salvar_abrir_proxima" and urlsplit(origem).path == "/hoje":
            proxima = proxima_tarefa_hoje(origem, excluir_proposta_id=proposta_id)
            if proxima:
                flash("Proposta salva com sucesso.", "ok")
                return redirect(url_for("detalhe_proposta", proposta_id=proxima["id"], origem=origem))
            flash("Todas as tarefas da fila foram concluídas.", "ok")
            return redirect(url_interna_com_parametros(origem, "/hoje", destaque_proposta=proposta_id))
        flash("Proposta salva com sucesso.", "ok")
        if acao == "salvar_voltar":
            if urlsplit(origem).path in {"/funil", "/hoje"}:
                return redirect(url_interna_com_parametros(origem, "/propostas", destaque_proposta=proposta_id))
            return redirect(origem)
        return redirect(url_for("detalhe_proposta", proposta_id=proposta_id, origem=origem))

    return render_template("editar_proposta.html", proposta=proposta)




def mapear_status_encerradas(rotulo: str) -> dict[str, str] | None:
    mapa = {
        "Pago - falta cair na promotora": {"status": "Pago", "valor_caiu_promotora": "NÃO", "valor_sacado": "NÃO"},
        "Pago - disponível para saque": {"status": "Pago", "valor_caiu_promotora": "SIM", "valor_sacado": "NÃO"},
        "Pago - já sacado": {"status": "Pago", "valor_caiu_promotora": "SIM", "valor_sacado": "SIM"},
        "Perdido / Cancelado": {"status": "Perdido / Cancelado", "valor_caiu_promotora": "NÃO", "valor_sacado": "NÃO"},
    }
    return mapa.get(rotulo)


@app.route("/proposta/<int:proposta_id>/status", methods=["POST"])
def mudar_status(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    is_fetch = request.headers.get("X-Requested-With") == "fetch"
    if not proposta:
        if is_fetch:
            return resposta_status_json(False, "Proposta não encontrada.", proposta_id=proposta_id, status_http=404)
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))
    novo_status = limpar_texto(request.form.get("status"))
    observacao = limpar_texto(request.form.get("observacao")) or "Status atualizado"
    origem = limpar_texto(request.form.get("origem")) or "index"
    status_anterior = proposta["status"]

    dados_encerrada = mapear_status_encerradas(novo_status)
    if dados_encerrada:
        status_final = dados_encerrada["status"]
        caiu = dados_encerrada["valor_caiu_promotora"]
        sacado = dados_encerrada["valor_sacado"]
        mudou_algo = (
            status_final != proposta["status"]
            or caiu != (proposta["valor_caiu_promotora"] or "NÃO")
            or sacado != (proposta["valor_sacado"] or "NÃO")
        )
        if mudou_algo:
            data_encerramento = data_encerramento_para_status(
                proposta["status"],
                status_final,
                proposta["data_encerramento"] if "data_encerramento" in proposta.keys() else None,
            )
            get_db().execute(
                """
                UPDATE propostas
                SET status = ?, valor_caiu_promotora = ?, valor_sacado = ?, data_atualizacao = ?, data_encerramento = ?
                WHERE id = ?
                """,
                (status_final, caiu, sacado, agora_iso(), data_encerramento, proposta_id),
            )
            get_db().commit()
            registrar_historico(proposta_id, proposta["status"], status_final, observacao)
        if is_fetch:
            return resposta_status_json(
                True,
                "Proposta movida com sucesso",
                proposta_id=proposta_id,
                etapa_origem=status_anterior,
                etapa_destino=status_final,
            )
        flash("Situação da encerrada atualizada.", "ok")
    else:
        if not status_valido(novo_status):
            if is_fetch:
                return resposta_status_json(False, "Status inválido.", proposta_id=proposta_id, status_http=400)
            flash("Status inválido.", "erro")
            return redirect(request.referrer or url_for("index"))
        if novo_status != proposta["status"]:
            data_encerramento = data_encerramento_para_status(
                proposta["status"],
                novo_status,
                proposta["data_encerramento"] if "data_encerramento" in proposta.keys() else None,
            )
            get_db().execute(
                "UPDATE propostas SET status = ?, data_atualizacao = ?, data_encerramento = ? WHERE id = ?",
                (novo_status, agora_iso(), data_encerramento, proposta_id),
            )
            get_db().commit()
            registrar_historico(proposta_id, proposta["status"], novo_status, observacao)
            if is_fetch:
                return resposta_status_json(
                    True,
                    "Proposta movida com sucesso",
                    proposta_id=proposta_id,
                    etapa_origem=status_anterior,
                    etapa_destino=novo_status,
                )
            flash("Status atualizado.", "ok")
        elif is_fetch:
            return resposta_status_json(
                True,
                "Proposta já estava nesta etapa",
                proposta_id=proposta_id,
                etapa_origem=status_anterior,
                etapa_destino=novo_status,
            )
    if origem == "funil":
        return redirect(url_for("funil"))
    if origem == "encerradas":
        return redirect(url_for("encerradas"))
    if origem == "vendedor":
        return redirect(url_for("vendedor"))
    if origem == "administrativo":
        return redirect(url_for("administrativo"))
    return redirect(request.referrer or url_for("index"))




@app.route("/proposta/<int:proposta_id>/financeiro-rapido", methods=["POST"])
def atualizar_financeiro_rapido(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    is_fetch = request.headers.get("X-Requested-With") == "fetch"
    if not proposta:
        if is_fetch:
            return jsonify({"success": False, "message": "Proposta não encontrada."}), 404
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("encerradas"))

    valor = parse_moeda(request.form.get("troco"))
    comissao = parse_moeda(request.form.get("comissao"))
    percentual = round((comissao / valor) * 100, 4) if valor > 0 and comissao > 0 else 0.0

    get_db().execute(
        """
        UPDATE propostas
        SET troco = ?, comissao = ?, comissao_percentual = ?, data_atualizacao = ?
        WHERE id = ?
        """,
        (valor, comissao, percentual, agora_iso(), proposta_id),
    )
    get_db().commit()
    registrar_historico(
        proposta_id,
        proposta["status"],
        proposta["status"],
        f"Valor/comissão atualizados em Encerradas. Percentual calculado: {percentual:.2f}%",
    )
    if is_fetch:
        return jsonify({
            "success": True,
            "message": "Valores atualizados.",
            "proposta_id": proposta_id,
            "troco": br_moeda(valor),
            "comissao": br_moeda(comissao),
            "comissao_percentual": br_percentual(percentual),
            "comissao_numero": comissao,
        })
    flash("Valor e comissão atualizados.", "ok")
    return redirect(request.form.get("next") or url_for("encerradas"))


@app.route("/proposta/<int:proposta_id>/excluir", methods=["POST"])
def excluir_proposta(proposta_id: int):
    proposta = buscar_proposta(proposta_id)
    if not proposta:
        flash("Proposta não encontrada.", "erro")
        return redirect(url_for("index"))

    db = get_db()
    anexos = db.execute("SELECT * FROM anexos WHERE proposta_id = ?", (proposta_id,)).fetchall()
    registrar_notificacao_importante(
        proposta_id=None,
        proposta_nome=proposta["nome"],
        proposta_numero=proposta["numero_proposta"],
        tipo="lead_excluido",
        titulo="Lead excluído",
        mensagem=f"{proposta['nome']} foi removido do CRM.",
    )

    # Remove arquivos físicos dos anexos, quando existirem.
    # Se algum arquivo estiver aberto/bloqueado pelo Windows, o registro ainda será removido do CRM.
    arquivos_com_erro = 0
    for anexo in anexos:
        caminho = Path(anexo["caminho"])
        try:
            if caminho.exists():
                caminho.unlink()
        except OSError:
            arquivos_com_erro += 1

    # Remove registros relacionados mesmo quando o SQLite antigo não estiver com cascade ativo.
    db.execute("DELETE FROM anexos WHERE proposta_id = ?", (proposta_id,))
    db.execute("DELETE FROM anotacoes WHERE proposta_id = ?", (proposta_id,))
    db.execute("DELETE FROM historico WHERE proposta_id = ?", (proposta_id,))
    db.execute("DELETE FROM propostas WHERE id = ?", (proposta_id,))
    db.commit()

    # Tenta remover a pasta do cliente se ela ficar vazia.
    try:
        pasta = pasta_cliente(proposta)
        if pasta.exists() and not any(pasta.iterdir()):
            pasta.rmdir()
    except OSError:
        pass

    if arquivos_com_erro:
        flash(f"Proposta excluída. {arquivos_com_erro} arquivo(s) não puderam ser apagados da pasta.", "erro")
    else:
        flash("Lead/proposta excluído com sucesso.", "ok")

    destino = request.form.get("next") or url_for("index")
    return redirect(destino)


def renderizar_funil(status_visiveis: list[str], titulo: str, subtitulo: str, modulo: str):
    if not status_visiveis:
        status_visiveis = status_ativos()
    placeholders = ",".join("?" for _ in status_visiveis)
    propostas = get_db().execute(
        f"SELECT * FROM propostas WHERE status IN ({placeholders}) ORDER BY data_retorno ASC, id DESC",
        status_visiveis,
    ).fetchall()
    por_status = {status: [] for status in status_visiveis}
    for proposta in propostas:
        por_status.setdefault(proposta["status"], []).append(proposta)
    totais_status = {
        status: {
            "qtd": len(itens),
            "comissao": sum(float(item["comissao"] or 0) for item in itens),
        }
        for status, itens in por_status.items()
    }
    return render_template(
        "funil.html",
        por_status=por_status,
        totais_status=totais_status,
        status_opcoes=nomes_status(),
        titulo=titulo,
        subtitulo=subtitulo,
        modulo=modulo,
    )


@app.route("/funil")
def funil():
    return renderizar_funil(
        status_ativos(),
        "Funil em andamento",
        "Mostra apenas propostas abertas. Pagas e perdidas ficam no menu Encerradas.",
        "geral",
    )


def normalizar_status_tarefa(valor: Any) -> str:
    status = limpar_texto(valor).lower()
    return status if status in TAREFA_STATUS else "pendente"


def normalizar_prioridade_tarefa(valor: Any) -> str:
    prioridade = limpar_texto(valor).lower()
    return prioridade if prioridade in TAREFA_PRIORIDADES else "normal"


def normalizar_categoria_tarefa(valor: Any) -> str:
    categoria = limpar_texto(valor)
    return categoria if categoria in TAREFA_CATEGORIAS else "Outro"


def normalizar_horario(valor: Any) -> str:
    texto = limpar_texto(valor)
    if not texto:
        return ""
    try:
        return datetime.strptime(texto[:5], "%H:%M").strftime("%H:%M")
    except ValueError:
        return ""


def tarefa_vazia(proposta: sqlite3.Row | None = None) -> dict[str, Any]:
    return {
        "titulo": "",
        "descricao": "",
        "data_tarefa": hoje_iso(),
        "horario": "",
        "prioridade": "normal",
        "status": "pendente",
        "categoria": "Retorno",
        "notificar": 1 if agenda_notificar_padrao() else 0,
        "proposta_id": proposta["id"] if proposta else "",
    }


def buscar_tarefa(tarefa_id: int) -> sqlite3.Row | None:
    return get_db().execute(
        """
        SELECT t.*, p.nome AS proposta_nome, p.cpf AS proposta_cpf, p.telefone AS proposta_telefone,
               p.numero_proposta AS proposta_numero, p.status AS proposta_status
        FROM tarefas t
        LEFT JOIN propostas p ON p.id = t.proposta_id
        WHERE t.id = ?
        """,
        (tarefa_id,),
    ).fetchone()


def dados_formulario_tarefa() -> tuple[dict[str, Any], list[str]]:
    titulo = limpar_texto(request.form.get("titulo"))
    data_tarefa = parse_data_iso(request.form.get("data_tarefa"))
    prioridade = normalizar_prioridade_tarefa(request.form.get("prioridade"))
    status = normalizar_status_tarefa(request.form.get("status"))
    categoria = normalizar_categoria_tarefa(request.form.get("categoria"))
    proposta_id_txt = limpar_texto(request.form.get("proposta_id"))
    proposta_id = int(proposta_id_txt) if proposta_id_txt.isdigit() else None
    notificar = 1 if request.form.get("notificar") == "1" else 0
    erros = []
    if not titulo:
        erros.append("Informe o título da tarefa.")
    if not data_tarefa:
        erros.append("Informe uma data válida.")
    if notificar and not normalizar_horario(request.form.get("horario")):
        erros.append("Informe um horário para gerar a notificação.")
    if proposta_id and not buscar_proposta(proposta_id):
        erros.append("A proposta vinculada não foi encontrada.")
        proposta_id = None
    return {
        "titulo": titulo,
        "descricao": limpar_texto(request.form.get("descricao")),
        "data_tarefa": data_tarefa or hoje_iso(),
        "horario": normalizar_horario(request.form.get("horario")),
        "prioridade": prioridade,
        "status": status,
        "categoria": categoria,
        "proposta_id": proposta_id,
        "notificar": notificar,
    }, erros


def filtros_agenda_de_args(args: Any) -> dict[str, str]:
    return {
        "data": parse_data_iso(args.get("agenda_data")) if args.get("agenda_data") else "",
        "status": normalizar_status_tarefa(args.get("agenda_status")) if args.get("agenda_status") else "",
        "prioridade": normalizar_prioridade_tarefa(args.get("agenda_prioridade")) if args.get("agenda_prioridade") else "",
        "categoria": normalizar_categoria_tarefa(args.get("agenda_categoria")) if args.get("agenda_categoria") else "",
        "vinculo": limpar_texto(args.get("agenda_vinculo")),
    }


def prioridade_ordem_sql() -> str:
    return "CASE prioridade WHEN 'alta' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END"


def carregar_tarefas_agenda(filtros: dict[str, str]) -> list[sqlite3.Row]:
    where = []
    params: list[Any] = []
    if filtros["data"]:
        where.append("t.data_tarefa = ?")
        params.append(filtros["data"])
    if filtros["status"]:
        where.append("t.status = ?")
        params.append(filtros["status"])
    if filtros["prioridade"]:
        where.append("t.prioridade = ?")
        params.append(filtros["prioridade"])
    if filtros["categoria"]:
        where.append("t.categoria = ?")
        params.append(filtros["categoria"])
    if filtros["vinculo"] == "com_proposta":
        where.append("t.proposta_id IS NOT NULL")
    if filtros["vinculo"] == "sem_proposta":
        where.append("t.proposta_id IS NULL")
    sql = "WHERE " + " AND ".join(where) if where else ""
    return get_db().execute(
        f"""
        SELECT t.*, p.nome AS proposta_nome, p.cpf AS proposta_cpf, p.telefone AS proposta_telefone,
               p.numero_proposta AS proposta_numero, p.status AS proposta_status
        FROM tarefas t
        LEFT JOIN propostas p ON p.id = t.proposta_id
        {sql}
        ORDER BY {prioridade_ordem_sql()}, COALESCE(NULLIF(t.horario, ''), '99:99') ASC, t.data_tarefa ASC, t.id ASC
        """,
        params,
    ).fetchall()


def tarefas_vinculadas_pendentes(proposta_id: int) -> list[sqlite3.Row]:
    return get_db().execute(
        f"""
        SELECT *
        FROM tarefas
        WHERE proposta_id = ?
          AND status IN ('pendente', 'adiada')
        ORDER BY data_tarefa ASC, {prioridade_ordem_sql()}, COALESCE(NULLIF(horario, ''), '99:99') ASC, id ASC
        """,
        (proposta_id,),
    ).fetchall()


def contexto_agenda(filtros: dict[str, str]) -> dict[str, Any]:
    hoje_data = hoje_iso()
    tarefas = carregar_tarefas_agenda(filtros)
    ativas = [tarefa for tarefa in tarefas if tarefa["status"] in ("pendente", "adiada")]
    concluidas_hoje = [
        tarefa for tarefa in tarefas
        if tarefa["status"] == "concluida" and limpar_texto(tarefa["concluido_em"])[:10] == hoje_data
    ]
    secoes = [
        {
            "key": "agenda_atrasadas",
            "titulo": "Atrasadas",
            "vazia": "Nenhuma tarefa manual atrasada.",
            "tarefas": [tarefa for tarefa in ativas if tarefa["data_tarefa"] < hoje_data],
        },
        {
            "key": "agenda_hoje",
            "titulo": "Hoje",
            "vazia": "Nenhuma tarefa manual para hoje.",
            "tarefas": [tarefa for tarefa in ativas if tarefa["data_tarefa"] == hoje_data],
        },
        {
            "key": "agenda_proximas",
            "titulo": "Próximas",
            "vazia": "Nenhuma próxima tarefa manual.",
            "tarefas": [tarefa for tarefa in ativas if tarefa["data_tarefa"] > hoje_data],
        },
        {
            "key": "agenda_concluidas",
            "titulo": "Concluídas hoje",
            "vazia": "Nenhuma tarefa concluída hoje.",
            "tarefas": concluidas_hoje,
        },
    ]
    contadores = {
        "atrasadas": len(secoes[0]["tarefas"]),
        "hoje": len(secoes[1]["tarefas"]),
        "proximas": len(secoes[2]["tarefas"]),
        "concluidas_hoje": len(concluidas_hoje),
    }
    return {
        "agenda_secoes": secoes,
        "agenda_contadores": contadores,
        "agenda_filtros": filtros,
    }


def mes_agenda_de_args(args: Any) -> date:
    """Retorna o primeiro dia do mês solicitado, com fallback para o atual."""
    valor = limpar_texto(args.get("mes"))
    try:
        return datetime.strptime(valor, "%Y-%m").date().replace(day=1)
    except ValueError:
        return date.today().replace(day=1)


def carregar_retornos_agenda(inicio: date, fim: date) -> list[sqlite3.Row]:
    """Retorna uma previsão por operação, sem duplicar portabilidade + refin."""
    return get_db().execute(
        """
        SELECT p.id, p.nome, p.data_retorno, p.proxima_acao, p.status
        FROM propostas p
        WHERE p.data_retorno >= ? AND p.data_retorno < ?
          AND p.status NOT IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
          AND NOT (
              UPPER(TRIM(COALESCE(p.produto, ''))) = 'REFINANCIAMENTO'
              AND COALESCE(TRIM(p.numero_port_vinculada), '') <> ''
              AND EXISTS (
                  SELECT 1
                  FROM propostas port
                  WHERE port.id <> p.id
                    AND UPPER(TRIM(COALESCE(port.numero_proposta, ''))) = UPPER(TRIM(p.numero_port_vinculada))
                    AND COALESCE(TRIM(port.data_retorno), '') <> ''
                    AND port.status NOT IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
              )
          )
        ORDER BY p.data_retorno, p.nome, p.id
        """,
        (inicio.isoformat(), fim.isoformat()),
    ).fetchall()


def contexto_calendario(mes: date) -> dict[str, Any]:
    """Monta os eventos mensais sem criar retornos duplicados no banco."""
    if mes.month == 12:
        proximo_mes = mes.replace(year=mes.year + 1, month=1)
    else:
        proximo_mes = mes.replace(month=mes.month + 1)
    mes_anterior = (mes - timedelta(days=1)).replace(day=1)
    tarefas = get_db().execute(
        """SELECT t.*, p.nome AS proposta_nome FROM tarefas t
           LEFT JOIN propostas p ON p.id = t.proposta_id
           WHERE t.data_tarefa >= ? AND t.data_tarefa < ?
           ORDER BY COALESCE(NULLIF(t.horario, ''), '99:99'), t.id""",
        (mes.isoformat(), proximo_mes.isoformat()),
    ).fetchall()
    retornos = carregar_retornos_agenda(mes, proximo_mes)
    eventos_por_data: dict[str, list[dict[str, Any]]] = {}
    for tarefa in tarefas:
        eventos_por_data.setdefault(tarefa["data_tarefa"], []).append({
            "tipo": "tarefa", "id": tarefa["id"], "titulo": tarefa["titulo"],
            "horario": tarefa["horario"], "prioridade": tarefa["prioridade"],
            "status": tarefa["status"], "proposta_nome": tarefa["proposta_nome"],
        })
    for retorno in retornos:
        eventos_por_data.setdefault(retorno["data_retorno"][:10], []).append({
            "tipo": "retorno", "id": retorno["id"],
            "titulo": retorno["proxima_acao"] or "Retorno programado",
            "horario": "", "prioridade": "normal", "status": retorno["status"],
            "proposta_nome": retorno["nome"],
        })
    semanas = []
    for semana in calendar.Calendar(firstweekday=6).monthdatescalendar(mes.year, mes.month):
        semanas.append([{
            "data": dia.isoformat(), "numero": dia.day, "do_mes": dia.month == mes.month,
            "hoje": dia.isoformat() == hoje_iso(), "eventos": eventos_por_data.get(dia.isoformat(), []),
        } for dia in semana])
    meses_pt = ("janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro")
    return {
        "semanas_calendario": semanas, "mes_agenda": mes.strftime("%Y-%m"),
        "mes_anterior": mes_anterior.strftime("%Y-%m"), "proximo_mes": proximo_mes.strftime("%Y-%m"),
        "titulo_mes_agenda": f"{meses_pt[mes.month - 1].capitalize()} de {mes.year}",
        "total_tarefas_mes": len(tarefas), "total_retornos_mes": len(retornos),
    }


def data_agenda_de_args(args: Any) -> date:
    valor = limpar_texto(args.get("data"))
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError:
        return date.today()


def contexto_semana_agenda(referencia: date) -> dict[str, Any]:
    """Monta a visão semanal da Agenda, de domingo a sábado."""
    inicio = referencia - timedelta(days=(referencia.weekday() + 1) % 7)
    fim = inicio + timedelta(days=7)
    tarefas = get_db().execute(
        """SELECT t.*, p.nome AS proposta_nome FROM tarefas t
           LEFT JOIN propostas p ON p.id = t.proposta_id
           WHERE t.data_tarefa >= ? AND t.data_tarefa < ?
           ORDER BY t.data_tarefa, COALESCE(NULLIF(t.horario, ''), '99:99'), t.id""",
        (inicio.isoformat(), fim.isoformat()),
    ).fetchall()
    retornos = carregar_retornos_agenda(inicio, fim)
    eventos_horarios: dict[str, list[dict[str, Any]]] = {}
    eventos_dia_inteiro: dict[str, list[dict[str, Any]]] = {}
    for tarefa in tarefas:
        evento = {
            "tipo": "tarefa", "id": tarefa["id"], "titulo": tarefa["titulo"], "horario": tarefa["horario"],
            "prioridade": tarefa["prioridade"], "status": tarefa["status"], "proposta_nome": tarefa["proposta_nome"],
        }
        try:
            horario_tarefa = datetime.strptime(tarefa["horario"] or "", "%H:%M")
            evento["minuto"] = horario_tarefa.hour * 60 + horario_tarefa.minute
            evento["top"] = max(0, round((evento["minuto"] - 420) * 64 / 60))
            eventos_horarios.setdefault(tarefa["data_tarefa"], []).append(evento)
        except ValueError:
            eventos_dia_inteiro.setdefault(tarefa["data_tarefa"], []).append(evento)
    for retorno in retornos:
        eventos_dia_inteiro.setdefault(retorno["data_retorno"][:10], []).append({
            "tipo": "retorno", "id": retorno["id"], "titulo": retorno["proxima_acao"] or "Retorno programado",
            "horario": "", "prioridade": "normal", "status": retorno["status"], "proposta_nome": retorno["nome"],
        })
    nomes = ("Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb")
    dias = []
    for indice in range(7):
        dia = inicio + timedelta(days=indice)
        chave = dia.isoformat()
        dias.append({
            "data": chave, "nome": nomes[indice], "numero": dia.day, "hoje": dia == date.today(),
            "eventos_dia_inteiro": eventos_dia_inteiro.get(chave, []),
            "eventos_horarios": eventos_horarios.get(chave, []),
        })
    titulo = f"{inicio.strftime('%d/%m')} a {(fim - timedelta(days=1)).strftime('%d/%m/%Y')}"
    return {
        "dias_semana": dias, "semana_anterior": (inicio - timedelta(days=7)).isoformat(),
        "proxima_semana": (inicio + timedelta(days=7)).isoformat(), "data_semana": referencia.isoformat(),
        "titulo_semana_agenda": titulo, "total_tarefas_semana": len(tarefas), "total_retornos_semana": len(retornos),
        "horas_semana": list(range(7, 21)),
    }


def tarefa_json(tarefa_id: int, mensagem: str) -> dict[str, Any]:
    tarefa = buscar_tarefa(tarefa_id)
    alerta_agenda = agenda_alerta_info()
    return {
        "success": True,
        "message": mensagem,
        "tarefa_id": tarefa_id,
        "status": tarefa["status"] if tarefa else "",
        "data_tarefa": tarefa["data_tarefa"] if tarefa else "",
        "concluido_em": tarefa["concluido_em"] if tarefa else "",
        "agenda_alerta_ativo": alerta_agenda["ativo"],
        "agenda_alerta_total": alerta_agenda["total"],
    }


def resposta_tarefa(tarefa_id: int, mensagem: str, destino: str | None = None):
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify(tarefa_json(tarefa_id, mensagem))
    flash(mensagem, "ok")
    return redirect(url_interna_segura(destino or request.form.get("next") or request.referrer, "/hoje"))


@app.route("/api/agenda/lembretes")
def lembretes_agenda():
    """Entrega uma vez os lembretes já vencidos para o navegador ativo."""
    lembretes = get_db().execute(
        """SELECT t.id, t.titulo, t.descricao, t.data_tarefa, t.horario, p.nome AS proposta_nome
           FROM tarefas t LEFT JOIN propostas p ON p.id = t.proposta_id
           WHERE t.notificar = 1 AND t.notificado_em IS NULL
             AND t.status IN ('pendente', 'adiada') AND COALESCE(t.horario, '') <> ''
             AND datetime(t.data_tarefa || ' ' || t.horario) <= datetime(?)
           ORDER BY t.data_tarefa, t.horario, t.id""",
        (agora_iso(),),
    ).fetchall()
    alerta_agenda = agenda_alerta_info()
    return jsonify({
        "lembretes": [dict(item) for item in lembretes],
        "agenda_alerta_ativo": alerta_agenda["ativo"],
        "agenda_alerta_total": alerta_agenda["total"],
        "agenda_antecedencia_minutos": alerta_agenda["antecedencia_minutos"],
    })


@app.route("/api/agenda/lembretes/confirmar", methods=["POST"])
def confirmar_lembretes_agenda():
    dados = request.get_json(silent=True) or {}
    ids = [int(item) for item in dados.get("ids", []) if str(item).isdigit()]
    if ids:
        marcadores = ",".join("?" for _ in ids)
        get_db().execute(
            f"UPDATE tarefas SET notificado_em = ?, atualizado_em = ? WHERE id IN ({marcadores})",
            [agora_iso(), agora_iso(), *ids],
        )
        get_db().commit()
    return jsonify({"success": True})


def data_retorno_curta(proposta: sqlite3.Row) -> str:
    return limpar_texto(proposta["data_retorno"])[:10]


def data_operacional(valor: Any) -> date | None:
    texto = limpar_texto(valor)
    if not texto:
        return None
    for formato in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto[:19], formato).date()
        except ValueError:
            continue
    return None


def dias_desde(valor: Any) -> int | None:
    data = data_operacional(valor)
    if not data:
        return None
    return max(0, (date.today() - data).days)


def data_mais_recente(*valores: Any) -> str:
    datas = []
    for valor in valores:
        data = data_operacional(valor)
        if data:
            datas.append((data, limpar_texto(valor)))
    if not datas:
        return ""
    return max(datas, key=lambda item: item[0])[1]


def texto_dias_operacionais(dias: int | None) -> str:
    if dias is None:
        return "Sem registro"
    if dias == 0:
        return "Hoje"
    if dias == 1:
        return "Ontem"
    return f"{dias} dias"


def interacao_hoje(proposta: sqlite3.Row | dict[str, Any]) -> bool:
    return dias_desde(proposta.get("ultima_interacao_data") if isinstance(proposta, dict) else None) == 0


def carregar_propostas_abertas_hoje() -> list[sqlite3.Row]:
    status_visiveis = status_ativos()
    placeholders = ",".join("?" for _ in status_visiveis)
    return get_db().execute(
        f"""
        SELECT * FROM propostas
        WHERE status IN ({placeholders})
        ORDER BY
            CASE WHEN COALESCE(data_retorno, '') = '' THEN 1 ELSE 0 END,
            data_retorno ASC,
            data_atualizacao ASC,
            id ASC
        """,
        status_visiveis,
    ).fetchall()


def carregar_propostas_operacao() -> list[dict[str, Any]]:
    status_visiveis = status_ativos()
    placeholders = ",".join("?" for _ in status_visiveis)
    rows = get_db().execute(
        f"""
        SELECT p.*,
               (
                   SELECT MAX(h.data_hora)
                   FROM historico h
                   WHERE h.proposta_id = p.id
                     AND h.status_novo = p.status
               ) AS data_entrada_etapa,
               (
                   SELECT MAX(a.data_hora)
                   FROM anotacoes a
                   WHERE a.proposta_id = p.id
               ) AS ultima_anotacao,
               (
                   SELECT MAX(h.data_hora)
                   FROM historico h
                   WHERE h.proposta_id = p.id
               ) AS ultimo_historico
        FROM propostas p
        WHERE p.status IN ({placeholders})
        """,
        status_visiveis,
    ).fetchall()
    propostas = [preparar_proposta_operacao(row) for row in rows]
    return sorted(
        propostas,
        key=lambda p: (
            -int(p.get("dias_sem_interacao") or 0),
            -int(p.get("dias_etapa") or 0),
            limpar_texto(p.get("nome")),
            int(p.get("id") or 0),
        ),
    )


def preparar_proposta_operacao(row: sqlite3.Row) -> dict[str, Any]:
    proposta = dict(row)
    entrada_etapa = (
        proposta.get("data_entrada_etapa")
        or proposta.get("data_criacao")
        or proposta.get("data_atualizacao")
    )
    ultima_interacao = data_mais_recente(
        proposta.get("ultima_anotacao"),
        proposta.get("ultimo_historico"),
        proposta.get("data_verificacao"),
        proposta.get("data_atualizacao"),
        proposta.get("data_criacao"),
    )
    dias_etapa = dias_desde(entrada_etapa)
    dias_interacao = dias_desde(ultima_interacao)
    proposta["dias_etapa"] = dias_etapa if dias_etapa is not None else 0
    proposta["dias_sem_interacao"] = dias_interacao if dias_interacao is not None else 0
    proposta["entrada_etapa_data"] = entrada_etapa
    proposta["ultima_interacao_data"] = ultima_interacao
    proposta["ultima_interacao_texto"] = texto_dias_operacionais(dias_interacao)
    proposta["dias_etapa_texto"] = texto_dias_operacionais(dias_etapa)
    proposta["banco_operacao"] = banco_digitado_exibicao(proposta) or limpar_texto(proposta.get("banco_atual")) or "-"
    return proposta


def filtros_hoje_de_args(args: Any) -> dict[str, str]:
    return {
        "tipo_tarefa": limpar_texto(args.get("tipo_tarefa")),
        "etapa": limpar_texto(args.get("etapa")),
        "banco": limpar_texto(args.get("banco")),
        "verificacao": limpar_texto(args.get("verificacao")),
    }


def filtros_hoje_de_url(origem: Any) -> dict[str, str]:
    partes = urlsplit(url_interna_segura(origem, "/hoje"))
    return filtros_hoje_de_args(dict(parse_qsl(partes.query, keep_blank_values=True)))


def aplicar_filtros_hoje(propostas: list[sqlite3.Row], filtros: dict[str, str]) -> list[sqlite3.Row]:
    filtradas = []
    for proposta in propostas:
        banco = proposta.get("banco_operacao") if isinstance(proposta, dict) else banco_digitado_exibicao(proposta) or limpar_texto(proposta["banco_atual"])
        contato_hoje = interacao_hoje(proposta) if isinstance(proposta, dict) else verificado_hoje(proposta)
        if filtros["etapa"] and proposta["status"] != filtros["etapa"]:
            continue
        if filtros["banco"] and banco != filtros["banco"]:
            continue
        if filtros["verificacao"] == "pendente" and contato_hoje:
            continue
        if filtros["verificacao"] == "verificada" and not contato_hoje:
            continue
        filtradas.append(proposta)
    return filtradas


def classificar_bloco_operacao(proposta: dict[str, Any]) -> str:
    status = limpar_texto(proposta.get("status"))
    if int(proposta.get("dias_sem_interacao") or 0) >= DIAS_PARADA_OPERACIONAL:
        return "paradas"
    if status == "Aguardando CIP":
        return "cip"
    if status == "Aguardando Averbação":
        return "averbacao"
    if status == "Aguardando Pagamento":
        return "pagamento"
    if status == "Aguardando Reapresentação":
        return "reapresentacao"
    if normalizar_sim_nao(proposta.get("beneficio_bloqueado")) == "SIM":
        return "bloqueado"
    return "acompanhamento"


def indicadores_bloco(propostas: list[dict[str, Any]]) -> dict[str, Any]:
    tempos = [int(proposta.get("dias_etapa") or 0) for proposta in propostas]
    media = round(sum(tempos) / len(tempos), 1) if tempos else 0
    return {
        "quantidade": len(propostas),
        "tempo_medio": media,
        "maior_tempo": max(tempos, default=0),
    }


def secoes_hoje(propostas: list[dict[str, Any]], filtros: dict[str, str], motivos_atencao: dict[int, list[str]] | None = None) -> list[dict[str, Any]]:
    agrupadas = {key: [] for key in TAREFAS_HOJE_PRIORIDADE}
    for proposta in propostas:
        agrupadas[classificar_bloco_operacao(proposta)].append(proposta)

    secoes = []
    for key in TAREFAS_HOJE_PRIORIDADE:
        info = TAREFAS_HOJE_INFO[key]
        propostas_bloco = agrupadas[key]
        if not propostas_bloco and filtros["tipo_tarefa"] != key:
            continue
        secoes.append({
            "key": key,
            "titulo": info["titulo"],
            "vazia": info["vazia"],
            "propostas": propostas_bloco,
            "indicadores": indicadores_bloco(propostas_bloco),
        })
    if filtros["tipo_tarefa"]:
        return [secao for secao in secoes if secao["key"] == filtros["tipo_tarefa"]]
    return secoes


def fila_hoje(
    propostas: list[sqlite3.Row],
    filtros: dict[str, str],
    motivos_atencao: dict[int, list[str]],
    excluir_proposta_id: int | None = None,
) -> list[sqlite3.Row]:
    prioridades = [filtros["tipo_tarefa"]] if filtros["tipo_tarefa"] in TAREFAS_HOJE_PRIORIDADE else list(TAREFAS_HOJE_PRIORIDADE)
    fila = []
    vistos: set[int] = set()
    for key in prioridades:
        for proposta in propostas:
            proposta_id = int(proposta["id"])
            if proposta_id == excluir_proposta_id or proposta_id in vistos:
                continue
            if classificar_bloco_operacao(proposta) != key:
                continue
            vistos.add(proposta_id)
            fila.append(proposta)
    return fila


def contexto_hoje(filtros: dict[str, str], excluir_proposta_id: int | None = None) -> dict[str, Any]:
    propostas_base = carregar_propostas_operacao()
    bancos = sorted({
        proposta.get("banco_operacao") or "-"
        for proposta in propostas_base
        if proposta.get("banco_operacao") and proposta.get("banco_operacao") != "-"
    })
    propostas = aplicar_filtros_hoje(propostas_base, filtros)
    secoes_todas = secoes_hoje(propostas, {"tipo_tarefa": "", "etapa": "", "banco": "", "verificacao": ""}, {})
    secoes = secoes_hoje(propostas, filtros, {})
    fila = fila_hoje(propostas, filtros, {}, excluir_proposta_id)
    primeira_proposta = fila[0] if fila else None
    interacoes_hoje = sum(1 for proposta in fila if interacao_hoje(proposta))

    contadores = {
        "paradas": len(next((secao["propostas"] for secao in secoes_todas if secao["key"] == "paradas"), [])),
        "cip": len(next((secao["propostas"] for secao in secoes_todas if secao["key"] == "cip"), [])),
        "averbacao": len(next((secao["propostas"] for secao in secoes_todas if secao["key"] == "averbacao"), [])),
        "pagamento": len(next((secao["propostas"] for secao in secoes_todas if secao["key"] == "pagamento"), [])),
        "total": len(fila),
        "interacoes_hoje": interacoes_hoje,
        "pendentes": max(0, len(fila) - interacoes_hoje),
    }
    return {
        "secoes": secoes,
        "secoes_todas": secoes_todas,
        "contadores": contadores,
        "filtros": filtros,
        "bancos": bancos,
        "status_opcoes": status_ativos(),
        "motivos_atencao": {},
        "primeira_proposta": primeira_proposta,
        "fila": fila,
    }


def proxima_tarefa_hoje(origem: Any, excluir_proposta_id: int | None = None) -> sqlite3.Row | None:
    filtros = filtros_hoje_de_url(origem)
    contexto = contexto_hoje(filtros, excluir_proposta_id)
    return contexto["primeira_proposta"]


@app.route("/hoje")
def hoje():
    contexto = contexto_hoje(filtros_hoje_de_args(request.args))

    return render_template(
        "hoje.html",
        **contexto,
        titulo="Hoje",
        subtitulo="Propostas que precisam de atenção agora.",
    )


@app.route("/agenda")
def agenda():
    visao = "semana" if limpar_texto(request.args.get("visao")) == "semana" else "mes"
    if visao == "semana":
        contexto = contexto_semana_agenda(data_agenda_de_args(request.args))
    else:
        contexto = contexto_calendario(mes_agenda_de_args(request.args))
    return render_template(
        "agenda.html",
        **contexto,
        visao_agenda=visao,
        titulo="Agenda",
        subtitulo="Compromissos, tarefas e retornos das propostas em um só calendário.",
    )


@app.route("/encerradas")
def encerradas():
    mes = limpar_texto(request.args.get("mes")) or mes_atual()
    propostas = get_db().execute(
        """
        SELECT * FROM propostas
        WHERE status IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
          AND substr(COALESCE(data_encerramento, data_atualizacao, data_criacao), 1, 7) = ?
        ORDER BY COALESCE(data_encerramento, data_atualizacao) DESC, id DESC
        """,
        (mes,),
    ).fetchall()
    grupos = {
        "Pago - falta cair na promotora": [],
        "Pago - disponível para saque": [],
        "Pago - já sacado": [],
        "Perdido / Cancelado": [],
    }
    for proposta in propostas:
        if proposta["status"] == "Pago":
            caiu = (proposta["valor_caiu_promotora"] or "NÃO")
            sacado = (proposta["valor_sacado"] or "NÃO")
            if sacado == "SIM":
                grupos["Pago - já sacado"].append(proposta)
            elif caiu == "SIM":
                grupos["Pago - disponível para saque"].append(proposta)
            else:
                grupos["Pago - falta cair na promotora"].append(proposta)
        else:
            grupos["Perdido / Cancelado"].append(proposta)
    totais_status = {
        status: {
            "qtd": len(itens),
            "comissao": sum(float(item["comissao"] or 0) for item in itens),
        }
        for status, itens in grupos.items()
    }
    return render_template(
        "encerradas.html",
        por_status=grupos,
        totais_status=totais_status,
        status_opcoes=nomes_status(),
        titulo="Propostas encerradas",
        subtitulo="Pagas ficam separadas por comissão: falta cair, disponível para saque e já sacada.",
        modulo="encerradas",
        mes=mes,
    )


@app.route("/vendedor")
def vendedor():
    # Mantido para links antigos/favoritos. Agora há um único funil compartilhado.
    return redirect(url_for("funil"))


@app.route("/administrativo")
def administrativo():
    # Mantido para links antigos/favoritos. Agora há um único funil compartilhado.
    return redirect(url_for("funil"))


@app.route("/configuracoes", methods=["GET", "POST"])
def configuracoes():
    if request.method == "POST":
        acao = limpar_texto(request.form.get("acao")) or "agenda"
        if acao == "documentos":
            caminho, erro = validar_pasta_base_anexos(request.form.get("anexos_base_dir"))
            if erro:
                flash(erro, "erro")
                return redirect(url_for("configuracoes"))
            salvar_configuracao("anexos_base_dir", str(caminho))
            get_db().commit()
            flash("Pasta base dos documentos atualizada. Novos anexos serão salvos nesse local.", "ok")
            return redirect(url_for("configuracoes"))
        if acao == "restaurar_documentos":
            get_db().execute("DELETE FROM configuracoes WHERE chave = 'anexos_base_dir'")
            get_db().commit()
            flash("Pasta base restaurada para o padrão da instalação.", "ok")
            return redirect(url_for("configuracoes"))

        try:
            antecedencia = int(request.form.get("agenda_antecedencia_minutos") or 0)
        except (TypeError, ValueError):
            antecedencia = 0
        if antecedencia not in AGENDA_ANTECEDENCIAS_MINUTOS:
            flash("Escolha uma antecedência válida para os avisos da agenda.", "erro")
            return redirect(url_for("configuracoes"))

        salvar_configuracao("agenda_antecedencia_minutos", antecedencia)
        salvar_configuracao("agenda_notificar_padrao", 1 if request.form.get("agenda_notificar_padrao") == "1" else 0)
        get_db().commit()
        flash("Configurações salvas.", "ok")
        return redirect(url_for("configuracoes"))

    return render_template(
        "configuracoes.html",
        antecedencias=AGENDA_ANTECEDENCIAS_MINUTOS,
        antecedencia_atual=agenda_antecedencia_minutos(),
        notificar_padrao=agenda_notificar_padrao(),
        pasta_anexos_atual=str(pasta_base_anexos()),
        pasta_anexos_padrao=str(ANEXOS_BASE_DIR_PADRAO),
        pasta_anexos_personalizada=bool(obter_configuracao("anexos_base_dir", "")),
        titulo="Configurações",
        subtitulo="Personalize os alertas, o funil e a aparência do CRM.",
    )


def normalizar_formula_dashboard(configuracao: dict[str, Any], tipo_resultado: str = "numero") -> dict[str, Any]:
    primeiro = limpar_texto(configuracao.get("primeiro") or configuracao.get("indicador_a"))
    operacoes_brutas = configuracao.get("operacoes")
    if not isinstance(operacoes_brutas, list):
        operador_antigo = limpar_texto(configuracao.get("operador"))
        if tipo_resultado == "percentual" and operador_antigo == "dividir":
            operador_antigo = "percentual_de"
        operacoes_brutas = [{
            "operador": operador_antigo,
            "operando_tipo": "indicador",
            "indicador": limpar_texto(configuracao.get("indicador_b")),
        }] if operador_antigo else []

    operacoes: list[dict[str, Any]] = []
    for operacao in operacoes_brutas[:10]:
        if not isinstance(operacao, dict):
            continue
        operando_tipo = limpar_texto(operacao.get("operando_tipo")) or "indicador"
        try:
            valor = float(operacao.get("valor") or 0)
        except (TypeError, ValueError):
            valor = 0.0
        valor_exibicao = valor * 100 if operando_tipo == "percentual_fixo" else valor
        texto_valor = f"{valor_exibicao:.6f}".rstrip("0").rstrip(".").replace(".", ",")
        operacoes.append({
            "operador": limpar_texto(operacao.get("operador")),
            "operando_tipo": operando_tipo,
            "indicador": limpar_texto(operacao.get("indicador")),
            "valor": valor,
            "valor_exibicao": texto_valor or "0",
        })
    return {"primeiro": primeiro, "operacoes": operacoes, "limitar_zero": bool(configuracao.get("limitar_zero"))}


def carregar_dashboard_campos(apenas_ativos: bool = False) -> list[dict[str, Any]]:
    where = "WHERE ativo = 1" if apenas_ativos else ""
    rows = get_db().execute(
        f"SELECT * FROM dashboard_campos {where} ORDER BY ordem, id"
    ).fetchall()
    campos: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        try:
            configuracao = json.loads(item["configuracao_json"] or "{}")
        except (TypeError, ValueError, json.JSONDecodeError):
            configuracao = {}
        item["configuracao"] = configuracao if isinstance(configuracao, dict) else {}
        if item["modalidade"] == "formula":
            item["formula"] = normalizar_formula_dashboard(item["configuracao"], item["tipo"])
        campos.append(item)
    return campos


def indicadores_dashboard_disponiveis(excluir_id: int | None = None) -> list[dict[str, str]]:
    indicadores = [
        {"chave": chave, "nome": info["nome"], "tipo": info["tipo"]}
        for chave, info in DASHBOARD_INDICADORES_BASE.items()
    ]
    for campo in carregar_dashboard_campos():
        if excluir_id and campo["id"] == excluir_id:
            continue
        indicadores.append({
            "chave": f"personalizado:{campo['id']}",
            "nome": campo["nome"],
            "tipo": campo["tipo"],
        })
    return indicadores


def dashboard_filtros_valores() -> dict[str, list[str]]:
    valores: dict[str, list[str]] = {"": []}
    for campo in DASHBOARD_FILTRO_CAMPOS:
        if not campo:
            continue
        rows = get_db().execute(
            f"""SELECT DISTINCT TRIM(COALESCE({campo}, '')) AS valor
                FROM propostas
                WHERE TRIM(COALESCE({campo}, '')) <> ''
                ORDER BY valor COLLATE NOCASE"""
        ).fetchall()
        valores[campo] = [row["valor"] for row in rows]
    return valores


def dados_campo_dashboard_form(campo_id: int | None = None) -> tuple[dict[str, Any], list[str]]:
    erros: list[str] = []
    nome = limpar_texto(request.form.get("nome"))
    tipo = limpar_texto(request.form.get("tipo"))
    modalidade = limpar_texto(request.form.get("modalidade"))
    cor = limpar_texto(request.form.get("cor"))
    try:
        ordem = max(1, int(request.form.get("ordem") or 1))
    except (TypeError, ValueError):
        ordem = 1
    ativo = 1 if request.form.get("ativo") == "1" else 0

    if not nome:
        erros.append("Informe o nome do campo do Dashboard.")
    if tipo not in DASHBOARD_CAMPO_TIPOS:
        erros.append("Escolha um tipo de valor válido.")
    if modalidade not in DASHBOARD_CAMPO_MODALIDADES:
        erros.append("Escolha uma origem válida para o indicador.")
    if cor not in DASHBOARD_CAMPO_CORES:
        cor = "azul"

    configuracao: dict[str, Any] = {}
    if modalidade == "agregado":
        operacao = limpar_texto(request.form.get("agregacao"))
        campo_valor = limpar_texto(request.form.get("campo_valor"))
        base_data = limpar_texto(request.form.get("base_data"))
        filtro_campo = limpar_texto(request.form.get("filtro_campo"))
        filtro_valor = limpar_texto(request.form.get("filtro_valor"))
        if operacao not in DASHBOARD_AGREGACOES:
            erros.append("Escolha como os dados das propostas serão calculados.")
        if operacao != "contagem" and campo_valor not in DASHBOARD_AGREGACAO_CAMPOS:
            erros.append("Escolha qual valor das propostas será usado no cálculo.")
        if base_data not in ("criacao", "encerramento"):
            erros.append("Escolha a referência mensal do indicador.")
        if filtro_campo not in DASHBOARD_FILTRO_CAMPOS:
            erros.append("Escolha um filtro de propostas válido.")
        if filtro_campo and not filtro_valor:
            erros.append("Escolha o valor do filtro de propostas.")
        configuracao = {
            "agregacao": operacao,
            "campo_valor": campo_valor if operacao != "contagem" else "id",
            "base_data": base_data,
            "filtro_campo": filtro_campo,
            "filtro_valor": filtro_valor if filtro_campo else "",
        }
    elif modalidade == "formula":
        indicador_a = limpar_texto(request.form.get("indicador_a"))
        chaves_validas = {item["chave"] for item in indicadores_dashboard_disponiveis(campo_id)}
        if indicador_a not in chaves_validas:
            erros.append("Escolha um indicador inicial válido para a fórmula.")
        operadores = request.form.getlist("operador")[:10]
        tipos_operandos = request.form.getlist("operando_tipo")[:10]
        indicadores_b = request.form.getlist("indicador_b")[:10]
        valores_fixos = request.form.getlist("valor_fixo")[:10]
        if not operadores:
            erros.append("Adicione pelo menos uma operação à fórmula.")
        operacoes = []
        for indice, operador in enumerate(operadores):
            operador = limpar_texto(operador)
            operando_tipo = limpar_texto(tipos_operandos[indice] if indice < len(tipos_operandos) else "indicador")
            indicador_b = limpar_texto(indicadores_b[indice] if indice < len(indicadores_b) else "")
            valor_texto = limpar_texto(valores_fixos[indice] if indice < len(valores_fixos) else "")
            if operador not in DASHBOARD_FORMULA_OPERACOES:
                erros.append(f"A operação {indice + 1} não é válida.")
            if operando_tipo not in DASHBOARD_FORMULA_OPERANDO_TIPOS:
                erros.append(f"Escolha a origem do valor na operação {indice + 1}.")
                operando_tipo = "indicador"
            valor = 0.0
            if operando_tipo == "indicador":
                if indicador_b not in chaves_validas:
                    erros.append(f"Escolha um indicador válido na operação {indice + 1}.")
            else:
                if not valor_texto:
                    erros.append(f"Informe o valor fixo da operação {indice + 1}.")
                valor = parse_percentual(valor_texto) if operando_tipo == "percentual_fixo" else parse_moeda(valor_texto)
                if operando_tipo == "percentual_fixo":
                    valor /= 100
            operacoes.append({
                "operador": operador,
                "operando_tipo": operando_tipo,
                "indicador": indicador_b if operando_tipo == "indicador" else "",
                "valor": valor,
            })
        configuracao = {
            "primeiro": indicador_a,
            "operacoes": operacoes,
            "limitar_zero": request.form.get("limitar_zero") == "1",
        }

    return {
        "nome": nome,
        "tipo": tipo,
        "modalidade": modalidade,
        "configuracao_json": json.dumps(configuracao, ensure_ascii=False),
        "configuracao": configuracao,
        "cor": cor,
        "ordem": ordem,
        "ativo": ativo,
    }, erros


def referencias_formula_dashboard(configuracao: dict[str, Any], tipo_resultado: str = "numero") -> list[str]:
    formula = normalizar_formula_dashboard(configuracao, tipo_resultado)
    referencias = [formula["primeiro"]] if formula["primeiro"] else []
    referencias.extend(
        operacao["indicador"]
        for operacao in formula["operacoes"]
        if operacao["operando_tipo"] == "indicador" and operacao["indicador"]
    )
    return referencias


def formula_dashboard_tem_ciclo(campo_id: int, configuracao_nova: dict[str, Any]) -> bool:
    if not campo_id:
        return False
    grafo: dict[int, list[int]] = {}
    for campo in carregar_dashboard_campos():
        config = configuracao_nova if campo["id"] == campo_id else campo["configuracao"]
        if (campo["id"] == campo_id or campo["modalidade"] == "formula") and config:
            dependencias = []
            for chave in referencias_formula_dashboard(config, campo["tipo"]):
                if limpar_texto(chave).startswith("personalizado:"):
                    try:
                        dependencias.append(int(str(chave).split(":", 1)[1]))
                    except (TypeError, ValueError):
                        pass
            grafo[campo["id"]] = dependencias

    visitando: set[int] = set()
    visitados: set[int] = set()

    def visitar(no: int) -> bool:
        if no in visitando:
            return True
        if no in visitados:
            return False
        visitando.add(no)
        if any(visitar(dependencia) for dependencia in grafo.get(no, [])):
            return True
        visitando.remove(no)
        visitados.add(no)
        return False

    return any(visitar(no) for no in grafo)


@app.route("/configuracoes/dashboard", methods=["GET", "POST"])
def configurar_dashboard():
    db = get_db()
    editar_id = int(request.values.get("editar") or 0) if str(request.values.get("editar") or "").isdigit() else 0
    campo_edicao = next((campo for campo in carregar_dashboard_campos() if campo["id"] == editar_id), None)

    if request.method == "POST":
        acao = limpar_texto(request.form.get("acao"))
        campo_id = int(request.form.get("campo_id") or 0) if str(request.form.get("campo_id") or "").isdigit() else 0
        campo_existente = db.execute("SELECT * FROM dashboard_campos WHERE id = ?", (campo_id,)).fetchone() if campo_id else None

        if acao in ("criar", "salvar"):
            if acao == "salvar" and not campo_existente:
                flash("O campo do Dashboard não foi encontrado.", "erro")
                return redirect(url_for("configurar_dashboard"))
            dados, erros = dados_campo_dashboard_form(campo_id or None)
            if dados["modalidade"] == "formula" and formula_dashboard_tem_ciclo(campo_id, dados["configuracao"]):
                erros.append("Essa fórmula criaria uma dependência circular entre os campos.")
            if erros:
                for erro in erros:
                    flash(erro, "erro")
                destino = url_for("configurar_dashboard", editar=campo_id) if campo_id else url_for("configurar_dashboard")
                return redirect(destino)
            agora = agora_iso()
            if acao == "criar":
                db.execute(
                    """INSERT INTO dashboard_campos
                       (nome, tipo, modalidade, configuracao_json, cor, ordem, ativo, criado_em, atualizado_em)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (dados["nome"], dados["tipo"], dados["modalidade"], dados["configuracao_json"],
                     dados["cor"], dados["ordem"], dados["ativo"], agora, agora),
                )
                mensagem = "Campo adicionado ao Dashboard."
            else:
                db.execute(
                    """UPDATE dashboard_campos
                       SET nome = ?, tipo = ?, modalidade = ?, configuracao_json = ?, cor = ?,
                           ordem = ?, ativo = ?, atualizado_em = ?
                       WHERE id = ?""",
                    (dados["nome"], dados["tipo"], dados["modalidade"], dados["configuracao_json"],
                     dados["cor"], dados["ordem"], dados["ativo"], agora, campo_id),
                )
                mensagem = "Campo do Dashboard atualizado."
            db.commit()
            flash(mensagem, "ok")
            return redirect(url_for("configurar_dashboard"))

        if acao == "excluir":
            if not campo_existente:
                flash("O campo do Dashboard não foi encontrado.", "erro")
                return redirect(url_for("configurar_dashboard"))
            chave = f"personalizado:{campo_id}"
            dependentes = []
            for campo in carregar_dashboard_campos():
                config = campo["configuracao"]
                if campo["modalidade"] == "formula" and chave in referencias_formula_dashboard(config, campo["tipo"]):
                    dependentes.append(campo["nome"])
            if dependentes:
                flash(f"Este campo é usado por: {', '.join(dependentes)}. Edite essas fórmulas antes de excluir.", "erro")
                return redirect(url_for("configurar_dashboard"))
            db.execute("DELETE FROM dashboard_campos WHERE id = ?", (campo_id,))
            db.commit()
            flash("Campo removido do Dashboard.", "ok")
            return redirect(url_for("configurar_dashboard"))

    if campo_edicao:
        formulario = dict(campo_edicao)
    else:
        proxima_ordem = get_db().execute("SELECT COALESCE(MAX(ordem), 0) + 1 AS ordem FROM dashboard_campos").fetchone()["ordem"]
        formulario = {
            "id": 0, "nome": "", "tipo": "moeda", "modalidade": "manual",
            "configuracao": {},
            "formula": {
                "primeiro": "total",
                "operacoes": [{
                    "operador": "somar", "operando_tipo": "indicador",
                    "indicador": "pagas", "valor": 0, "valor_exibicao": "0",
                }],
                "limitar_zero": False,
            },
            "cor": "azul", "ordem": proxima_ordem, "ativo": 1,
        }
    formulario.setdefault("formula", {
        "primeiro": "total",
        "operacoes": [{
            "operador": "somar", "operando_tipo": "indicador",
            "indicador": "pagas", "valor": 0, "valor_exibicao": "0",
        }],
        "limitar_zero": False,
    })
    return render_template(
        "dashboard_config.html",
        campos=carregar_dashboard_campos(),
        campo=formulario,
        indicadores=indicadores_dashboard_disponiveis(formulario.get("id") or None),
        filtros_valores=dashboard_filtros_valores(),
        tipos=DASHBOARD_CAMPO_TIPOS,
        modalidades=DASHBOARD_CAMPO_MODALIDADES,
        cores=DASHBOARD_CAMPO_CORES,
        agregacoes=DASHBOARD_AGREGACOES,
        campos_agregacao=DASHBOARD_AGREGACAO_CAMPOS,
        campos_filtro=DASHBOARD_FILTRO_CAMPOS,
        formula_operacoes=DASHBOARD_FORMULA_OPERACOES,
        formula_operando_tipos=DASHBOARD_FORMULA_OPERANDO_TIPOS,
        titulo="Campos do Dashboard",
        subtitulo="Crie indicadores manuais, cálculos sobre propostas e fórmulas personalizadas.",
    )



@app.route("/configuracoes/status", methods=["GET", "POST"])
def configurar_status():
    db = get_db()
    if request.method == "POST":
        acao = limpar_texto(request.form.get("acao"))

        if acao == "adicionar":
            nome = limpar_texto(request.form.get("nome"))
            grupo = "geral"
            if not nome:
                flash("Informe o nome da etapa.", "erro")
                return redirect(url_for("configurar_status"))
            existe = db.execute("SELECT id FROM status_etapas WHERE nome = ?", (nome,)).fetchone()
            if existe:
                flash("Já existe uma etapa com esse nome.", "erro")
                return redirect(url_for("configurar_status"))
            proxima_ordem = db.execute("SELECT COALESCE(MAX(ordem), 0) + 1 AS ordem FROM status_etapas").fetchone()["ordem"]
            db.execute(
                "INSERT INTO status_etapas (nome, grupo, ordem, ativo) VALUES (?, ?, ?, 1)",
                (nome, grupo, proxima_ordem),
            )
            db.commit()
            flash("Etapa adicionada.", "ok")
            return redirect(url_for("configurar_status"))

        if acao == "salvar":
            # Compatibilidade com versões antigas da tela de etapas.
            etapa_id = int(request.form.get("etapa_id") or 0)
            etapa = db.execute("SELECT * FROM status_etapas WHERE id = ?", (etapa_id,)).fetchone()
            if not etapa:
                flash("Etapa não encontrada.", "erro")
                return redirect(url_for("configurar_status"))
            nome = limpar_texto(request.form.get("nome"))
            grupo = "geral"
            ordem = int(request.form.get("ordem") or etapa["ordem"] or 0)
            ativo = 1 if request.form.get("ativo") == "1" else 0
            if not nome:
                flash("O nome da etapa não pode ficar vazio.", "erro")
                return redirect(url_for("configurar_status"))
            duplicada = db.execute(
                "SELECT id FROM status_etapas WHERE nome = ? AND id <> ?", (nome, etapa_id)
            ).fetchone()
            if duplicada:
                flash("Já existe outra etapa com esse nome.", "erro")
                return redirect(url_for("configurar_status"))
            db.execute(
                "UPDATE status_etapas SET nome = ?, grupo = ?, ordem = ?, ativo = ? WHERE id = ?",
                (nome, grupo, ordem, ativo, etapa_id),
            )
            if nome != etapa["nome"]:
                db.execute("UPDATE propostas SET status = ? WHERE status = ?", (nome, etapa["nome"]))
                db.execute("UPDATE historico SET status_anterior = ? WHERE status_anterior = ?", (nome, etapa["nome"]))
                db.execute("UPDATE historico SET status_novo = ? WHERE status_novo = ?", (nome, etapa["nome"]))
            db.commit()
            flash("Etapa atualizada.", "ok")
            return redirect(url_for("configurar_status"))

        if acao == "salvar_todas":
            ids = request.form.getlist("etapa_id")
            nomes = request.form.getlist("nome")
            ordens = request.form.getlist("ordem")
            ativos = request.form.getlist("ativo")

            if not ids:
                flash("Nenhuma etapa enviada para salvar.", "erro")
                return redirect(url_for("configurar_status"))

            total = min(len(ids), len(nomes), len(ordens), len(ativos))
            dados = []
            nomes_normalizados = []

            for idx in range(total):
                try:
                    etapa_id = int(ids[idx])
                except (TypeError, ValueError):
                    continue
                nome = limpar_texto(nomes[idx])
                if not nome:
                    flash("Nenhuma etapa pode ficar sem nome.", "erro")
                    return redirect(url_for("configurar_status"))
                nome_chave = nome.casefold()
                if nome_chave in nomes_normalizados:
                    flash(f"Existe etapa duplicada com o nome: {nome}.", "erro")
                    return redirect(url_for("configurar_status"))
                nomes_normalizados.append(nome_chave)
                try:
                    ordem = int(ordens[idx] or idx + 1)
                except (TypeError, ValueError):
                    ordem = idx + 1
                ativo = 1 if str(ativos[idx]) == "1" else 0
                dados.append((etapa_id, nome, ordem, ativo))

            etapas_atuais = {
                row["id"]: row
                for row in db.execute("SELECT * FROM status_etapas").fetchall()
            }

            for etapa_id, nome, ordem, ativo in dados:
                etapa = etapas_atuais.get(etapa_id)
                if not etapa:
                    continue
                duplicada = db.execute(
                    "SELECT id FROM status_etapas WHERE nome = ? AND id <> ?",
                    (nome, etapa_id),
                ).fetchone()
                if duplicada:
                    flash(f"Já existe outra etapa com o nome: {nome}.", "erro")
                    return redirect(url_for("configurar_status"))

            for etapa_id, nome, ordem, ativo in dados:
                etapa = etapas_atuais.get(etapa_id)
                if not etapa:
                    continue
                db.execute(
                    "UPDATE status_etapas SET nome = ?, grupo = ?, ordem = ?, ativo = ? WHERE id = ?",
                    (nome, "geral", ordem, ativo, etapa_id),
                )
                if nome != etapa["nome"]:
                    db.execute("UPDATE propostas SET status = ? WHERE status = ?", (nome, etapa["nome"]))
                    db.execute("UPDATE historico SET status_anterior = ? WHERE status_anterior = ?", (nome, etapa["nome"]))
                    db.execute("UPDATE historico SET status_novo = ? WHERE status_novo = ?", (nome, etapa["nome"]))

            db.commit()
            flash("Etapas atualizadas com sucesso.", "ok")
            return redirect(url_for("configurar_status"))

        if acao == "excluir":
            etapa_id = int(request.form.get("etapa_id") or 0)
            destino = limpar_texto(request.form.get("destino_status"))
            etapa = db.execute("SELECT * FROM status_etapas WHERE id = ?", (etapa_id,)).fetchone()
            if not etapa:
                flash("Etapa não encontrada.", "erro")
                return redirect(url_for("configurar_status"))
            if not destino or destino == etapa["nome"] or not status_valido(destino):
                flash("Escolha uma etapa de destino válida antes de excluir.", "erro")
                return redirect(url_for("configurar_status"))
            db.execute("UPDATE propostas SET status = ? WHERE status = ?", (destino, etapa["nome"]))
            db.execute("DELETE FROM status_etapas WHERE id = ?", (etapa_id,))
            db.commit()
            flash("Etapa excluída e propostas movidas para a etapa escolhida.", "ok")
            return redirect(url_for("configurar_status"))

    etapas = carregar_status_etapas()
    return render_template(
        "status_config.html",
        etapas=etapas,
        titulo="Configurar etapas",
        subtitulo="Edite os nomes, a ordem e quais etapas ficam ativas no funil.",
    )


def consulta_dashboard(mes: str) -> dict[str, Any]:
    db = get_db()
    propostas = db.execute(
        "SELECT * FROM propostas WHERE substr(data_criacao, 1, 7) = ? ORDER BY data_criacao DESC", (mes,)
    ).fetchall()
    propostas_encerradas_mes = db.execute(
        """
        SELECT *
        FROM propostas
        WHERE status IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')
          AND substr(COALESCE(data_encerramento, data_atualizacao, data_criacao), 1, 7) = ?
        """,
        (mes,),
    ).fetchall()
    placeholders_comissao_prevista = ", ".join("?" for _ in STATUS_COMISSAO_PREVISTA)
    propostas_comissao_prevista = db.execute(
        f"""
        SELECT status, comissao
        FROM propostas
        WHERE status IN ({placeholders_comissao_prevista})
        """,
        STATUS_COMISSAO_PREVISTA,
    ).fetchall()
    total = len(propostas)
    pagas = [p for p in propostas_encerradas_mes if p["status"] == "Pago"]
    perdidas = [p for p in propostas_encerradas_mes if p["status"] in ("Perdido", "Cancelado", "Perdido / Cancelado")]
    troco_previsto = sum(float(p["troco"] or 0) for p in propostas)
    troco_pago = sum(float(p["troco"] or 0) for p in pagas)
    comissao_prevista = sum(float(p["comissao"] or 0) for p in propostas_comissao_prevista)
    comissao_paga = sum(float(p["comissao"] or 0) for p in pagas)
    valor_a_sacar = sum(float(p["comissao"] or 0) for p in pagas if (p["valor_caiu_promotora"] or "NÃO") == "SIM" and (p["valor_sacado"] or "NÃO") != "SIM")
    falta_cair_promotora = sum(float(p["comissao"] or 0) for p in pagas if (p["valor_caiu_promotora"] or "NÃO") != "SIM")
    valor_ja_sacado = sum(float(p["comissao"] or 0) for p in pagas if (p["valor_sacado"] or "NÃO") == "SIM")
    saldo_atual = saldo_em_conta()
    valor_a_receber = valor_a_sacar + falta_cair_promotora + saldo_atual
    valor_previsto = comissao_prevista + valor_a_receber

    def agrupar(campo: str) -> list[dict[str, Any]]:
        rows = db.execute(
            f"""
            SELECT COALESCE(NULLIF({campo}, ''), 'Não informado') AS nome, COUNT(*) AS qtd
            FROM propostas
            WHERE substr(data_criacao, 1, 7) = ?
            GROUP BY COALESCE(NULLIF({campo}, ''), 'Não informado')
            ORDER BY qtd DESC, nome ASC
            """,
            (mes,),
        ).fetchall()
        maior = max([r["qtd"] for r in rows], default=1)
        return [{"nome": r["nome"], "qtd": r["qtd"], "percentual": round((r["qtd"] / maior) * 100, 1)} for r in rows]

    return {
        "mes": mes,
        "total": total,
        "pagas": len(pagas),
        "perdidas": len(perdidas),
        "troco_previsto": troco_previsto,
        "troco_pago": troco_pago,
        "comissao_prevista": comissao_prevista,
        "comissao_prevista_status": STATUS_COMISSAO_PREVISTA,
        "comissao_paga": comissao_paga,
        "valor_a_sacar": valor_a_sacar,
        "falta_cair_promotora": falta_cair_promotora,
        "valor_ja_sacado": valor_ja_sacado,
        "saldo_em_conta": saldo_atual,
        "valor_a_receber": valor_a_receber,
        "valor_previsto": valor_previsto,
        "por_status": agrupar("status"),
        "por_banco": agrupar("banco_digitado"),
        "por_produto": agrupar("produto"),
    }


def saldo_em_conta() -> float:
    registro = get_db().execute(
        "SELECT valor FROM configuracoes WHERE chave = 'saldo_em_conta'"
    ).fetchone()
    return parse_moeda(registro["valor"]) if registro else 0.0


def calcular_agregado_dashboard(configuracao: dict[str, Any], mes: str) -> float:
    agregacao = configuracao.get("agregacao")
    campo_valor = configuracao.get("campo_valor")
    base_data = configuracao.get("base_data")
    filtro_campo = configuracao.get("filtro_campo") or ""
    filtro_valor = limpar_texto(configuracao.get("filtro_valor"))

    if agregacao == "contagem":
        expressao = "COUNT(*)"
    elif agregacao in ("soma", "media") and campo_valor in DASHBOARD_AGREGACAO_CAMPOS:
        funcao = "SUM" if agregacao == "soma" else "AVG"
        expressao = f"{funcao}(COALESCE({campo_valor}, 0))"
    else:
        return 0.0

    params: list[Any] = [mes]
    if base_data == "encerramento":
        where = [
            "status IN ('Pago', 'Perdido / Cancelado', 'Perdido', 'Cancelado')",
            "substr(COALESCE(data_encerramento, data_atualizacao, data_criacao), 1, 7) = ?",
        ]
    else:
        where = ["substr(data_criacao, 1, 7) = ?"]
    if filtro_campo in DASHBOARD_FILTRO_CAMPOS and filtro_campo:
        where.append(f"UPPER(TRIM(COALESCE({filtro_campo}, ''))) = UPPER(?)")
        params.append(filtro_valor)
    registro = get_db().execute(
        f"SELECT {expressao} AS valor FROM propostas WHERE {' AND '.join(where)}",
        params,
    ).fetchone()
    return float(registro["valor"] or 0)


def formatar_valor_dashboard(valor: float, tipo: str) -> str:
    if tipo == "moeda":
        return br_moeda(valor)
    if tipo == "percentual":
        return br_percentual(valor)
    if abs(valor - round(valor)) < 0.00001:
        return f"{int(round(valor)):,}".replace(",", ".")
    texto = f"{valor:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def calcular_campos_dashboard(dados: dict[str, Any], mes: str) -> list[dict[str, Any]]:
    todos_campos = carregar_dashboard_campos()
    campos_por_id = {campo["id"]: campo for campo in todos_campos}
    cache: dict[str, float] = {}
    calculando: set[str] = set()

    def resolver(chave: str) -> float:
        if chave in cache:
            return cache[chave]
        if chave in DASHBOARD_INDICADORES_BASE:
            valor = float(dados.get(chave) or 0)
            cache[chave] = valor
            return valor
        if not chave.startswith("personalizado:"):
            return 0.0
        try:
            campo_id = int(chave.split(":", 1)[1])
        except (TypeError, ValueError):
            return 0.0
        campo = campos_por_id.get(campo_id)
        if not campo or chave in calculando:
            return 0.0
        calculando.add(chave)
        config = campo["configuracao"]
        if campo["modalidade"] == "manual":
            registro = get_db().execute(
                "SELECT valor FROM dashboard_valores_manuais WHERE campo_id = ? AND mes = ?",
                (campo_id, mes),
            ).fetchone()
            valor = float(registro["valor"] or 0) if registro else 0.0
        elif campo["modalidade"] == "agregado":
            valor = calcular_agregado_dashboard(config, mes)
        elif campo["modalidade"] == "formula":
            formula = normalizar_formula_dashboard(config, campo["tipo"])
            valor = resolver(formula["primeiro"])
            for operacao in formula["operacoes"]:
                operando = (
                    resolver(operacao["indicador"])
                    if operacao["operando_tipo"] == "indicador"
                    else float(operacao["valor"] or 0)
                )
                operador = operacao["operador"]
                if operador == "somar":
                    valor += operando
                elif operador == "subtrair":
                    valor -= operando
                elif operador == "multiplicar":
                    valor *= operando
                elif operador == "dividir":
                    valor = (valor / operando) if operando else 0.0
                elif operador == "percentual_de":
                    valor = (valor / operando * 100) if operando else 0.0
                elif operador == "variacao_percentual":
                    valor = ((valor - operando) / operando * 100) if operando else 0.0
                elif operador == "media":
                    valor = (valor + operando) / 2
                elif operador == "maior":
                    valor = max(valor, operando)
                elif operador == "menor":
                    valor = min(valor, operando)
                elif operador == "diferenca_absoluta":
                    valor = abs(valor - operando)
            if formula["limitar_zero"]:
                valor = max(0.0, valor)
        else:
            valor = 0.0
        calculando.discard(chave)
        cache[chave] = valor
        return valor

    resultado: list[dict[str, Any]] = []
    for campo in todos_campos:
        if not campo["ativo"]:
            continue
        item = dict(campo)
        item["valor"] = resolver(f"personalizado:{campo['id']}")
        item["valor_formatado"] = formatar_valor_dashboard(item["valor"], campo["tipo"])
        resultado.append(item)
    return resultado


@app.route("/dashboard")
def dashboard():
    mes = limpar_texto(request.args.get("mes")) or mes_atual()
    dados = consulta_dashboard(mes)
    return render_template(
        "dashboard.html",
        dados=dados,
        saldo_em_conta=dados["saldo_em_conta"],
        campos_dashboard=calcular_campos_dashboard(dados, mes),
    )


@app.route("/dashboard/campos/<int:campo_id>/valor", methods=["POST"])
def atualizar_valor_dashboard(campo_id: int):
    campo = get_db().execute(
        "SELECT * FROM dashboard_campos WHERE id = ? AND modalidade = 'manual'",
        (campo_id,),
    ).fetchone()
    mes = limpar_texto(request.form.get("mes")) or mes_atual()
    if not campo:
        flash("O campo manual do Dashboard não foi encontrado.", "erro")
        return redirect(url_for("dashboard", mes=mes))
    valor_texto = limpar_texto(request.form.get("valor"))
    if not valor_texto:
        flash("Informe o valor do campo.", "erro")
        return redirect(url_for("dashboard", mes=mes))
    valor = parse_percentual(valor_texto) if campo["tipo"] == "percentual" else parse_moeda(valor_texto)
    get_db().execute(
        """INSERT INTO dashboard_valores_manuais (campo_id, mes, valor, atualizado_em)
           VALUES (?, ?, ?, ?)
           ON CONFLICT(campo_id, mes) DO UPDATE
           SET valor = excluded.valor, atualizado_em = excluded.atualizado_em""",
        (campo_id, mes, valor, agora_iso()),
    )
    get_db().commit()
    flash("Valor mensal atualizado.", "ok")
    return redirect(url_for("dashboard", mes=mes))


@app.route("/dashboard/saldo", methods=["POST"])
def atualizar_saldo_dashboard():
    valor_informado = limpar_texto(request.form.get("saldo_em_conta"))
    if not valor_informado:
        flash("Informe o saldo em conta.", "erro")
    else:
        get_db().execute(
            """INSERT INTO configuracoes (chave, valor, atualizado_em) VALUES ('saldo_em_conta', ?, ?)
               ON CONFLICT(chave) DO UPDATE SET valor = excluded.valor, atualizado_em = excluded.atualizado_em""",
            (str(parse_moeda(valor_informado)), agora_iso()),
        )
        get_db().commit()
        flash("Saldo em conta atualizado.", "ok")
    mes = limpar_texto(request.form.get("mes")) or mes_atual()
    return redirect(url_for("dashboard", mes=mes))


@app.route("/exportar/csv")
def exportar_csv():
    propostas = get_db().execute("SELECT * FROM propostas ORDER BY data_criacao DESC").fetchall()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    colunas = [desc[0] for desc in get_db().execute("SELECT * FROM propostas LIMIT 1").description if desc[0] != "banco_destino"]
    writer.writerow(colunas)
    for p in propostas:
        writer.writerow([p[c] for c in colunas])
    memoria = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(memoria, mimetype="text/csv", as_attachment=True, download_name="propostas_crm.csv")


@app.route("/exportar/xlsx")
def exportar_xlsx():
    propostas = get_db().execute("SELECT * FROM propostas ORDER BY data_criacao DESC").fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Propostas"
    colunas = [desc[0] for desc in get_db().execute("SELECT * FROM propostas LIMIT 1").description if desc[0] != "banco_destino"]
    ws.append(colunas)
    for p in propostas:
        ws.append([p[c] for c in colunas])
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return send_file(memoria, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="propostas_crm.xlsx")


def normalizar_cabecalho(cabecalho: Any) -> str:
    texto = limpar_texto(cabecalho).lower()
    texto = texto.replace("ç", "c").replace("ã", "a").replace("õ", "o").replace("á", "a").replace("à", "a").replace("é", "e").replace("ê", "e").replace("í", "i").replace("ó", "o").replace("ô", "o").replace("ú", "u")
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    aliases = {
        "nome_do_cliente": "nome",
        "cliente": "nome",
        "nb": "nb_matricula",
        "matricula": "nb_matricula",
        "nb_ou_matricula": "nb_matricula",
        "tipo": "tipo_cliente",
        "banco_origem": "banco_atual",
        "banco_atual": "banco_atual",
        "banco": "banco_digitado",
        "banco_digitado": "banco_digitado",
        "banco_de_digitacao": "banco_digitado",
        "digitado": "banco_digitado",
        "destino": "banco_digitado",
        "banco_destino": "banco_digitado",
        "promotora": "promotora",
        "promotora_responsavel": "promotora",
        "bloqueado": "beneficio_bloqueado",
        "beneficio_bloqueado": "beneficio_bloqueado",
        "beneficio": "beneficio_bloqueado",
        "valor": "troco",
        "parcela": "parcela_atual",
        "valor_da_parcela_atual": "parcela_atual",
        "valor_da_nova_parcela": "nova_parcela",
        "nova_parcela": "nova_parcela",
        "valor_de_troco": "troco",
        "comissao_percentual": "comissao_percentual",
        "percentual_comissao": "comissao_percentual",
        "percentual_da_comissao": "comissao_percentual",
        "porcentagem_comissao": "comissao_percentual",
        "porcentagem_da_comissao": "comissao_percentual",
        "comissao_%": "comissao_percentual",
        "comissao": "comissao",
        "comissão": "comissao",
        "valor_da_comissao": "comissao",
        "valor_da_comissão": "comissao",
        "pontos": "comissao",
        "saldo_pmt": "valor_caiu_promotora",
        "saldo_pmt_": "valor_caiu_promotora",
        "valor_caiu_na_promotora": "valor_caiu_promotora",
        "caiu_na_promotora": "valor_caiu_promotora",
        "valor_caiu_promotora": "valor_caiu_promotora",
        "acerto": "valor_sacado",
        "sacado": "valor_sacado",
        "valor_ja_foi_sacado": "valor_sacado",
        "valor_sacado": "valor_sacado",
        "margem": "margem_apos",
        "whatsapp": "telefone",
        "telefone_whatsapp": "telefone",
        "retorno": "data_retorno",
        "data_de_retorno": "data_retorno",
        "previsao_de_saldo": "data_retorno",
        "previsao_saldo": "data_retorno",
        "data_previsao_saldo": "data_retorno",
        "observacao": "observacoes",
        "motivo": "observacoes",
        "motivos": "observacoes",
        "numero_proposta": "numero_proposta",
        "n_proposta": "numero_proposta",
        "n_da_proposta": "numero_proposta",
        "proposta": "numero_proposta",
        "codigo_proposta": "numero_proposta",
        "numero_port": "numero_port_vinculada",
        "proposta_port": "numero_port_vinculada",
        "proposta_portabilidade": "numero_port_vinculada",
        "numero_proposta_port": "numero_port_vinculada",
        "numero_refin": "numero_refin_vinculada",
        "proposta_refin": "numero_refin_vinculada",
        "proposta_refinanciamento": "numero_refin_vinculada",
        "numero_proposta_refin": "numero_refin_vinculada",
        "dia": "data_criacao",
        "data": "data_criacao",
        "endereco": "endereco",
        "endereço": "endereco",
        "dados_bancarios": "dados_bancarios",
        "dados_bancarios_cliente": "dados_bancarios",
        "banco_cliente": "dados_bancarios",
        "conta_cliente": "dados_bancarios",
        "dados_da_conta": "dados_bancarios",
    }
    return aliases.get(texto, texto)




def localizar_coluna_contatos(headers: list[Any], candidatos: set[str]) -> int | None:
    """Localiza coluna por cabeçalho normalizado, aceitando variações como TELEFONE1 e TELEFONE 1."""
    for idx, header in enumerate(headers):
        normalizado = normalizar_cabecalho(header)
        compactado = re.sub(r"[^a-z0-9]", "", normalizado)
        if normalizado in candidatos or compactado in candidatos:
            return idx
    return None


def normalizar_telefone_contato(valor: Any) -> str:
    """Mantém apenas dígitos e adiciona DDI 55 quando ainda não estiver presente."""
    digitos = re.sub(r"\D", "", limpar_texto(valor))
    if not digitos:
        return ""
    # Evita duplicar o DDI se a base já vier com 55 no início.
    if digitos.startswith("55") and len(digitos) >= 12:
        return digitos
    return f"55{digitos}"


def linhas_contatos_convertidas(arquivo: Any) -> tuple[list[dict[str, str]], int]:
    """Lê CSV/XLSX e retorna linhas no modelo: nome, numero.

    A base original deve ter NOME e TELEFONE1/TELEFONE 1. Telefones vazios são ignorados.
    """
    filename = (arquivo.filename or "").lower()
    linhas: list[dict[str, str]] = []
    total_lidas = 0

    if filename.endswith(".csv"):
        conteudo = arquivo.read().decode("utf-8-sig", errors="replace")
        sample = conteudo[:4096]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(io.StringIO(conteudo), delimiter=delimiter)
        rows = list(reader)
        if not rows:
            return [], 0
        headers = rows[0]
        idx_nome = localizar_coluna_contatos(headers, {"nome"})
        idx_tel = localizar_coluna_contatos(headers, {"telefone1", "telefone_1"})
        if idx_nome is None or idx_tel is None:
            raise ValueError("Não encontrei as colunas NOME e TELEFONE1/TELEFONE 1 no arquivo.")
        for row in rows[1:]:
            total_lidas += 1
            nome = limpar_texto(row[idx_nome] if idx_nome < len(row) else "").strip('"')
            numero = normalizar_telefone_contato(row[idx_tel] if idx_tel < len(row) else "")
            if nome and numero:
                linhas.append({"nome": nome, "numero": numero})
        return linhas, total_lidas

    if filename.endswith(".xlsx"):
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], 0
        header_idx = None
        idx_nome = idx_tel = None
        for i, possivel_header in enumerate(rows[:10]):
            nome_idx = localizar_coluna_contatos(list(possivel_header), {"nome"})
            tel_idx = localizar_coluna_contatos(list(possivel_header), {"telefone1", "telefone_1"})
            if nome_idx is not None and tel_idx is not None:
                header_idx = i
                idx_nome = nome_idx
                idx_tel = tel_idx
                break
        if header_idx is None or idx_nome is None or idx_tel is None:
            raise ValueError("Não encontrei as colunas NOME e TELEFONE1/TELEFONE 1 no arquivo.")
        for row in rows[header_idx + 1:]:
            total_lidas += 1
            nome = limpar_texto(row[idx_nome] if idx_nome < len(row) else "").strip('"')
            numero = normalizar_telefone_contato(row[idx_tel] if idx_tel < len(row) else "")
            if nome and numero:
                linhas.append({"nome": nome, "numero": numero})
        return linhas, total_lidas

    raise ValueError("Formato inválido. Use CSV ou XLSX.")


def gerar_excel_contatos_convertidos(linhas: list[dict[str, str]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatos"
    ws.append(["nome", "numero"])
    for row in linhas:
        ws.append([row["nome"], row["numero"]])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    for cell in ws["B"]:
        cell.number_format = "@"
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return memoria


@app.route("/importar", methods=["POST"])
def importar():
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo CSV ou XLSX.", "erro")
        return redirect(url_for("index"))

    nome = arquivo.filename.lower()
    mes_importacao = limpar_texto(request.form.get("mes_importacao")) or mes_atual()
    aba_preferida = nome_mes_pt(mes_importacao)
    linhas: list[dict[str, Any]] = []
    try:
        if nome.endswith(".csv"):
            conteudo = arquivo.read().decode("utf-8-sig")
            sample = conteudo[:2048]
            delimiter = ";" if sample.count(";") >= sample.count(",") else ","
            reader = csv.DictReader(io.StringIO(conteudo), delimiter=delimiter)
            for row in reader:
                linhas.append({normalizar_cabecalho(k): v for k, v in row.items()})
        elif nome.endswith(".xlsx"):
            wb = load_workbook(arquivo, data_only=True)
            planilha_encontrada = False
            worksheets = wb.worksheets
            if aba_preferida:
                preferidas = [ws for ws in worksheets if remover_acentos(ws.title).upper() == remover_acentos(aba_preferida).upper()]
                outras = [ws for ws in worksheets if ws not in preferidas]
                worksheets = preferidas + outras
            for ws in worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header_idx = None
                for idx, possivel_header in enumerate(rows[:10]):
                    headers_teste = [normalizar_cabecalho(h) for h in possivel_header]
                    if "nome" in headers_teste and ("cpf" in headers_teste or "status" in headers_teste):
                        header_idx = idx
                        break
                if header_idx is None:
                    continue
                headers = [normalizar_cabecalho(h) for h in rows[header_idx]]
                for row in rows[header_idx + 1:]:
                    linhas.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
                planilha_encontrada = True
                break
            if not planilha_encontrada:
                flash("Não encontrei uma aba com cabeçalhos como NOME, CPF e STATUS.", "erro")
                return redirect(url_for("index"))
        else:
            flash("Formato inválido. Use CSV ou XLSX.", "erro")
            return redirect(url_for("index"))
    except Exception as exc:
        flash(f"Erro ao ler arquivo: {exc}", "erro")
        return redirect(url_for("index"))

    linhas_filtradas = [row for row in linhas if linha_no_mes(row, mes_importacao)]

    importadas = 0
    ignoradas_mes = len(linhas) - len(linhas_filtradas)
    db = get_db()
    for row in linhas_filtradas:
        nome_cliente = limpar_texto(row.get("nome"))
        if not nome_cliente:
            continue
        status_original = limpar_texto(row.get("status"))
        status = normalizar_status_importacao(status_original)
        if not status_valido(status):
            status = status_padrao()
        produto_importado = normalizar_produto_importacao(row.get("produto"))
        banco_digitado_importado = texto_planilha(row.get("banco_digitado"))
        banco_destino_importado = ""
        agora = agora_iso()
        cursor = db.execute(
            """
            INSERT INTO propostas (
                cliente_id, nome, cpf, nb_matricula, numero_proposta, numero_port_vinculada, numero_refin_vinculada, tipo_cliente, banco_atual, banco_destino, banco_digitado, produto,
                promotora, beneficio_bloqueado, valor_caiu_promotora, valor_sacado, data_verificacao, parcela_atual, nova_parcela, troco, comissao_percentual, comissao, margem_apos, status, responsavel,
                telefone, endereco, dados_bancarios, data_criacao, data_atualizacao, proxima_acao, data_retorno, observacoes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                salvar_cliente_dos_dados({
                    "nome": nome_cliente,
                    "cpf": formatar_cpf(texto_planilha(row.get("cpf"))),
                    "nb_matricula": texto_planilha(row.get("nb_matricula")),
                    "telefone": texto_planilha(row.get("telefone")),
                    "tipo_cliente": texto_planilha(row.get("tipo_cliente")),
                    "endereco": texto_planilha(row.get("endereco")),
                    "dados_bancarios": texto_planilha(row.get("dados_bancarios")),
                }),
                nome_cliente,
                formatar_cpf(texto_planilha(row.get("cpf"))),
                texto_planilha(row.get("nb_matricula")),
                texto_planilha(row.get("numero_proposta")),
                texto_planilha(row.get("numero_port_vinculada")),
                texto_planilha(row.get("numero_refin_vinculada")),
                texto_planilha(row.get("tipo_cliente")),
                texto_planilha(row.get("banco_atual")),
                banco_destino_importado,
                banco_digitado_importado,
                produto_importado,
                texto_planilha(row.get("promotora")),
                normalizar_bloqueado(row.get("beneficio_bloqueado")),
                normalizar_sim_nao(row.get("valor_caiu_promotora")),
                normalizar_sim_nao(row.get("valor_sacado")),
                hoje_iso(),
                parse_moeda(row.get("parcela_atual")),
                parse_moeda(row.get("nova_parcela")),
                parse_moeda(row.get("troco")),
                parse_percentual(row.get("comissao_percentual")),
                parse_moeda(row.get("comissao")),
                limpar_texto(row.get("margem_apos")),
                status,
                texto_planilha(row.get("responsavel")),
                texto_planilha(row.get("telefone")),
                texto_planilha(row.get("endereco")),
                texto_planilha(row.get("dados_bancarios")),
                parse_data_iso(row.get("data_criacao")) or agora,
                agora,
                texto_planilha(row.get("proxima_acao")),
                parse_data_iso(row.get("data_retorno")) or extrair_data_do_status(limpar_texto(row.get("status")), row.get("data_criacao")),
                texto_planilha(row.get("observacoes")),
            ),
        )
        db.commit()
        registrar_historico(cursor.lastrowid, None, status, "Proposta importada")
        observacao_importada = texto_planilha(row.get("observacoes"))
        if observacao_importada:
            registrar_anotacao(cursor.lastrowid, observacao_importada, agora)
        importadas += 1

    complemento = f" Ignoradas por mês diferente: {ignoradas_mes}." if ignoradas_mes else ""
    flash(f"Importação concluída: {importadas} proposta(s) importada(s) de {mes_importacao}.{complemento}", "ok")
    return redirect(url_for("index"))


if __name__ == "__main__":
    os.environ.setdefault("FLASK_ENV", "development")
    with app.app_context():
        init_db()
        garantir_modelos()
        criar_backup_automatico()
    app.run(host="0.0.0.0", port=5000, debug=False)
